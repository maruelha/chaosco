"""Core South Sustainphase Monitoring (build plan step 2, 2026-08-27):
importer — tab-name pattern → (day, stream), parent/detail split via
Task ID + outline level, replace-per-tab write.

The fixture builds the CURRENT (2026-08-31) layout: header row 5 plus
the Comments/Observations column M. `header_row`/`comments` parametrise
it so the pre-2026-08-31 layout (header row 6, no column M) is covered
too -- the header row is located, not hardcoded."""
from pathlib import Path

import openpyxl
import pytest

from app import database
from app.db import sustain as db_sustain
from app.sustain_importer import (ParseError, parse_sustain_workbook,
                                  run_sustain_import)

HEADERS = ["Task ID", "L4 Taxonomy", "Process / Task", "Cadence",
           "Due Today", "Country", "Provider / Partner / Financial Account",
           "France Result", "Italy Result", "Portugal Result", "Spain Result",
           "Task Overall (DO NOT EDIT)"]
COMMENTS_HEADER = "Comments/Observations"


def _sheet(wb, title, header_row=5, comments=True):
    ws = wb.create_sheet(title)
    ws["A1"] = "Retail Daily Operations Checklist Template"
    ws.cell(header_row - 2, 8, "DUE TASKS")
    ws.cell(header_row - 1, 8, 2)
    headers = HEADERS + ([COMMENTS_HEADER] if comments else [])
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    return ws


def _row(ws, r, values, outline_level=0):
    for col, v in enumerate(values, start=1):
        ws.cell(r, col, v)
    if outline_level:
        ws.row_dimensions[r].outline_level = outline_level
        ws.row_dimensions[r].hidden = True


def _workbook(tmp_path, header_row=5, comments=True) -> Path:
    """The data rows sit at header_row+1 .. header_row+5, so the same
    fixture describes both workbook layouts."""
    r0 = header_row + 1
    com = (lambda v: [v]) if comments else (lambda v: [])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    retail = _sheet(wb, "Retail_2026-09-01", header_row, comments)
    _row(retail, r0, ["1", "Settlement", "Monitor files", "Daily", "Yes",
                      None, "Adyen (POS)", "OK ", "N/A", "N/A", "N/A", "OK"]
         + com("watch the cut-off "))
    _row(retail, r0 + 1, [None, None, "↳ Detail check", "Daily", "Yes",
                          "France", "Adyen for cards", "diff 12,50", "N/A",
                          "N/A", "N/A", None] + com(None), outline_level=1)
    _row(retail, r0 + 2, [None, None, "↳ Detail check", "Monthly", "No",
                          "France", "Voucher CadhoC", "Not due", "N/A",
                          "N/A", "N/A", None] + com("monthly, next Friday"),
         outline_level=1)
    _row(retail, r0 + 3, [2, "Clearing", "Monitor items", "Daily", "Yes",
                          None, "Sundry", None, None, None, None, "Pending"]
         + com(None))
    _row(retail, r0 + 5, [None] * 12 + com(None))  # trailing blank: ignored
    ecom = _sheet(wb, "eCom_2026-09-01", header_row, comments)
    _row(ecom, r0, ["1", "Settlement", "Monitor files", "Daily", "Yes", None,
                    "AAEFR-ADIDAS_FR", "OK", "N/A", "N/A", "N/A", "OK"]
         + com(None))
    wb.create_sheet("Instructions")["A1"] = "not a day tab"
    path = tmp_path / "1_0109_0409-O2C DTC_GBS Operations_checklist.xlsx"
    wb.save(path)
    return path


@pytest.mark.parametrize("header_row, comments", [
    (5, True),    # current layout (2026-08-31): header row 5 + column M
    (6, False),   # previous layout: header row 6, no comments column
])
def test_parse_splits_tabs_parents_and_details(tmp_path, header_row, comments):
    tabs = parse_sustain_workbook(_workbook(tmp_path, header_row, comments))
    assert [(t["day"], t["stream"]) for t in tabs] == [
        ("2026-09-01", "Retail"), ("2026-09-01", "eCom")]
    retail = tabs[0]["tasks"]
    assert [t["task_id"] for t in retail] == ["1", "2"]
    assert retail[0]["excel_row"] == header_row + 1
    assert retail[0]["result_fr"] == "OK"      # stripped
    assert retail[0]["provider"] == "Adyen (POS)"
    details = retail[0]["details"]
    assert [(d["country"], d["due_today"]) for d in details] \
        == [("France", "Yes"), ("France", "No")]
    assert details[0]["result_fr"] == "diff 12,50"
    assert retail[1]["details"] == []
    # numeric Task ID (2) is normalised to text
    assert retail[1]["task_id"] == "2"


def test_parse_reads_comments_column(tmp_path):
    """Column M is imported for parents and details, stripped, and is
    simply None in the older layout that has no such column."""
    retail = parse_sustain_workbook(_workbook(tmp_path))[0]["tasks"]
    assert retail[0]["comments"] == "watch the cut-off"     # stripped
    assert [d["comments"] for d in retail[0]["details"]] == [
        None, "monthly, next Friday"]
    assert retail[1]["comments"] is None

    old = parse_sustain_workbook(
        _workbook(tmp_path, header_row=6, comments=False))[0]["tasks"]
    assert old[0]["comments"] is None


def test_parse_rejects_wrong_structure(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retail_2026-09-01"   # matching name, no Task ID header
    ws["A1"] = "something else"
    path = tmp_path / "x DTC_GBS Operations_checklist.xlsx"
    wb.save(path)
    with pytest.raises(ParseError):
        parse_sustain_workbook(path)


def test_parse_no_day_tabs(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Instructions"
    path = tmp_path / "other.xlsx"
    wb.save(path)
    assert parse_sustain_workbook(path) == []


def test_run_sustain_import_writes_per_tab(tmp_path):
    db_path = tmp_path / "s.db"
    database.init_db(db_path).close()
    cfg = {"database_path": str(db_path)}
    result = run_sustain_import(cfg, _workbook(tmp_path))
    assert result["ok"], result["error"]
    assert result["tabs"] == 2
    assert result["tasks"] == 3
    assert result["details"] == 2

    conn = database.get_connection(db_path)
    try:
        assert [(t["day"], t["stream"], t["task_count"])
                for t in db_sustain.list_tabs(conn)] == [
            ("2026-09-01", "Retail", 2), ("2026-09-01", "eCom", 1)]
        # the free-text detail entry rolls up to attention
        counts = db_sustain.summary_counts(conn, "2026-09-01", "Retail")
        assert counts == {"due": 2, "completed": 0, "pending": 1,
                          "attention": 1}
        # comments survive the round trip into storage
        tasks = db_sustain.list_tasks(conn, "2026-09-01", "Retail")
        assert tasks[0]["comments"] == "watch the cut-off"
        assert tasks[0]["details"][1]["comments"] == "monthly, next Friday"
    finally:
        conn.close()


def test_run_sustain_import_error_on_wrong_workbook(tmp_path):
    db_path = tmp_path / "s.db"
    database.init_db(db_path).close()
    wb = openpyxl.Workbook()
    wb.active.title = "Instructions"
    path = tmp_path / "other.xlsx"
    wb.save(path)
    result = run_sustain_import({"database_path": str(db_path)}, path)
    assert not result["ok"]
    assert "day tabs" in result["error"]
