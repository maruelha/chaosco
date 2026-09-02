"""Delegated Testing card: XML upload + buckets on the page, authored
fields (blocked reason / next step), detail page tabs, both reports."""
import io

import pytest

from app import database
from app.db import blockers as db_blockers
from app.db import delegated as db_delegated
from app.db import jira as db_jira
import app.web_blockers as web_blockers
import app.web_core as web_core
import app.web_delegated as web_delegated
import app.web_notes as web_notes
from app.web import app

XML = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="0.92"><channel>
  <item>
    <key id="1">S4ECOM-2001</key>
    <summary>SM2001_Blocked settlement case</summary>
    <type id="17">Story</type>
    <status id="3">Blocked</status>
    <assignee username="JIRAUSER2">Tester, Tom</assignee>
    <link>https://jira.example.com/browse/S4ECOM-2001</link>
    <labels><label>settlement</label><label>fr_scope</label></labels>
    <comments>
      <comment id="1" created="Mon, 24 Aug 2026 10:00:00 +0200">Order Number - TBY_SS_ADE0001111</comment>
      <comment id="2" created="Tue, 25 Aug 2026 10:00:00 +0200">Return Order: 6000084252</comment>
    </comments>
  </item>
  <item>
    <key id="2">S4ECOM-2002</key>
    <summary>SM2002_Marina gatekeeper case</summary>
    <type id="18">User Story</type>
    <status id="3">In Progress</status>
    <assignee username="JIRAUSER1">Haase, Marina [External]</assignee>
    <labels><label>settlement</label></labels>
  </item>
  <item>
    <key id="3">S4ECOM-2003</key>
    <summary>SM2003_Team case</summary>
    <status id="3">Accepted</status>
    <assignee username="JIRAUSER2">Tester, Tom</assignee>
  </item>
  <item>
    <key id="4">S4ECOM-2004</key>
    <summary>SM2004_Odd status case</summary>
    <status id="3">Ready for Verification</status>
    <assignee username="JIRAUSER2">Tester, Tom</assignee>
  </item>
  <item>
    <key id="5">S4DEF-3001</key>
    <summary>SM3001_Unregistered defect in the export</summary>
    <type id="1">Defect</type>
    <status id="3">Open</status>
    <assignee username="JIRAUSER2">Tester, Tom</assignee>
  </item>
  <item>
    <key id="6">S4EPIC-4001</key>
    <summary>SM4001_An epic that slipped into the export</summary>
    <type id="6">Epic</type>
    <status id="3">Open</status>
  </item>
</channel></rss>
"""


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_path = tmp_path / "delegated.db"
    database.init_db(db_path).close()
    db_jira.init_schema(db_path)
    db_delegated.init_schema(db_path)
    db_blockers.init_schema(db_path)
    monkeypatch.setattr(web_delegated, "_get_conn",
                        lambda: database.get_connection(db_path))
    monkeypatch.setitem(web_delegated._cfg, "database_path", str(db_path))
    monkeypatch.setitem(web_delegated._cfg, "jira_gatekeeper_assignee", "Haase")
    monkeypatch.setattr(web_delegated, "_UPLOAD_FOLDER", tmp_path / "uploads")
    # the generic /report-comments/... routes go through web_core._get_conn —
    # without this they would write to the REAL dev DB (bit us 2026-08-27)
    monkeypatch.setattr(web_core, "_db_path", db_path)
    monkeypatch.setattr(web_blockers, "_get_conn",
                        lambda: database.get_connection(db_path))
    # the Ways of Working page adds notes via the generic /n/... routes
    monkeypatch.setattr(web_notes, "_db_path", db_path)
    return app.test_client()


def _upload(client, data=XML, filename="delegated.xml"):
    return client.post("/delegated/upload", data={
        "file": (io.BytesIO(data.encode("utf-8")), filename)})


def test_upload_imports_and_page_shows_buckets(client):
    html = client.get("/delegated/").get_data(as_text=True)
    assert "No delegated tickets yet" in html

    resp = _upload(client)
    assert resp.status_code == 302 and "jira_ok=1" in resp.headers["Location"]

    html = client.get("/delegated/?jira_ok=1&jira_msg=x").get_data(as_text=True)
    assert "S4ECOM-2001" in html                       # blocked ticket present
    assert "\U0001f534 Blocked" in html
    assert "Marina gatekeeper check" in html
    assert "Testing team creating order" in html
    assert "Unexpected status" in html                 # odd status is visible
    assert "Return Order: 6000084252" in html          # LATEST comment's order
    # every section carries a ui-section color modifier — a bare rt-section
    # summary renders WHITE on white (invisible title) [USER 2026-08-26]
    assert 'class="rt-section "' not in html
    assert "ui-section--red" in html                   # Blocked section color
    # count strip at the top [USER 2026-09-02]: every section heading with
    # its number + the total, backlog included
    strip = html.split('class="ui-stackbar-legend"')[1].split("</div>")[0]
    assert "🔴 Blocked: <b>1</b>" in strip
    assert "📦 Backlog: <b>" in strip
    assert "Total: <b>" in strip


def test_only_user_stories_on_board_report_and_numbers(client):
    """[USER 2026-08-27]: "the main page should only have jira user
    stories" — the export deliberately carries the blocker defect issues
    (one upload refreshes everything), so a Defect-type issue must never
    surface as a testing ticket, registered as a blocker or not. 'Story'
    matches by SUBSTRING ("User Story" counts — an exact match emptied
    Marina's real board). Items without a <type> are tolerated too."""
    _upload(client)
    for url in ("/delegated/", "/delegated/report"):
        html = client.get(url).get_data(as_text=True)
        assert "S4DEF-3001" not in html, url
    # numbers: not in the bucket table — it MAY appear below it, in the
    # blocker overview (auto-registered as a blocker on upload)
    numbers_html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "S4DEF-3001" not in numbers_html.split("Blocker overview")[0]
    for url in ("/delegated/", "/delegated/report"):  # numbers shows counts, not keys
        html = client.get(url).get_data(as_text=True)
        assert "S4ECOM-2001" in html, url   # explicit "Story" stays
        assert "S4ECOM-2002" in html, url   # "User Story" stays too (substring)
        assert "S4ECOM-2003" in html, url   # no <type> at all stays too


def test_board_shows_what_the_story_filter_hides(client):
    """The stories-only filter must never empty the board SILENTLY
    (2026-08-27) — the page names the hidden types and counts. Defects
    auto-register as blockers on upload, so only genuinely unhandled
    types (the Epic) appear in the hint."""
    _upload(client)
    html = client.get("/delegated/").get_data(as_text=True)
    assert "Not shown (not a user story)" in html
    assert "Epic ×1" in html
    assert "Defect" not in html.split("Not shown (not a user story)")[1].split("</p>")[0]


def test_upload_auto_registers_defects_as_blockers(client):
    """[USER 2026-08-27: "why cant i see all the defects I uploaded in
    the list of blockers?"] — Defect/Bug/Task issues in the export become
    blocker rows automatically; stories and Epics don't; re-upload never
    duplicates."""
    resp = _upload(client)
    assert "1+blockers+registered" in resp.headers["Location"].replace("%20", "+") \
        or "1 blockers registered" in resp.headers["Location"]
    conn = web_delegated._get_conn()
    try:
        rows = db_blockers.list_blockers(conn)
    finally:
        conn.close()
    assert len(rows) == 1
    b = rows[0]
    assert b["type"] == "defect"
    assert b["jira_key"] == "S4DEF-3001"
    assert b["name"] == "SM3001_Unregistered defect in the export"
    assert b["solman_id"] == "SM3001"  # from the summary prefix
    # blockers list page shows it
    html = client.get("/blockers/").get_data(as_text=True)
    assert "S4DEF-3001" in html
    # re-upload: no duplicate
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        assert len(db_blockers.list_blockers(conn)) == 1
    finally:
        conn.close()


def test_reupload_refreshes_blocker_jira_status_and_autocloses(client):
    """Reproduction of [USER 2026-09-01: "the Status value in Jira is
    never updating" on the Blockers page] — the exact flow: the defect is
    auto-registered on the first upload, gets Resolved in Jira, and the
    next export upload must refresh the shown status AND auto-close the
    blocker."""
    _upload(client)  # S4DEF-3001 arrives as Open, auto-registered
    html = client.get("/blockers/").get_data(as_text=True)
    assert "Open" in html

    resolved = XML.replace(
        "<key id=\"5\">S4DEF-3001</key>\n    "
        "<summary>SM3001_Unregistered defect in the export</summary>\n    "
        "<type id=\"1\">Defect</type>\n    <status id=\"3\">Open</status>",
        "<key id=\"5\">S4DEF-3001</key>\n    "
        "<summary>SM3001_Unregistered defect in the export</summary>\n    "
        "<type id=\"1\">Defect</type>\n    <status id=\"3\">Resolved</status>")
    assert "Resolved" in resolved  # the replace really hit
    _upload(client, data=resolved)

    conn = web_delegated._get_conn()
    try:
        row = conn.execute(
            "SELECT jira_status FROM jira_issues WHERE jira_key='S4DEF-3001'"
        ).fetchone()
    finally:
        conn.close()
    assert row[0] == "Resolved"          # the store refreshed
    html = client.get("/blockers/").get_data(as_text=True)
    assert "Resolved" in html            # the page shows the fresh status
    assert "(closed in Jira)" in html    # and the blocker auto-closed
    # staleness diagnostic: the status carries an "as of <date>" stamp
    # (last_seen = the last upload that contained this key)
    from datetime import date
    assert f"as of {date.today().isoformat()}" in html


def test_dashboard_badge_counts_match_the_board(client):
    """The card badge mirrors the board: no defects, no registered
    blockers [2026-08-27]."""
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        counts = db_delegated.delegated_counts(conn)
        assert counts == {"total": 4, "blocked": 1}  # defect not counted
        b = db_blockers.create_blocker(conn, "defect", "Now a blocker", "S4ECOM-2004")
        counts = db_delegated.delegated_counts(conn)
        assert counts == {"total": 3, "blocked": 1}  # registered blocker drops out
        assert b
    finally:
        conn.close()


def test_reupload_backfills_type_on_existing_rows(client):
    """"can this be fixed for already uploaded issues?" [USER 2026-08-27]
    — yes: the upsert refresh now also writes type, so one normal upload
    corrects rows that were imported without one."""
    # first upload with NO type on S4ECOM-2003
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        row = conn.execute("SELECT type FROM jira_issues WHERE jira_key='S4ECOM-2003'").fetchone()
        assert row[0] is None
    finally:
        conn.close()
    # second upload where S4ECOM-2003 now carries a type
    typed = XML.replace(
        "<key id=\"3\">S4ECOM-2003</key>",
        "<key id=\"3\">S4ECOM-2003</key>\n    <type id=\"1\">Defect</type>")
    _upload(client, data=typed)
    conn = web_delegated._get_conn()
    try:
        row = conn.execute("SELECT type FROM jira_issues WHERE jira_key='S4ECOM-2003'").fetchone()
        assert row[0] == "Defect"
    finally:
        conn.close()
    html = client.get("/delegated/").get_data(as_text=True)
    assert "S4ECOM-2003" not in html  # now filtered as a defect


def test_upload_rejects_non_xml_and_missing_file(client):
    resp = _upload(client, filename="delegated.xlsx")
    assert "jira_ok=0" in resp.headers["Location"]
    resp = client.post("/delegated/upload", data={})
    assert "jira_ok=0" in resp.headers["Location"]


def test_upload_tags_delegated_without_touching_other_sources(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        rows = conn.execute(
            "SELECT seen_in_delegated, seen_in_gatekeeper, seen_in_ecom"
            " FROM jira_issues").fetchall()
    finally:
        conn.close()
    assert len(rows) == 6  # incl. the Defect + Epic issues — tagged in the STORE, filtered in the views
    assert all(r[0] == 1 and r[1] == 0 and r[2] == 0 for r in rows)


def test_blocked_reason_and_next_step_save(client):
    _upload(client)
    assert client.post("/delegated/ticket/S4ECOM-2001/blocked-reason",
                       data={"blocked_reason": "waiting for settlement file"}
                       ).get_json()["ok"]
    assert client.post("/delegated/ticket/S4ECOM-2001/next-step",
                       data={"next_step": "chase GBS on Friday"}).get_json()["ok"]
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_blocked_reason(
            conn, "S4ECOM-2001") == "waiting for settlement file"
        assert db_delegated.get_delegated_next_step(
            conn, "S4ECOM-2001") == "chase GBS on Friday"
    finally:
        conn.close()
    # why-blocked AND the next step left the board for BLOCKED tickets
    # 2026-08-28 [USER: "I need the next step for the blockers - not for
    # the blocked test cases"] — both stay editable on the detail page
    html = client.get("/delegated/").get_data(as_text=True)
    assert "waiting for settlement file" not in html
    assert "chase GBS on Friday" not in html   # S4ECOM-2001 is BLOCKED
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "waiting for settlement file" in detail
    assert "chase GBS on Friday" in detail
    # a NON-blocked ticket keeps its inline next-step field
    client.post("/delegated/ticket/S4ECOM-2003/next-step",
                data={"next_step": "ping the team"})
    html = client.get("/delegated/").get_data(as_text=True)
    assert "ping the team" in html


def test_detail_page_has_tabs_and_working_fields(client):
    _upload(client)
    html = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "Details" in html and "Messages (2)" in html
    assert "Why blocked" in html                       # blocked ticket
    html = client.get("/delegated/ticket/S4ECOM-2003").get_data(as_text=True)
    assert "Why blocked" not in html                   # not blocked
    assert client.get("/delegated/ticket/NOPE-1").status_code == 404


def test_detail_post_saves_both_fields(client):
    _upload(client)
    resp = client.post("/delegated/ticket/S4ECOM-2001", data={
        "blocked_reason": "missing master data", "next_step": "retest Monday"})
    assert resp.status_code == 302
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_blocked_reason(
            conn, "S4ECOM-2001") == "missing master data"
        assert db_delegated.get_delegated_next_step(
            conn, "S4ECOM-2001") == "retest Monday"
    finally:
        conn.close()


def test_status_report_renders_buckets(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2001/blocked-reason",
                data={"blocked_reason": "no settlement file yet"})
    html = client.get("/delegated/report").get_data(as_text=True)
    assert "Delegated Testing Report" in html
    assert "\U0001f534 Blocked" in html
    assert "Marina gatekeeper check" in html
    assert "no settlement file yet" in html            # why-blocked column
    assert "Return Order: 6000084252" in html          # LATEST comment's order
    # the report renders no comment bodies, so the OLDER comment's order
    # must not appear anywhere — proves the latest-comment-only rule
    assert "TBY_SS_ADE0001111" not in html


def test_numbers_report_counts(client):
    _upload(client)
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "Management Summary Status Report" in html
    # blocked 1 · marina 1 · accepted 1 · unexpected 1 — all sections listed,
    # grouped under the 3 review stages [USER 2026-08-27]
    assert "Until Gatekeeper Check" in html
    assert "Past Gatekeeper Check" in html
    assert "Settlement file to be created" in html
    assert "Test case completed" in html
    assert "Unexpected status" in html
    assert "Blocker overview" in html
    # goal box present and editable on the screen page
    assert 'id="goal-input"' in html


def test_numbers_report_goal_actual_counts_post_gatekeeper_and_flagged_blocked(client):
    _upload(client)
    client.post("/delegated/numbers/goal", data={"goal": "5"})
    client.post("/delegated/ticket/S4ECOM-2001/counts-toward-goal", data={"value": "1"})
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert 'value="5"' in html
    # 4 fixture tickets: S4ECOM-2001 Blocked+flagged, S4ECOM-2002/2003 land
    # in the pre-gatekeeper stage, S4ECOM-2004 is unexpected — none reach
    # post-gatekeeper, so actual = 0 (post-gatekeeper) + 1 (flagged blocked)
    assert '<span class="goal-val">1</span>' in html


def test_report_download_is_standalone_attachment(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2001/blocked-reason",
                data={"blocked_reason": "no settlement file yet"})
    resp = client.get("/delegated/report/download")
    assert resp.status_code == 200
    assert 'attachment; filename="delegated_report_' in resp.headers["Content-Disposition"]
    html = resp.get_data(as_text=True)
    assert "Delegated Testing Report" in html
    assert "no settlement file yet" in html            # content intact
    # interactive chrome is stripped: toolbar, filter bar, scripts, inputs
    assert 'class="toolbar"' not in html
    assert 'class="filterbar"' not in html
    assert "<script>" not in html
    assert "co-input" not in html.split("</style>")[1]  # no call-out inputs in the body


def test_report_shows_blocker_chips_and_filter(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        b1 = db_blockers.create_blocker(conn, "defect", "Pricing bug", "S4DEF-1")
        db_blockers.link_blocker(conn, b1["blocker_id"], "S4ECOM-2001")
    finally:
        conn.close()
    html = client.get("/delegated/report").get_data(as_text=True)
    assert "Pricing bug (S4DEF-1)" in html           # chip on the blocked row
    assert f'value="{b1["blocker_id"]}">Pricing bug (S4DEF-1)' in html  # filter option
    assert f'data-blockers="{b1["blocker_id"]}"' in html  # row carries the id for JS filtering


def test_report_download_includes_blocker_chips_but_no_filter(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        b1 = db_blockers.create_blocker(conn, "task", "Backfill master data", None)
        db_blockers.link_blocker(conn, b1["blocker_id"], "S4ECOM-2001")
    finally:
        conn.close()
    html = client.get("/delegated/report/download").get_data(as_text=True)
    assert "Backfill master data" in html   # chip still shown — static display, not interactive
    assert "rf-blocker" not in html         # filter dropdown stripped like the rest of the filterbar


def test_report_page_keeps_its_buttons(client):
    _upload(client)
    html = client.get("/delegated/report").get_data(as_text=True)
    assert "📊 Management Summary" in html
    assert "⬇ Download HTML" in html
    # no Print button on the report [USER 2026-08-26] — Ctrl+P still works
    assert 'onclick="window.print()"' not in html


def test_numbers_download_is_standalone_attachment(client):
    _upload(client)
    resp = client.get("/delegated/numbers/download")
    assert resp.status_code == 200
    assert 'attachment; filename="delegated_numbers_' in resp.headers["Content-Disposition"]
    html = resp.get_data(as_text=True)
    assert "Management Summary Status Report" in html
    assert "Settlement file to be created" in html
    # goal renders as static text, not an editable input, in the download
    assert 'id="goal-input"' not in html
    # toolbar (app-local links) and the copy script are stripped
    assert 'class="toolbar"' not in html
    assert "<script>" not in html


def test_email_choices_and_attachments_include_delegated(client):
    from app import emailer
    assert ("delegated", "Delegated Testing Report") in emailer.REPORT_CHOICES
    assert ("delegated_numbers", "Delegated Testing — Management Summary") in emailer.REPORT_CHOICES
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        atts = emailer.gather_attachments(
            conn, {}, app, ["delegated", "delegated_numbers"], "2026-08-26")
    finally:
        conn.close()
    assert [name for name, _ in atts] == [
        "delegated_report_2026-08-26.html", "delegated_numbers_2026-08-26.html"]
    for _name, html in atts:
        assert 'class="toolbar"' not in html
        assert "<script>" not in html


def test_counts_toward_goal_toggle_via_checkbox(client):
    _upload(client)
    resp = client.post("/delegated/ticket/S4ECOM-2001/counts-toward-goal",
                       data={"value": "1"})
    assert resp.get_json()["ok"]
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_counts_toward_goal(conn, "S4ECOM-2001") is True
    finally:
        conn.close()
    html = client.get("/delegated/").get_data(as_text=True)
    assert 'data-key="S4ECOM-2001"' in html and "checked" in html


def test_counts_toward_goal_saved_via_detail_form(client):
    _upload(client)
    resp = client.post("/delegated/ticket/S4ECOM-2001", data={
        "blocked_reason": "waiting", "next_step": "chase",
        "counts_toward_goal": "1"})
    assert resp.status_code == 302
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_counts_toward_goal(conn, "S4ECOM-2001") is True
    finally:
        conn.close()
    # unchecking (checkbox simply absent from the POST) clears it
    client.post("/delegated/ticket/S4ECOM-2001", data={
        "blocked_reason": "waiting", "next_step": "chase"})
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_counts_toward_goal(conn, "S4ECOM-2001") is False
    finally:
        conn.close()


def test_board_and_detail_show_attached_blocker_chip(client):
    from app.db import blockers as db_blockers
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        b1 = db_blockers.create_blocker(conn, "defect", "Pricing bug", None)
        db_blockers.link_blocker(conn, b1["blocker_id"], "S4ECOM-2001")
    finally:
        conn.close()
    board_html = client.get("/delegated/").get_data(as_text=True)
    assert "Pricing bug" in board_html
    detail_html = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "Pricing bug" in detail_html
    assert "Manage blockers" in detail_html


def test_reimport_refreshes_status(client):
    _upload(client)
    updated = XML.replace(">Blocked<", ">In Review<")
    _upload(client, updated, "delegated_v2.xml")
    html = client.get("/delegated/").get_data(as_text=True)
    assert "ECOM BPO test" in html
    conn = web_delegated._get_conn()
    try:
        row = conn.execute("SELECT jira_status FROM jira_issues"
                           " WHERE jira_key='S4ECOM-2001'").fetchone()
    finally:
        conn.close()
    assert row[0] == "In Review"


# ---------------------------------------------------------------------------
# Management Summary: call-outs + open-blockers-only overview [USER 2026-08-27]

def test_numbers_calls_out_section_and_add_route(client):
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "callout-section" in html
    assert "+ Add call-out" in html
    # regression: 'delegated' was missing from the add-route allowlist since
    # 2026-08-26 — both delegated keys must be accepted now
    for key in ("delegated", "delegated_numbers"):
        resp = client.post(f"/report-comments/{key}/add", data={"comment": "note"})
        assert resp.get_json()["ok"], key
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert 'value="note"' in html            # delegated_numbers call-out shows
    report_html = client.get("/delegated/report").get_data(as_text=True)
    assert 'value="note"' in report_html     # delegated call-out on ITS report


def test_numbers_download_shows_callouts_static(client):
    client.post("/report-comments/delegated_numbers/add",
                data={"comment": "big note for mgmt"})
    html = client.get("/delegated/numbers/download").get_data(as_text=True)
    assert "big note for mgmt" in html
    assert "co-input" not in html.split("</style>")[1]  # static text, no inputs
    assert "+ Add call-out" not in html


def test_numbers_blocker_overview_hides_closed_blockers(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        open_b = db_blockers.create_blocker(conn, "defect", "OpenBlocker", "S4XYZ-1")
        manual = db_blockers.create_blocker(conn, "task", "ManuallyClosed", None)
        db_blockers.set_blocker_closed(conn, manual["blocker_id"], True)
        # jira-done blocker: S4ECOM-2004 is in the store (fixture) — set done
        auto = db_blockers.create_blocker(conn, "defect", "JiraDoneBlocker", "S4ECOM-2004")
        conn.execute("UPDATE jira_issues SET jira_status='Resolved'"
                     " WHERE jira_key='S4ECOM-2004'")
        conn.commit()
    finally:
        conn.close()
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "OpenBlocker" in html
    assert "ManuallyClosed" not in html
    assert "JiraDoneBlocker" not in html
    assert open_b and auto


# ---------------------------------------------------------------------------
# Backlog [USER 2026-08-27]: park tickets in their own section, excluded
# from the Management Summary

def test_backlog_ticket_moves_to_own_section_on_board_and_report(client):
    _upload(client)
    resp = client.post("/delegated/ticket/S4ECOM-2003/backlog", data={"value": "1"})
    assert resp.get_json()["ok"]

    html = client.get("/delegated/").get_data(as_text=True)
    assert "<summary>📦 Backlog" in html  # the section header, not a tooltip
    backlog_part = html.split("<summary>📦 Backlog")[1]
    assert "S4ECOM-2003" in backlog_part
    # gone from its old bucket (Testing team creating order)
    team_part = html.split("<summary>Testing team creating order")[1].split("<summary>📦 Backlog")[0]
    assert 'data-key="S4ECOM-2003"' not in team_part

    report_html = client.get("/delegated/report").get_data(as_text=True)
    assert ">📦 Backlog</span>" in report_html
    assert "S4ECOM-2003" in report_html.split(">📦 Backlog</span>")[1]


def test_backlog_excluded_from_management_summary(client):
    _upload(client)
    before = client.get("/delegated/numbers").get_data(as_text=True)
    client.post("/delegated/ticket/S4ECOM-2003/backlog", data={"value": "1"})
    after = client.get("/delegated/numbers").get_data(as_text=True)

    def total_of(html):
        return int(html.split('rh-stat-val">')[1].split("<")[0])
    assert total_of(after) == total_of(before) - 1
    assert "📦 Backlog" not in after  # no backlog row/section in the summary

    # unpark → back in the numbers
    client.post("/delegated/ticket/S4ECOM-2003/backlog", data={"value": "0"})
    restored = client.get("/delegated/numbers").get_data(as_text=True)
    assert total_of(restored) == total_of(before)


def test_backlog_wins_over_blocked_and_button_is_the_one_control(client):
    _upload(client)
    # S4ECOM-2001 is Blocked — parking it moves it OUT of the Blocker section
    client.post("/delegated/ticket/S4ECOM-2001/backlog", data={"value": "1"})
    html = client.get("/delegated/").get_data(as_text=True)
    blocked_part = html.split("<summary>🔴 Blocked")[1].split("<summary>")[0]
    assert "S4ECOM-2001" not in blocked_part
    assert "S4ECOM-2001" in html.split("<summary>📦 Backlog")[1]
    # no backlog checkbox column on the board [USER 2026-09-01]
    assert "dlg-backlog" not in html

    # detail page shows the parked state + the way back — no form checkbox
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "Move back to board" in detail
    assert 'name="backlog"' not in detail

    # saving the working-fields form must NOT unpark the ticket
    client.post("/delegated/ticket/S4ECOM-2001", data={
        "next_step": "", "blocked_reason": "", "counts_toward_goal": "0"})
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_backlog(conn, "S4ECOM-2001") is True
    finally:
        conn.close()

    # unpark via the button's route — the detail page offers parking again
    client.post("/delegated/ticket/S4ECOM-2001/backlog", data={"value": "0"})
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "Move to backlog" in detail
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_backlog(conn, "S4ECOM-2001") is False
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# nextInLine label rule (2026-09-01 [USER]): only labeled tickets go into
# "Not started yet" — the rest wait in Backlog; manual override wins.

OPEN_XML = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="0.92"><channel>
  <item>
    <key id="1">S4ECOM-3101</key>
    <summary>SM3101_Pulled next</summary>
    <type id="17">Story</type>
    <status id="1">Open</status>
    <labels><label>nextInLine</label></labels>
  </item>
  <item>
    <key id="2">S4ECOM-3102</key>
    <summary>SM3102_Waiting its turn</summary>
    <type id="17">Story</type>
    <status id="1">Open</status>
    <labels><label>fr_scope</label></labels>
  </item>
</channel></rss>
"""


def test_next_in_line_label_decides_board_placement(client):
    _upload(client, data=OPEN_XML)
    html = client.get("/delegated/").get_data(as_text=True)
    open_part = html.split("<summary>Not started yet")[1].split("<summary>")[0]
    backlog_part = html.split("<summary>📦 Backlog")[1]
    assert "S4ECOM-3101" in open_part        # labeled → board
    assert "S4ECOM-3102" not in open_part
    assert "S4ECOM-3102" in backlog_part     # unlabeled → parked
    # the label-parked ticket is out of the Management Summary total
    summary = client.get("/delegated/numbers").get_data(as_text=True)
    assert "S4ECOM-3102" not in summary


def test_manual_unpark_beats_the_missing_label(client):
    _upload(client, data=OPEN_XML)
    # she pulls the unlabeled ticket onto the board by hand
    client.post("/delegated/ticket/S4ECOM-3102/backlog", data={"value": "0"})
    html = client.get("/delegated/").get_data(as_text=True)
    open_part = html.split("<summary>Not started yet")[1].split("<summary>")[0]
    assert "S4ECOM-3102" in open_part
    # and the override survives the next upload (labels refresh, manual wins)
    _upload(client, data=OPEN_XML)
    html = client.get("/delegated/").get_data(as_text=True)
    open_part = html.split("<summary>Not started yet")[1].split("<summary>")[0]
    assert "S4ECOM-3102" in open_part


def test_label_added_in_jira_moves_ticket_on_next_upload(client):
    _upload(client, data=OPEN_XML)
    # the team labels the waiting ticket in Jira → next export upload
    labeled = OPEN_XML.replace("<labels><label>fr_scope</label></labels>",
                               "<labels><label>fr_scope</label>"
                               "<label>nextInLine</label></labels>")
    _upload(client, data=labeled)
    html = client.get("/delegated/").get_data(as_text=True)
    open_part = html.split("<summary>Not started yet")[1].split("<summary>")[0]
    assert "S4ECOM-3102" in open_part


def test_detail_page_shows_label_parked_state(client):
    _upload(client, data=OPEN_XML)
    detail = client.get("/delegated/ticket/S4ECOM-3102").get_data(as_text=True)
    assert "Parked in Backlog" in detail     # effective state, no manual flag set
    detail = client.get("/delegated/ticket/S4ECOM-3101").get_data(as_text=True)
    assert "Move to backlog" in detail       # labeled → on the board


def test_backlog_tristate_migration_rebuilds_not_null_column(tmp_path):
    """One-time migration: the old NOT NULL DEFAULT 0 column becomes
    nullable; old 0 (mostly 'never touched') → NULL so the label rule can
    take over, manual parks (1) survive."""
    import sqlite3 as sq
    db_path = tmp_path / "mig.db"
    conn = sq.connect(db_path)
    conn.executescript("""
        CREATE TABLE delegated_annotations (
            jira_key TEXT PRIMARY KEY, blocked_reason TEXT, next_step TEXT,
            updated_at TEXT, counts_toward_goal INTEGER NOT NULL DEFAULT 0,
            backlog INTEGER NOT NULL DEFAULT 0,
            req_tool INTEGER NOT NULL DEFAULT 0, sales_xls TEXT);
        INSERT INTO delegated_annotations (jira_key, next_step, backlog, sales_xls)
        VALUES ('S4ECOM-1', 'call Tom', 0, 'maybe'), ('S4ECOM-2', NULL, 1, NULL);
    """)
    conn.commit()
    conn.close()
    db_delegated.init_schema(db_path)
    conn = database.get_connection(db_path)
    try:
        assert db_delegated.get_delegated_backlog(conn, "S4ECOM-1") is None
        assert db_delegated.get_delegated_backlog(conn, "S4ECOM-2") is True
        # the other authored fields came through untouched
        ann = db_delegated.get_delegated_annotations(conn)
        assert ann["S4ECOM-1"]["next_step"] == "call Tom"
        assert ann["S4ECOM-1"]["sales_xls"] == "maybe"
        # idempotent — a second init must not error or change anything
        conn.close()
        db_delegated.init_schema(db_path)
        conn = database.get_connection(db_path)
        assert db_delegated.get_delegated_backlog(conn, "S4ECOM-2") is True
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# ReqTool [USER 2026-08-29]: dashboard-only authored flag, filterable on the
# board — deliberately absent from both reports

def test_req_tool_toggle_shows_on_board_and_filters(client):
    _upload(client)
    resp = client.post("/delegated/ticket/S4ECOM-2003/req-tool", data={"value": "1"})
    assert resp.get_json()["ok"]

    html = client.get("/delegated/").get_data(as_text=True)
    assert 'class="dlg-reqtool"' in html  # the board column renders
    assert 'data-reqtool="1"' in html  # at least one row carries the checked marker
    assert 'ReqTool: checked' in html and 'ReqTool: unchecked' in html

    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_req_tool(conn, "S4ECOM-2003") is True
    finally:
        conn.close()

    client.post("/delegated/ticket/S4ECOM-2003/req-tool", data={"value": "0"})
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_req_tool(conn, "S4ECOM-2003") is False
    finally:
        conn.close()


def test_req_tool_detail_form_roundtrip(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2001/req-tool", data={"value": "1"})
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert 'name="req_tool" value="1" checked' in detail

    # unticking via the detail form's missing checkbox clears it
    client.post("/delegated/ticket/S4ECOM-2001", data={
        "next_step": "", "blocked_reason": "", "counts_toward_goal": "0"})
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_req_tool(conn, "S4ECOM-2001") is False
    finally:
        conn.close()


def test_req_tool_absent_from_both_reports(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2003/req-tool", data={"value": "1"})
    report_html = client.get("/delegated/report").get_data(as_text=True)
    numbers_html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "ReqTool" not in report_html
    assert "ReqTool" not in numbers_html


# ---------------------------------------------------------------------------
# SalesXLS [USER 2026-09-01]: dashboard-only TRI-STATE marker (yes/no/maybe,
# unset = not assessed) — cycling chip on board + detail page, filterable,
# deliberately absent from all three reports (like ReqTool)

def test_sales_xls_roundtrip_and_validation(client):
    _upload(client)
    for value in ("yes", "maybe", "no"):
        resp = client.post("/delegated/ticket/S4ECOM-2003/sales-xls",
                           data={"value": value})
        assert resp.get_json()["ok"]
        conn = web_delegated._get_conn()
        try:
            assert db_delegated.get_delegated_sales_xls(
                conn, "S4ECOM-2003") == value
        finally:
            conn.close()

    # empty value = back to not assessed (the cycle's fourth state)
    client.post("/delegated/ticket/S4ECOM-2003/sales-xls", data={"value": ""})
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2003") is None
    finally:
        conn.close()

    # garbage is rejected with 400 and nothing is stored
    resp = client.post("/delegated/ticket/S4ECOM-2003/sales-xls",
                       data={"value": "perhaps"})
    assert resp.status_code == 400
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2003") is None
    finally:
        conn.close()


def test_sales_xls_renders_on_board_and_detail_with_filter(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2003/sales-xls", data={"value": "maybe"})

    html = client.get("/delegated/").get_data(as_text=True)
    assert 'class="chip dlg-salesxls"' in html   # the chip column renders
    assert 'data-salesxls="maybe"' in html       # row marker feeds the filter
    # the filter dropdown with all four states
    assert "SalesXLS: all" in html
    for option in ("SalesXLS: ✓ Yes", "SalesXLS: ? Maybe",
                   "SalesXLS: ✗ No", "SalesXLS: — not set"):
        assert option in html

    detail = client.get("/delegated/ticket/S4ECOM-2003").get_data(as_text=True)
    assert 'dlg-salesxls' in detail and 'data-value="maybe"' in detail


def test_sales_xls_absent_from_all_three_reports(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2003/sales-xls", data={"value": "yes"})
    for path in ("/delegated/report", "/delegated/numbers",
                 "/delegated/overview"):
        assert "SalesXLS" not in client.get(path).get_data(as_text=True)


# ---------------------------------------------------------------------------
# Ways of Working [USER 2026-09-01]: migrated into the generic working-notes
# pages the same day (app/note_pages.py, tests/test_note_pages.py) — from
# the delegated side only the buttons and the old-URL redirects remain.

def test_wow_and_insights_buttons_and_old_url_redirects(client):
    html = client.get("/delegated/").get_data(as_text=True)
    assert "Ways of working" in html
    assert "Testing insights" in html
    resp = client.get("/delegated/wow")
    assert resp.status_code == 302
    assert "/notes-page/delegated_wow" in resp.headers["Location"]
    resp = client.get("/delegated/wow/download")
    assert resp.status_code == 302
    assert "/notes-page/delegated_wow/download" in resp.headers["Location"]


# ---------------------------------------------------------------------------
# Jira labels (2026-08-28) - imported, chips + filter on board/report/detail

def test_labels_imported_and_shown_with_filter(client):
    _upload(client)
    html = client.get("/delegated/").get_data(as_text=True)
    # chips on the row + data attribute for the client-side filter
    assert "settlement" in html and "fr_scope" in html
    assert 'data-labels="fr_scope settlement"' in html
    # the filter dropdown lists the distinct labels once each
    bar = html.split('id="dlg-label-filter"')[1].split("</select>")[0]
    assert bar.count('value="settlement"') == 1
    assert bar.count('value="fr_scope"') == 1


def test_labels_replaced_on_reimport(client):
    _upload(client)
    _upload(client, data=XML.replace(
        "<label>settlement</label><label>fr_scope</label>",
        "<label>renamed_label</label>"))
    html = client.get("/delegated/").get_data(as_text=True)
    assert "renamed_label" in html
    assert "fr_scope" not in html


def test_report_has_no_labels_but_detail_does(client):
    # labels left the report again 2026-08-28 [USER: "not interesting"];
    # the detail view lists them (Details tab)
    _upload(client)
    report = client.get("/delegated/report").get_data(as_text=True)
    assert 'id="rf-label"' not in report
    assert "fr_scope" not in report
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "Labels" in detail
    assert "fr_scope · settlement" in detail


# ---------------------------------------------------------------------------
# Blocker impact on the Management Summary (2026-08-28)

def test_blocker_impact_editable_on_numbers_page(client):
    _upload(client)   # auto-registers S4DEF-3001 as a defect blocker
    conn = web_delegated._get_conn()
    try:
        blocker = next(b for b in db_blockers.list_blockers(conn)
                       if b["jira_key"] == "S4DEF-3001")
    finally:
        conn.close()
    resp = client.post(f"/blockers/{blocker['blocker_id']}/impact",
                       data={"impact": "blocks FR settlement retests"})
    assert resp.get_json()["ok"]

    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "Impact" in html
    assert 'value="blocks FR settlement retests"' in html   # editable input
    download = client.get("/delegated/numbers/download").get_data(as_text=True)
    assert "blocks FR settlement retests" in download        # static text
    assert 'class="blk-impact"' not in download


# ---------------------------------------------------------------------------
# MB tracking join (2026-08-28) - ECOM tab of the UAT tracking workbook

def _tracking_xlsx(status_for_2001="Not Ready", extra_rows=None) -> bytes:
    import openpyxl
    headers = ["Status", "Assigned to", "Country", "Testcase Scenario",
               "Test Case ID", "Testcase Name", "Description Change",
               "Jira ID", "Date execution started",
               "Order Number/Transaction Number",
               "Defect ID (if applicable)", "S4 Sales Order",
               "S4 Billing Documents", "S4 Journal Invoice Entry",
               "Delivery Note (for TradeCo)",
               "Reason for pass with reservation",
               "Old Order Numbers/Transaction Numbers", "Comments"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ECOM"
    ws.append(headers)
    ws.append([status_for_2001, "Tom", "FR", "Scenario X", "TC-42",
               "Standard order FR", None, "S4ECOM-2001", None, "600123",
               "DEF-7", "SO-1", "BD-1", "JE-1", None, "minor diff", None,
               "watch the refund"])
    for r in (extra_rows or []):
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_tracking(client, data=None,
                     filename="DTC_UAT_testtracking_ROE(31).xlsx"):
    data = data if data is not None else _tracking_xlsx()
    return client.post("/delegated/upload-tracking", data={
        "file": (io.BytesIO(data), filename)})


def test_tracking_upload_imports_ecom_tab(client):
    _upload(client)
    resp = _upload_tracking(client)
    assert "jira_ok=1" in resp.headers["Location"]
    html = client.get(resp.headers["Location"]).get_data(as_text=True)
    assert "1 new" in html
    # wrong files rejected
    resp = _upload_tracking(client, filename="whatever.xlsx")
    assert "jira_ok=0" in resp.headers["Location"]
    resp = _upload_tracking(client, filename="notes.txt")
    assert "jira_ok=0" in resp.headers["Location"]


def test_board_mb_status_column_with_mismatch_color(client):
    _upload(client)
    # before any tracking upload: the blocked ticket has no ECOM row ->
    # neutral dash with explanatory title
    html = client.get("/delegated/").get_data(as_text=True)
    assert "no ECOM-tab row for this Jira ID yet" in html
    # S4ECOM-2001 is in the Blocker bucket; "Not Ready" is NOT an expected
    # blocked MB status -> red mismatch chip
    _upload_tracking(client)
    html = client.get("/delegated/").get_data(as_text=True)
    assert "MB Status" in html
    assert 'background:#c1121f' in html and "Not Ready" in html
    # now with an expected blocked wording -> plain text, no red chip on it
    _upload_tracking(client, data=_tracking_xlsx(
        status_for_2001="Blocked - returned to Sales"))
    html = client.get("/delegated/").get_data(as_text=True)
    assert "Blocked - returned to Sales" in html
    mb_cell = html.split("Blocked - returned to Sales")[0][-200:]
    assert "c1121f" not in mb_cell


def test_ticket_detail_shows_mb_card(client):
    _upload(client)
    _upload_tracking(client)
    html = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "MB tracking (ECOM tab)" in html
    for value in ("TC-42", "Standard order FR", "Not Ready", "DEF-7",
                  "SO-1", "BD-1", "JE-1", "minor diff", "watch the refund"):
        assert value in html, value
    # a ticket without an ECOM row shows no MB card
    html = client.get("/delegated/ticket/S4ECOM-2002").get_data(as_text=True)
    assert "MB tracking (ECOM tab)" not in html


# ---------------------------------------------------------------------------
# SalesXLS auto-match upload (2026-09-01 [USER]) — sales workbook's
# "All Countries Combined" tab, "Solman ID" column, matched by substring
# against each board ticket's raw Jira Summary.

def _sales_xls_xlsx(solman_ids=("SM2001", "SM2003"), header="Solman ID") -> bytes:
    # the real header is "Solman ID" [USER 2026-09-02: not "SolmanID"]
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Countries Combined"
    ws.append([header, "Other Column"])
    for sid in solman_ids:
        ws.append([sid, "x"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_sales_xls(client, data=None, filename="sales.xlsx"):
    data = data if data is not None else _sales_xls_xlsx()
    return client.post("/delegated/upload-sales-xls", data={
        "file": (io.BytesIO(data), filename)})


def test_sales_xls_upload_marks_matches_yes_and_unassessed_no(client):
    _upload(client)  # S4ECOM-2001..2004 board-visible (Blocker/Marina/Accepted/Unexpected)
    resp = _upload_sales_xls(client)
    assert "jira_ok=1" in resp.headers["Location"]
    html = client.get(resp.headers["Location"]).get_data(as_text=True)
    assert "2 marked Yes" in html
    assert "2 newly marked No" in html

    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2001") == "yes"  # SM2001
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2003") == "yes"  # SM2003
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2002") == "no"   # SM2002, no match
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2004") == "no"   # SM2004, no match
    finally:
        conn.close()


def test_sales_xls_upload_never_overwrites_an_existing_assessment_with_no(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        db_delegated.set_delegated_sales_xls(conn, "S4ECOM-2002", "maybe")
    finally:
        conn.close()
    _upload_sales_xls(client)  # SM2001/SM2003 match, SM2002 does not
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2002") == "maybe"
    finally:
        conn.close()


def test_sales_xls_upload_always_overwrites_a_match_regardless_of_prior_value(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        db_delegated.set_delegated_sales_xls(conn, "S4ECOM-2001", "no")
    finally:
        conn.close()
    _upload_sales_xls(client)  # SM2001 matches S4ECOM-2001
    conn = web_delegated._get_conn()
    try:
        assert db_delegated.get_delegated_sales_xls(conn, "S4ECOM-2001") == "yes"
    finally:
        conn.close()


def test_sales_xls_upload_rejects_non_xlsx_and_missing_file(client):
    resp = client.post("/delegated/upload-sales-xls", data={})
    assert "jira_ok=0" in resp.headers["Location"]
    resp = _upload_sales_xls(client, filename="notes.txt")
    assert "jira_ok=0" in resp.headers["Location"]


def test_sales_xls_upload_rejects_missing_sheet_or_column(client):
    _upload(client)
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong Sheet"
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload_sales_xls(client, data=buf.getvalue())
    assert "jira_ok=0" in resp.headers["Location"]
    html = client.get(resp.headers["Location"]).get_data(as_text=True)
    assert "All Countries Combined" in html


# ---------------------------------------------------------------------------
# Report tweaks + call-out archive (2026-08-28)

def test_report_blocker_chips_are_id_only(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        blocker = next(b for b in db_blockers.list_blockers(conn)
                       if b["jira_key"] == "S4DEF-3001")
        db_blockers.link_blocker(conn, blocker["blocker_id"], "S4ECOM-2001")
    finally:
        conn.close()
    html = client.get("/delegated/report").get_data(as_text=True)
    chips = html.split('class="rpt-blockers"')[1].split("</td>")[0]
    assert "S4DEF-3001" in chips
    # the name lives in the tooltip only
    assert "SM3001_Unregistered defect in the export" not in chips.replace(
        'title="defect: SM3001_Unregistered defect in the export"', "")


def test_callout_archive_keeps_dates_and_leaves_live_list(client):
    _upload(client)
    resp = client.post("/report-comments/delegated/add",
                       data={"comment": "settlement files delayed"})
    cid = resp.get_json()["row"]["id"]
    resp = client.post(f"/report-comments/{cid}/archive")
    assert resp.get_json()["ok"]

    html = client.get("/delegated/report").get_data(as_text=True)
    # out of the live editable list, present in the archive expander
    assert 'value="settlement files delayed"' not in html
    assert "Archived call-outs (1)" in html
    assert "settlement files delayed" in html
    assert "archived" in html.split("Archived call-outs")[1][:400]
    # the download shows neither (live only, no archive section)
    download = client.get("/delegated/report/download").get_data(as_text=True)
    assert "settlement files delayed" not in download


def test_blockers_list_has_editable_impact_column(client):
    _upload(client)
    html = client.get("/blockers/").get_data(as_text=True)
    assert "Impact" in html and 'class="rt-comment blk-impact"' in html


# ---------------------------------------------------------------------------
# Responsible team per blocker + Mgmt Summary call-out archive (2026-08-28)

def test_blocker_team_fixed_other_and_learned_options(client):
    _upload(client)   # registers the S4DEF-3001 defect blocker
    conn = web_delegated._get_conn()
    try:
        blocker = next(b for b in db_blockers.list_blockers(conn)
                       if b["jira_key"] == "S4DEF-3001")
        bid = blocker["blocker_id"]
    finally:
        conn.close()
    # fixed pick via the inline route
    resp = client.post(f"/blockers/{bid}/team", data={"team": "PDM"})
    assert resp.get_json()["ok"]
    html = client.get("/blockers/").get_data(as_text=True)
    assert "Team" in html
    assert f'<option value="PDM" selected>' in html
    # custom "Other" value via the detail form
    resp = client.post(f"/blockers/{bid}", data={
        "type": "defect", "name": blocker["name"],
        "jira_key": "S4DEF-3001", "team": "__other__",
        "team_other": "Warehouse IT"})
    assert resp.status_code == 302
    conn = web_delegated._get_conn()
    try:
        assert db_blockers.get_blocker(conn, bid)["team"] == "Warehouse IT"
        # the learned value joins the combobox after the fixed teams
        assert db_blockers.team_options(conn) == \
            db_blockers.FIXED_TEAMS + ["Warehouse IT"]
    finally:
        conn.close()
    # visible on the blockers list, the Mgmt Summary and the board chip
    assert "Warehouse IT" in client.get("/blockers/").get_data(as_text=True)
    numbers = client.get("/delegated/numbers").get_data(as_text=True)
    assert "Warehouse IT" in numbers
    conn = web_delegated._get_conn()
    try:
        db_blockers.link_blocker(conn, bid, "S4ECOM-2001")
    finally:
        conn.close()
    board = client.get("/delegated/").get_data(as_text=True)
    assert "Warehouse IT" in board
    report = client.get("/delegated/report").get_data(as_text=True)
    assert "Warehouse IT" in report


def test_numbers_callout_archive(client):
    _upload(client)
    resp = client.post("/report-comments/delegated_numbers/add",
                       data={"comment": "goal raised to 25"})
    cid = resp.get_json()["row"]["id"]
    client.post(f"/report-comments/{cid}/archive")
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert 'value="goal raised to 25"' not in html   # off the live list
    assert "Archived call-outs (1)" in html
    assert "goal raised to 25" in html
    download = client.get("/delegated/numbers/download").get_data(as_text=True)
    assert "goal raised to 25" not in download


# ---------------------------------------------------------------------------
# Board slimming + MB robustness (2026-08-28, follow-up batch)

def test_board_is_slim_but_filter_and_detail_keep_the_data(client):
    _upload(client)
    html = client.get("/delegated/").get_data(as_text=True)
    # gone from the board: label chips, Why blocked column, chat +
    # message buttons, the Orders POPUP button
    assert 'class="chip chip--none"' not in html
    assert ">Why blocked</th>" not in html
    assert "js-open-msg" not in html
    assert "js-open-orders" not in html
    # still there: the Orders COLUMN [USER 2026-08-28: "I like those"],
    # the Label filter (fed by data-labels), Details
    assert ">Orders</th>" in html
    assert "Return Order: 6000084252" in html
    assert 'id="dlg-label-filter"' in html
    assert 'data-labels="fr_scope settlement"' in html
    # detail view still carries labels + chat/message + order details
    detail = client.get("/delegated/ticket/S4ECOM-2001").get_data(as_text=True)
    assert "fr_scope" in detail
    assert "js-open-msg" in detail
    assert "js-open-orders" in detail


def test_gbs_accepts_ready_for_validation():
    from app.delegated_buckets import mb_status_state
    assert mb_status_state("gbs", {"status": "Ready for Validation"}) == "ok"


def test_mb_join_token_fallback(client):
    """The workbook Jira-ID cell may carry more than the bare key -> the
    exact match misses; the token scan still finds it."""
    _upload(client)
    xlsx = _tracking_xlsx()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["ECOM"]
    ws.cell(2, 8, "S4ECOM-2001 / S4ECOM-9999 (retest)")   # Jira ID column H
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload_tracking(client, data=buf.getvalue())
    assert "MB+rows+match+1" in resp.headers["Location"].replace("%20", "+")
    html = client.get("/delegated/").get_data(as_text=True)
    assert "Not Ready" in html   # matched despite the messy cell


def test_numbers_blocker_overview_shows_next_step(client):
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        blocker = next(b for b in db_blockers.list_blockers(conn)
                       if b["jira_key"] == "S4DEF-3001")
        db_blockers.set_blocker_next_step(conn, blocker["blocker_id"],
                                          "escalate to PDM lead")
    finally:
        conn.close()
    html = client.get("/delegated/numbers").get_data(as_text=True)
    assert "Next step" in html and "escalate to PDM lead" in html
    # blockers list lost the Notes column
    blockers_html = client.get("/blockers/").get_data(as_text=True)
    assert ">Notes</th>" not in blockers_html


# ---------------------------------------------------------------------------
# Delegated Testing Overview — the management report (2026-08-31)

def test_overview_report_renders_pipeline_bar_and_blockers_by_team(client):
    _upload(client)
    # register a blocker with a team, attached to the blocked ticket, so the
    # Blocked line can be staged by the team [USER 2026-08-31]
    conn = web_delegated._get_conn()
    try:
        bid = db_blockers.create_blocker(conn, "defect", "KibanaFeed", "S4XYZ-9")["blocker_id"]
        db_blockers.set_blocker_team(conn, bid, "Kibana")
        db_blockers.link_blocker(conn, bid, "S4ECOM-2001")
    finally:
        conn.close()

    html = client.get("/delegated/overview").get_data(as_text=True)
    assert "Delegated Testing Overview" in html
    # the four stage cards, with the owner line under each
    for label in ("TECH TEST EXECUTION", "MB EXECUTION &amp; VERIFICATION",
                  "ECOM BPO VERIFICATION", "COMPLETE"):
        assert label in html
    assert "owner · Sales Tech" in html
    # execution status bar + its legend
    assert "Execution status" in html
    for group in ("Passed", "In Progress", "Blocked", "Not Started"):
        assert group in html
    # blocker overview grouped by TEAM, not by type
    assert "Blockers by team" in html
    assert "Kibana" in html and "KibanaFeed" in html
    # call-outs live on their own key
    assert "/report-comments/delegated_overview/add" in html


def test_overview_stages_the_blocked_ticket_by_its_blocker_team(client):
    """The blocked ticket counts on the BPO card's Blocked line — its
    blocker's team is Kibana [USER 2026-08-31]."""
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        bid = db_blockers.create_blocker(conn, "defect", "KibanaFeed", "S4XYZ-9")["blocker_id"]
        db_blockers.set_blocker_team(conn, bid, "Kibana")
        db_blockers.link_blocker(conn, bid, "S4ECOM-2001")
    finally:
        conn.close()
    conn = web_delegated._get_conn()
    try:
        ctx = web_delegated.overview_context(conn)
    finally:
        conn.close()
    stages = {s["key"]: s for s in ctx["stages"]}
    assert stages["bpo"]["blocked"] == 1
    assert ctx["blocked_unassigned"] == 0
    # 2001 blocked · 2002 marina · 2003 accepted · 2004 unexpected
    assert ctx["total"] == 4
    assert stages["tech"]["in_progress"] == 1      # 2003 Accepted
    assert stages["mb"]["in_progress"] == 1        # 2002 In Progress
    assert ctx["unexpected"] == 1


def test_overview_teamless_blocker_is_reported_as_unstaged(client):
    """No team on the blocker -> the ticket is called out, never dropped."""
    _upload(client)
    conn = web_delegated._get_conn()
    try:
        bid = db_blockers.create_blocker(conn, "defect", "NoTeamBlocker", "S4XYZ-8")["blocker_id"]
        db_blockers.link_blocker(conn, bid, "S4ECOM-2001")
        ctx = web_delegated.overview_context(conn)
    finally:
        conn.close()
    assert ctx["blocked_unassigned"] == 1
    assert all(s["blocked"] == 0 for s in ctx["stages"])
    html = client.get("/delegated/overview").get_data(as_text=True)
    assert "could not be staged" in html
    assert "No team assigned" in html          # blocker table group


def test_overview_backlog_ticket_is_out_of_the_report(client):
    _upload(client)
    client.post("/delegated/ticket/S4ECOM-2003/backlog", data={"value": "1"})
    conn = web_delegated._get_conn()
    try:
        ctx = web_delegated.overview_context(conn)
    finally:
        conn.close()
    assert ctx["total"] == 3
    stages = {s["key"]: s for s in ctx["stages"]}
    assert stages["tech"]["in_progress"] == 0


def test_overview_download_is_a_standalone_attachment(client):
    _upload(client)
    resp = client.get("/delegated/overview/download")
    assert resp.status_code == 200
    assert 'attachment; filename="delegated_overview_' in resp.headers["Content-Disposition"]
    html = resp.get_data(as_text=True)
    assert "Delegated Testing Overview" in html
    assert "TECH TEST EXECUTION" in html
    # toolbar + scripts stripped, call-outs static
    assert "coSave(" not in html
    assert "Download HTML" not in html


def test_reports_name_the_unexpected_status(client):
    """[USER 2026-09-01] the fixture's S4ECOM-2004 sits on "Ready for
    Verification"; both aggregate reports must SAY so, not just count it."""
    _upload(client)
    numbers = client.get("/delegated/numbers").get_data(as_text=True)
    assert "Unexpected status" in numbers
    assert "Ready for Verification ×1" in numbers

    overview = client.get("/delegated/overview").get_data(as_text=True)
    assert "unexpected Jira status" in overview
    assert "Ready for Verification" in overview

    # and in the downloads, which is where management actually reads them
    dl = client.get("/delegated/overview/download").get_data(as_text=True)
    assert "Ready for Verification" in dl


def test_reopened_is_bucketed_like_open(client):
    """[USER 2026-09-01: "I have the status Reopened - which is to be
    treated exactly the same as opened"]. The ticket carries nextInLine
    here because since the label rule (same day) a not-started ticket
    without it waits in Backlog instead — that path has its own tests."""
    xml = XML.replace(">Ready for Verification<", ">Reopened<").replace(
        '<key id="4">S4ECOM-2004</key>',
        '<key id="4">S4ECOM-2004</key>\n'
        '    <labels><label>nextInLine</label></labels>')
    _upload(client, xml, "delegated_reopened.xml")
    conn = web_delegated._get_conn()
    try:
        ctx = web_delegated.overview_context(conn)
    finally:
        conn.close()
    assert ctx["unexpected"] == 0            # no longer an odd status
    assert ctx["unexpected_statuses"] == []
    stages = {s["key"]: s for s in ctx["stages"]}
    assert stages["tech"]["in_progress"] == 2   # Accepted + the Reopened one
    html = client.get("/delegated/").get_data(as_text=True)
    # split on the SECTION heading — the count strip names it first (2026-09-02)
    not_started = html.split("<summary>Not started yet")[1].split("<summary>")[0]
    assert 'data-key="S4ECOM-2004"' in not_started


def test_sales_xls_column_header_matches_with_or_without_the_space(tmp_path):
    """[USER 2026-09-02]: the column is called "Solman ID" (with a space) —
    the first build looked for "SolmanID" and never found it. Both spell
    the same column; case and spaces are ignored, anything else is loud."""
    from app.read_defects import ParseError
    from app.sales_xls_importer import parse_sales_xls
    for header in ("Solman ID", "SolmanID", "solman id ", "SOLMAN ID"):
        path = tmp_path / "sales.xlsx"
        path.write_bytes(_sales_xls_xlsx(("SM1", "SM2"), header=header))
        assert parse_sales_xls(path) == ["SM1", "SM2"], header
    path = tmp_path / "wrong.xlsx"
    path.write_bytes(_sales_xls_xlsx(("SM1",), header="Solman"))
    with pytest.raises(ParseError, match="Solman ID"):
        parse_sales_xls(path)
