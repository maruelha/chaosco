"""Core South Sustainphase Monitoring card (build plan step 3, 2026-08-27):
upload page + file-picker import wiring."""
import io

import openpyxl
import pytest

from app import database
from app.db import sustain as db_sustain
import app.web_sustain as web_sustain
from app.web import app

HEADERS = ["Task ID", "L4 Taxonomy", "Process / Task", "Cadence",
           "Due Today", "Country", "Provider / Partner / Financial Account",
           "France Result", "Italy Result", "Portugal Result", "Spain Result",
           "Task Overall (DO NOT EDIT)"]

FILENAME = "1_0109_0409-O2C DTC_GBS Operations_checklist.xlsx"


def _xlsx_bytes(retail_fr="OK", days=("2026-09-01",)) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tabs = [f"{stream}_{day}" for day in days
            for stream in ("Retail", "eCom")]
    for title in tabs:
        ws = wb.create_sheet(title)
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(6, col, header)
        ws.append([])  # row 7 filled below
        # results live at detail level when a task has details (like the
        # real workbook); retail_fr lands in the detail row's France cell
        fr = retail_fr if title.startswith("Retail_") else "OK"
        for col, v in enumerate(["1", "Settlement", "Monitor files", "Daily",
                                 "Yes", None, "Adyen", None, None, None,
                                 None, "Pending"], start=1):
            ws.cell(7, col, v)
        ws.cell(8, 3, "↳ Detail check")
        ws.cell(8, 4, "Daily")
        ws.cell(8, 5, "Yes")
        ws.cell(8, 6, "France")
        ws.cell(8, 7, "Adyen for cards")
        ws.cell(8, 8, fr)
        ws.row_dimensions[8].outline_level = 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_path = tmp_path / "sustain.db"
    database.init_db(db_path).close()
    db_sustain.init_schema(db_path)
    monkeypatch.setattr(web_sustain, "_get_conn",
                        lambda: database.get_connection(db_path))
    monkeypatch.setitem(web_sustain._cfg, "database_path", str(db_path))
    monkeypatch.setattr(web_sustain, "_UPLOAD_FOLDER", tmp_path / "uploads")
    return app.test_client()


def _upload(client, data=None, filename=FILENAME):
    data = data if data is not None else _xlsx_bytes()
    return client.post("/sustain/upload", data={
        "file": (io.BytesIO(data), filename)})


def test_home_shows_empty_state_before_import(client):
    html = client.get("/sustain/").get_data(as_text=True)
    assert "Sustainphase Monitoring" in html
    assert "No checklist imported yet" in html


def test_upload_imports_and_reports_counts(client):
    resp = _upload(client)
    assert resp.status_code == 302 and "sustain_ok=1" in resp.headers["Location"]
    html = client.get(resp.headers["Location"]).get_data(as_text=True)
    assert "2 day tabs" in html and "2 tasks" in html

    conn = web_sustain._get_conn()
    try:
        assert db_sustain.task_count(conn) == 2
    finally:
        conn.close()


def test_upload_rejects_wrong_files(client):
    resp = _upload(client, filename="notes.txt")
    assert "sustain_ok=0" in resp.headers["Location"]
    # right extension, wrong workbook family (suffix mismatch)
    resp = _upload(client, filename="DTC_UAT_testtracking_ROE.xlsx")
    assert "sustain_ok=0" in resp.headers["Location"]
    resp = client.post("/sustain/upload", data={})
    assert "sustain_ok=0" in resp.headers["Location"]


def test_upload_accepts_browser_duplicate_name(client):
    """'… checklist (1).xlsx' double-download copies must still import."""
    resp = _upload(
        client,
        filename="1_0109_0409-O2C DTC_GBS Operations_checklist (1).xlsx")
    assert "sustain_ok=1" in resp.headers["Location"]


def test_upload_dated_copy_kept_in_uploads_folder(client, tmp_path):
    _upload(client)
    saved = list((tmp_path / "uploads").glob("sustain_*.xlsx"))
    assert len(saved) == 1


# ---------------------------------------------------------------------------
# Day report (build plan step 4)

def test_day_report_mirrors_excel_with_expandable_details(client):
    _upload(client)
    html = client.get("/sustain/day/2026-09-01/Retail").get_data(as_text=True)
    assert "Monitor files" in html and "Settlement" in html
    # stat cards: due 1, completed 1, pending 0, attention 0
    assert 'class="num">1</div>' in html
    # detail row present but collapsed, with its country + provider
    assert 'class="sustain-detail" data-parent=' in html and "hidden" in html
    assert "France" in html and "Adyen for cards" in html
    # recomputed OK renders as a green pill
    assert 'pill pill--green">OK' in html
    # stream toggle to the same day's eCom tab
    assert '/sustain/day/2026-09-01/eCom' in html


def test_day_report_empty_tab_shows_hint(client):
    html = client.get("/sustain/day/2026-09-09/Retail").get_data(as_text=True)
    assert "No data for Retail 2026-09-09" in html


def test_home_links_days_to_day_report(client):
    _upload(client)
    html = client.get("/sustain/").get_data(as_text=True)
    assert 'href="/sustain/day/2026-09-01/Retail"' in html
    assert 'href="/sustain/day/2026-09-01/eCom"' in html


# ---------------------------------------------------------------------------
# Management summary (build plan step 5)

def test_summary_defaults_to_latest_day_and_lists_attention(client):
    _upload(client, data=_xlsx_bytes(
        retail_fr="acct 4711 unclear", days=("2026-09-01", "2026-09-02")))
    html = client.get("/sustain/summary").get_data(as_text=True)
    # latest day selected, both streams present
    assert "Retail — 2026-09-02" in html and "eCom — 2026-09-02" in html
    # the verbatim note is the discussion point
    assert "acct 4711 unclear" in html
    # eCom is all clear
    assert "Nothing needs attention" in html
    # trend has a row per tab (4) linking to the day reports
    assert html.count('href="/sustain/day/') >= 4
    # the recurring Retail note shows up as a repeat offender
    assert "Repeat offenders" in html
    # once in the selected day's attention list + once (deduped) in the
    # repeat-offenders table
    assert html.count("acct 4711 unclear") == 2


def test_summary_explicit_day_and_all_clear(client):
    _upload(client)
    html = client.get("/sustain/summary/2026-09-01").get_data(as_text=True)
    assert "Retail — 2026-09-01" in html
    assert "No recurring attention items" in html


def test_summary_empty_state(client):
    html = client.get("/sustain/summary").get_data(as_text=True)
    assert "No checklist imported yet" in html


def test_home_links_to_summary(client):
    html = client.get("/sustain/").get_data(as_text=True)
    assert 'href="/sustain/summary"' in html


def test_upload_reports_error_when_no_day_tabs(client):
    wb = openpyxl.Workbook()
    wb.active.title = "Instructions"
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, data=buf.getvalue())
    assert "sustain_ok=0" in resp.headers["Location"]
    html = client.get(resp.headers["Location"]).get_data(as_text=True)
    assert "day tabs" in html
