"""ECOM status report (buckets same as Retail [USER 2026-07-09]).

What must hold:
- status counts feed the SAME bucket definitions (one config); today's
  "Not Ready" rows land under known exclusions, never silently vanish
- impacted defects: ECOM channel only, passed family excluded, MB/Sales
  via the dtco2c flag (same rules as Retail)
- save-to-Excel appends a dated row to the ECOM sheet of the log workbook
- the email path renders the report as a standalone attachment
"""
import openpyxl
import pytest

from app import database, emailer
from app.db import ecom as db_ecom
import app.web_ecom as web_ecom
from app.reporter import compute_retail_report, load_status_mappings
from app.web import app


def _ecom_row(jira, status="Not Ready", defect_ref=""):
    return {"jira_id": jira, "status": status, "assigned_to": "",
            "country": "Germany (DE)", "testcase_scenario": "CANCELLATION",
            "test_case_id": jira, "testcase_name": f"{jira}_case",
            "description_change": "", "execution_started": "",
            "order_number": "", "old_order_numbers": "",
            "defect_id_ref": defect_ref, "s4_sales_order": "",
            "s4_billing_documents": "", "s4_journal_invoice_entry": "",
            "delivery_note": "", "reason_for_pass_with_reservation": "",
            "comment": "", "excel_row": 2, "_skip_reason": ""}


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_path = tmp_path / "ecom_report.db"
    database.init_db(db_path).close()
    db_ecom.init_schema(db_path)
    monkeypatch.setattr(web_ecom, "_db_path", db_path)
    conn = database.get_connection(db_path)
    try:
        db_ecom.upsert_ecom_rows(conn, [
            _ecom_row("S4ECOM-1"), _ecom_row("S4ECOM-2", status="In Progress"),
            _ecom_row("S4ECOM-3", status="Passed", defect_ref="DEF-9"),
            _ecom_row("S4ECOM-4", status="Blocked DTC", defect_ref="DEF-9"),
        ], "2026-07-09")
        conn.execute("INSERT INTO defects (defect_id, channel, solman_status)"
                     " VALUES ('DEF-9', 'ECOM', 'In Progress')")
        conn.execute("INSERT INTO defects (defect_id, channel, solman_status)"
                     " VALUES ('DEF-R', 'Retail', 'In Progress')")
        conn.commit()
    finally:
        conn.close()
    c = app.test_client()
    c.db_path = db_path
    return c


def test_status_counts_feed_shared_buckets(client):
    conn = database.get_connection(client.db_path)
    try:
        counts = db_ecom.get_ecom_status_counts(conn)
    finally:
        conn.close()
    assert counts == {"Not Ready": 1, "In Progress": 1, "Passed": 1,
                      "Blocked DTC": 1}
    report = compute_retail_report(counts, load_status_mappings())
    assert report["buckets"]["passed_with_dtc"] == 1
    assert report["buckets"]["blocked"] == 1
    # "Not Ready" is a known exclusion — visible, not silently dropped
    assert report["diagnostics"]["unmapped_known"] == {"Not Ready": 1}


def test_impacted_defects_ecom_channel_and_passed_excluded(client):
    conn = database.get_connection(client.db_path)
    try:
        rows = db_ecom.get_ecom_defects_impacted(
            conn, ["Passed", "conditionally passed"])
    finally:
        conn.close()
    assert [r["defect_id"] for r in rows] == ["DEF-9"]     # Retail defect excluded
    assert rows[0]["impacted_tc_count"] == 1               # Blocked DTC counts
    assert rows[0]["passed_tc_count"] == 1                 # Passed shown muted only


def test_report_page_renders(client):
    html = client.get("/ecom/report").get_data(as_text=True)
    assert "ECOM Status Report" in html
    assert "Not Ready: 1" in html                          # known-exclusion pill
    assert "DEF-9" in html
    assert "same bucket definitions as the Retail report" in html


def test_save_excel_appends_ecom_sheet(client, tmp_path, monkeypatch):
    log = tmp_path / "log.xlsx"
    monkeypatch.setitem(web_ecom._cfg, "retail_report_xlsx", str(log))
    d = client.post("/ecom/report/save-excel", data={"date": "2026-07-09"}).get_json()
    assert d["ok"], d.get("error")
    wb = openpyxl.load_workbook(log)
    ws = wb["ECOM"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(2, 1).value == "2026-07-09"
    assert ws.cell(2, 5).value == 1                        # Passed with DTC
    wb.close()


def test_email_attachment_and_choice(client):
    assert ("ecom", "ECOM Status Report") in emailer.REPORT_CHOICES
    conn = database.get_connection(client.db_path)
    try:
        atts = emailer.gather_attachments(conn, {}, app, ["ecom"], "2026-07-09")
    finally:
        conn.close()
    assert len(atts) == 1
    name, html = atts[0]
    assert name == "ecom_report_2026-07-09.html"
    assert "ECOM Status Report" in html
    assert "<script" not in html                           # standalone: scripts stripped
