"""Manual Test Cases pages + reports (2026-08-05).

What must hold:
- both streams' list and report pages render; unknown stream 404s
- status counts feed the SAME bucket definitions as Retail (one config)
- defects rule [USER 2026-08-05]: a defect appears ONLY if referenced in
  the tab's defect_id_ref AND its channel matches the vertical; referenced
  defects of another/blank channel surface in the off-channel data check,
  never silently vanish
- save-to-Excel appends a dated row to the vertical's own sheet of the
  shared report log workbook
"""
import openpyxl
import pytest

from app import database
from app.db import manual_tests as db_manual
import app.web_manual_tests as web_manual
from app.web import app


def _row(tc, country, status="Ready for Validation", defect_ref="", vertical="manual_retail"):
    base = {f: "" for f in db_manual.FIELDS[vertical]}
    base.update({"test_case_id": tc, "country": country, "status": status,
                 "testcase_name": f"{tc}_case",
                 "testcase_scenario": "Settlement File Validation",
                 "defect_id_ref": defect_ref,
                 "excel_row": 2, "_skip_reason": ""})
    return base


@pytest.fixture()
def client(tmp_path, monkeypatch):
    db_path = tmp_path / "manual_pages.db"
    database.init_db(db_path).close()
    db_manual.init_schema(db_path)
    monkeypatch.setattr(web_manual, "_db_path", db_path)
    conn = database.get_connection(db_path)
    try:
        db_manual.upsert_manual_rows(conn, "manual_retail", [
            _row("CDI0000MU01", "Germany"),
            _row("CDI0000MU01", "Poland", status="Passed", defect_ref="DEF-R1"),
            _row("CDI0000MU02", "Norway", status="Blocked DTC", defect_ref="DEF-R1"),
            _row("CDI0000MU03", "Greece", defect_ref="DEF-E1"),   # off-channel ref
        ], "2026-08-05")
        db_manual.upsert_manual_rows(conn, "manual_ecom", [
            _row("CDI0000MU01", "Austria", vertical="manual_ecom"),
        ], "2026-08-05")
        conn.execute("INSERT INTO defects (defect_id, channel, solman_status)"
                     " VALUES ('DEF-R1', 'Retail', 'In Progress')")
        conn.execute("INSERT INTO defects (defect_id, channel, solman_status)"
                     " VALUES ('DEF-E1', 'ecom', 'In Progress')")
        conn.execute("INSERT INTO defects (defect_id, channel, solman_status)"
                     " VALUES ('DEF-R2', 'Retail', 'In Progress')")  # NOT referenced
        conn.commit()
    finally:
        conn.close()
    c = app.test_client()
    c.db_path = db_path
    return c


def test_list_pages_render_and_unknown_stream_404s(client):
    html = client.get("/manual/retail").get_data(as_text=True)
    assert "Manual Test Cases Retail" in html
    assert "CDI0000MU02" in html
    html = client.get("/manual/ecom").get_data(as_text=True)
    assert "Manual Test Cases ECOM" in html
    assert "CDI0000MU01" in html
    assert client.get("/manual/nope").status_code == 404
    assert client.get("/manual/nope/report").status_code == 404


def test_list_filters(client):
    # (country names stay visible in the filter dropdown — assert on TC ids)
    html = client.get("/manual/retail?status=Passed").get_data(as_text=True)
    assert "CDI0000MU01" in html and "CDI0000MU02" not in html
    html = client.get("/manual/retail?search=MU02").get_data(as_text=True)
    assert "CDI0000MU02" in html and "CDI0000MU03" not in html


def test_settlement_filter_splits_by_test_case_name(client, tmp_path):
    """[USER 2026-08-06] "Test type" filter: name contains "settlement
    file" (case-insensitive) vs. not — independent of the scenario column."""
    conn = database.get_connection(client.db_path)
    try:
        rows = [
            {**_row("CDI0000MZ01", "Sweden"),
             "testcase_name": "CDI0000MZ01_Settlement File Validation"},
            {**_row("CDI0000MZ02", "Sweden"),
             "testcase_name": "CDI0000MZ02_Audit of Store Cash reconciliation"},
        ]
        db_manual.upsert_manual_rows(conn, "manual_retail", rows, "2026-08-06")
        conn.commit()
    finally:
        conn.close()

    html = client.get("/manual/retail?settlement=settlement").get_data(as_text=True)
    assert "CDI0000MZ01" in html and "CDI0000MZ02" not in html

    html = client.get("/manual/retail?settlement=other").get_data(as_text=True)
    assert "CDI0000MZ02" in html and "CDI0000MZ01" not in html

    html = client.get("/manual/retail?settlement=bogus").get_data(as_text=True)
    assert "CDI0000MZ01" in html and "CDI0000MZ02" in html  # unknown value -> no filter


def test_defects_rule_referenced_and_channel_match(client):
    conn = database.get_connection(client.db_path)
    try:
        rows = db_manual.get_manual_defects_impacted(
            conn, "manual_retail", ["Passed", "conditionally passed"])
        off = db_manual.get_manual_offchannel_defect_refs(conn, "manual_retail")
    finally:
        conn.close()
    # DEF-R1: referenced + Retail channel -> in. DEF-R2: right channel but
    # NOT referenced -> out. DEF-E1: referenced but ecom channel -> off-channel.
    assert [r["defect_id"] for r in rows] == ["DEF-R1"]
    assert rows[0]["impacted_tc_count"] == 1     # Blocked DTC counts
    assert rows[0]["passed_tc_count"] == 1       # Passed shown muted only
    assert [d["defect_id"] for d in off] == ["DEF-E1"]
    assert off[0]["ref_count"] == 1


def test_report_pages_render_with_defects_and_offchannel_warning(client):
    html = client.get("/manual/retail/report").get_data(as_text=True)
    assert "Manual Test Cases Retail Report" in html
    assert "DEF-R1" in html
    assert "DEF-R2" not in html                       # not referenced
    assert "DEF-E1" in html                           # off-channel warning box
    assert "outside the Retail channel" in html
    assert "same bucket definitions as the Retail report" in html

    html = client.get("/manual/ecom/report").get_data(as_text=True)
    assert "Manual Test Cases ECOM Report" in html
    # no ecom-channel defect referenced on the ECOM tab -> empty-state text
    assert "the Defect ID column of the tab is empty so far" in html


def test_save_excel_appends_own_sheets(client, tmp_path, monkeypatch):
    log = tmp_path / "log.xlsx"
    monkeypatch.setitem(web_manual._cfg, "retail_report_xlsx", str(log))
    d = client.post("/manual/retail/report/save-excel",
                    data={"date": "2026-08-05"}).get_json()
    assert d["ok"], d.get("error")
    d = client.post("/manual/ecom/report/save-excel",
                    data={"date": "2026-08-05"}).get_json()
    assert d["ok"], d.get("error")
    wb = openpyxl.load_workbook(log)
    assert "Manual Retail" in wb.sheetnames and "Manual ECOM" in wb.sheetnames
    ws = wb["Manual Retail"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(2, 1).value == "2026-08-05"
    assert ws.cell(2, 5).value == 1               # Passed with DTC (the Passed row)
    wb.close()


def test_email_attachments_and_choices(client):
    from app import emailer
    assert ("manual_retail", "Manual Test Cases Retail Report") in emailer.REPORT_CHOICES
    assert ("manual_ecom", "Manual Test Cases ECOM Report") in emailer.REPORT_CHOICES
    conn = database.get_connection(client.db_path)
    try:
        atts = emailer.gather_attachments(
            conn, {}, app, ["manual_retail", "manual_ecom"], "2026-08-05")
    finally:
        conn.close()
    assert [name for name, _ in atts] == [
        "manual_retail_report_2026-08-05.html",
        "manual_ecom_report_2026-08-05.html",
    ]
    assert "Manual Test Cases Retail Report" in atts[0][1]
    assert "<script" not in atts[0][1]             # standalone: scripts stripped


def test_download_is_standalone(client):
    resp = client.get("/manual/retail/report/download")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "Manual Test Cases Retail Report" in html
    assert "<script" not in html                  # standalone: scripts stripped
