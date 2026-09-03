"""Sustainphase Issues importer (rewritten 2026-09-03 [USER]): the Go-Live
defect tracker workbook — three tabs mapped by header name."""
import io
from datetime import datetime

import openpyxl
import pytest

from app.sustain_issues_importer import ParseError, parse_go_live_tracker

INCIDENT_HEADERS = ["Incident Number", "Date", "Requestor", "Title", "Status",
                    "Assigned To", "Latest comment/action"]
SOLUTION_HEADERS = ["Owner", "Interface", "Msg", "Text", "External Reference",
                    "INC reference, if any", "Reason", "Solution", "Status", None]
INTERFACES = [("/FINAC", "JRNLENTRIN"), ("/RFMPI", "SALES"), ("/SDSLS", "SO_BULK_I"),
              ("ZSD_I", "SO_MULTI")]


def workbook_bytes(incidents=None, solutions=None, interfaces=INTERFACES) -> bytes:
    """Mirrors Marina's file: ASPEN Incidents + Issue Solution tracker with
    headers in row 1; the Total tab with group titles in row 2 and the
    list header in row 3."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASPEN Incidents"
    ws.append(INCIDENT_HEADERS)
    for r in (incidents or []):
        ws.append(r)
    ws2 = wb.create_sheet("Issue Solution tracker")
    ws2.append(SOLUTION_HEADERS)
    for r in (solutions or []):
        ws2.append(r)
    ws3 = wb.create_sheet("Total")
    ws3.append([])
    ws3.append([None, "Interfaces", None, None, None, None, None, "Total Issue #", None])
    ws3.append([None, "Namespace", "Interface", "Version", "Name", "Variant in /aif/err",
                "Index tables", None, None])
    for ns, itf in interfaces:
        ws3.append([None, ns, itf, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write(tmp_path, data, name="Go-Live defect tracker (1).xlsx"):
    p = tmp_path / name
    p.write_bytes(data)
    return p


def test_parse_reads_all_three_tabs_and_skips_rows_without_incident_number(tmp_path):
    path = _write(tmp_path, workbook_bytes(
        incidents=[["INC001", datetime(2026, 9, 1), "Anna", "Invoice missing", "Open",
                    "Tom", "Checked SAP,\nwaiting for AIF"],
                   [None, datetime(2026, 9, 2), "Ben", "no number", "Open", None, None],
                   ["INC002", None, None, "Second", "Closed", None, ""]],
        solutions=[["Tom", "SALES", "E123", "Order rejected", "EXT-1", "INC001",
                    "Mapping", "Fix mapping", "Open", None],
                   ["Anna", "n/a", None, "Unknown", None, None, "Data", "Reload", "Closed", None]]))
    parsed = parse_go_live_tracker(path)
    inc = parsed["incidents"]
    assert [i["incident_number"] for i in inc] == ["INC001", "INC002"]
    assert parsed["skipped_incidents"] == 1
    assert inc[0]["date"] == "2026-09-01"
    assert inc[0]["latest_comment"] == "Checked SAP,\nwaiting for AIF"   # newlines kept
    assert inc[0]["assigned_to"] == "Tom" and inc[0]["excel_row"] == 2
    assert inc[1]["latest_comment"] is None
    sol = parsed["solutions"]
    assert [s["interface"] for s in sol] == ["SALES", "n/a"]
    assert sol[0]["inc_reference"] == "INC001" and sol[0]["external_reference"] == "EXT-1"
    assert sol[0]["reason"] == "Mapping" and sol[1]["status"] == "Closed"
    itf = parsed["interfaces"]
    assert [(i["namespace"], i["interface"]) for i in itf] == INTERFACES
    assert itf[0]["excel_row"] == 4


def test_parse_empty_template_is_fine(tmp_path):
    parsed = parse_go_live_tracker(_write(tmp_path, workbook_bytes()))
    assert parsed["incidents"] == [] and parsed["solutions"] == []
    assert len(parsed["interfaces"]) == len(INTERFACES)


def test_parse_missing_tab_is_loud(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Something else"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ParseError, match="ASPEN Incidents"):
        parse_go_live_tracker(_write(tmp_path, buf.getvalue()))
