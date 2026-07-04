"""Tests for the Retail Requirements Tracker importer (step 1).

Covers the parsing/resolution rules that would fail silently if wrong:
- test id split at the first underscore, whitespace (incl. newlines) collapsed
- '18' -> all-countries sentinel, plain int -> required_dtc,
  '1 or 5' -> stored raw, flagged as needing a decision
- return tab: countable id = return column when present (sale id kept as
  informational ref), else the sale column
- duplicate (name, test) rows across tabs are skipped, not double-imported
- unresolved names are flagged, never silently dropped
- country headers seed the country list with UK/Belguim fixes applied
"""
from pathlib import Path

import openpyxl
import pytest

from app.retail_tracker_importer import (
    collect_targets,
    parse_comment_targets,
    parse_name_targets,
    parse_tracking_workbook,
    resolve_requirements,
)

COUNTRIES = ["UK", "Germany", "Belguim"]


def _make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Tracking Sales"
    ws.append(["Scenarios", "Test ID Sale", "DTC required", "comment", "number"] + COUNTRIES)
    ws.append(["1. Retail Sale"])
    ws.append(["a. normal"])
    # newline inside the test id cell, like the real file
    ws.append(["plain vanilla sale", "GKP0029MU01_\nSale MLMQ with Barcode - Cash", 18, "", 2, "yes", "yes"])
    ws.append(["Book", "GKP0005MU01_Sale of the Book - Cash", 3, "", 1, "yes"])
    # requirement rows that carry their own sub-heading letter must NOT
    # inherit the previous sub-heading (real bug found 2026-07-04)
    ws.append(["c. Sale of giftcard"])
    ws.append(["giftcard sale", "GKP1012MU01_Gift Card Sale", 6, "", 0])
    ws.append(["d. sales cancellation", "GKP0013MU01_Verify Cancel transaction", 1, "", 0])
    ws.append(["7. Discounts"])
    ws.append(["Employee Discount: CS", "GKP1005MU01_Employee Discount CS", "1 or 5", "", 0])
    # section 8 repeats the Payment tab -> the PAYMENT tab is canonical
    ws.append(["8. Payment Methods General"])
    ws.append(["CASH", "GKP2003MU01_Sale w.Cash", 18, "", 1, "yes"])

    ws = wb.create_sheet("Tracking Return")
    ws.append(["Scenarios", "Test ID Sale", "Test ID Return", "DTC required", "comment", "number"] + COUNTRIES)
    ws.append(["1. Retail Sale"])
    # both ids -> return id counts, sale id is informational
    ws.append(["plain vanilla sale", "GKP0029MU01_Sale MLMQ with Barcode - Cash",
               "GKP0030MU01_Return Sale MLMQ with Barcode - Cash", 18, "ALL countries", 1, "yes"])
    # duplicate of the Sales tab row (same name + same test) -> must be skipped
    ws.append(["Book", "GKP0005MU01_Sale of the Book - Cash", None, 3, "", 0])
    # only the sale column holds an id -> that IS the countable test (exchange case)
    ws.append(["even exchange", "GKP1024MU01_Exchange Even", None, 18, "", 0])
    # unresolvable name
    ws.append(["suspend", None, "GKP0011MU01_Verify Suspend transaction functionality", 1, "", 0])

    ws = wb.create_sheet("Payment methods")
    ws.append(["", "Solman ID", "how many DTC required?", "comment", "number"] + COUNTRIES)
    ws.append(["Sales"])
    ws.append(["CASH", "GKP2003MU01_Sale w.Cash", 18, "", 1, "yes"])   # dup of sales section 8
    ws.append(["Return"])
    ws.append(["Blind return", "GKP1015MU01_Blind Return", 1, "", 0])  # unique -> folds into return

    ws = wb.create_sheet("Country Specific PaymentM")
    ws.append(["COUNTRY", "", "Card", "Voucher", "Tested Sale Card",
               "Tested Sale Voucher", "Tested Return Card", "Tested Return Voucher"])
    ws.append(["", "", "", "", "GKP2019MU01_Sale of Multi Tender",
               "GKP2021MU01_Sale of Payment Voucher",
               "GKP2020MU01_Return of Multi Tender",
               "GKP2022MU01_Return of Payment Voucher"])
    ws.append(["DE", "53340 AMEX", "X"])                                   # card
    ws.append(["DE", "130913 vouchers", "checking if applicable", "X"])    # voucher + note
    ws.append(["GB", "53430 Offline Transactions"])                        # no mark -> unknown
    ws.append(["AU", "53341 VISA", "x"])                                   # lowercase x, AU=Austria

    wb.save(path)


LOOKUP = {
    "sale mlmq with barcode - cash": "GKPMU000025",
    "return sale mlmq with barcode - cash": "GKPMU000026",
    "sale of the book - cash": "GKPMU000005",
    "exchange even": "GKPMU000053",
    "employee discount cs": "GKPMU000030",
    "sale w.cash": "GKPMU000040",
    "blind return": "GKPMU000048",
    "gift card sale": "GKPMU000044",
    "verify cancel transaction": "GKPMU000011",
}


@pytest.fixture()
def parsed(tmp_path):
    xlsx = tmp_path / "tracking_TEST.xlsx"
    _make_workbook(xlsx)
    return parse_tracking_workbook(xlsx)


def _by_name(parsed, area, name):
    return next(r for r in parsed["requirements"] if r["area"] == area and r["name"] == name)


def test_parses_expected_requirement_counts(parsed):
    # no payment area: CASH lives in sales (section 8), Blind return folds into return
    assert parsed["per_tab"] == {"sales": 6, "return": 4}


def test_test_id_split_and_whitespace_normalised(parsed):
    r = _by_name(parsed, "sales", "plain vanilla sale")
    assert r["excel_test_ref"] == "GKP0029MU01"
    assert r["test_name"] == "sale mlmq with barcode - cash"  # newline collapsed


def test_required_18_becomes_all_countries_sentinel(parsed):
    r = _by_name(parsed, "sales", "plain vanilla sale")
    assert r["all_countries"] == 1
    assert r["required_dtc"] is None


def test_required_plain_int_stored_as_count(parsed):
    r = _by_name(parsed, "sales", "Book")
    assert r["all_countries"] == 0
    assert r["required_dtc"] == 3


def test_required_1_or_5_counts_as_5(parsed):
    # user decision 2026-07-04: the Excel's '1 or 5' means required = 5
    r = _by_name(parsed, "sales", "Employee Discount: CS")
    assert r["required_raw"] == "1 or 5"  # provenance kept
    assert r["required_dtc"] == 5
    assert r["all_countries"] == 0


def test_return_tab_counts_return_id_and_keeps_sale_ref(parsed):
    r = _by_name(parsed, "return", "plain vanilla sale")
    assert r["excel_test_ref"] == "GKP0030MU01"          # the RETURN test counts
    assert r["test_name"] == "return sale mlmq with barcode - cash"
    assert r["sale_test_ref"].startswith("GKP0029MU01")  # informational only


def test_return_tab_sale_column_counts_when_return_empty(parsed):
    r = _by_name(parsed, "return", "even exchange")
    assert r["excel_test_ref"] == "GKP1024MU01"
    assert r["sale_test_ref"] is None


def test_duplicate_rows_skipped_and_reported(parsed):
    return_names = [r["name"] for r in parsed["requirements"] if r["area"] == "return"]
    assert "Book" not in return_names  # duplicate of the Sales tab row
    dups = {d["name"]: d for d in parsed["skipped_dups"]}
    assert "sales" in dups["Book"]["dup_of"]


def test_payment_tab_folds_into_sales_and_return(parsed):
    # user decision 2026-07-04: no payment area. The payment tab's duplicates
    # are skipped (sales/return tabs win); its unique rows fold into
    # sales/return based on their section heading.
    areas = {r["area"] for r in parsed["requirements"]}
    assert "payment_general" not in areas
    dups = {d["name"]: d for d in parsed["skipped_dups"]}
    assert "sales" in dups["CASH"]["dup_of"]          # payment CASH = dup of sales section 8
    blind = _by_name(parsed, "return", "Blind return")
    assert blind["scenario_label"] == "Payment methods — Return"
    assert blind["excel_row"] > 1000                  # own key space, no collision


def test_scenario_headings_flatten(parsed):
    assert _by_name(parsed, "sales", "plain vanilla sale")["scenario_label"] == "1. Retail Sale — a. normal"
    assert _by_name(parsed, "sales", "Employee Discount: CS")["scenario_label"] == "7. Discounts"
    assert _by_name(parsed, "sales", "CASH")["scenario_label"] == "8. Payment Methods General"


def test_lettered_requirement_row_replaces_sub_heading(parsed):
    # "d. sales cancellation" must NOT inherit "c. Sale of giftcard"
    assert _by_name(parsed, "sales", "giftcard sale")["scenario_label"] \
        == "1. Retail Sale — c. Sale of giftcard"
    r = _by_name(parsed, "sales", "sales cancellation")  # letter prefix stripped
    assert r["scenario_label"] == "1. Retail Sale — d. sales cancellation"


def test_country_headers_seed_with_name_fixes(parsed):
    assert parsed["countries"] == [
        ("United Kingdom", "UK"), ("Germany", "Germany"), ("Belgium", "Belguim")]


def test_tab4_rows_parsed_with_categories_and_country_names(parsed):
    rows = {(r["country"], r["method_name"]): r for r in parsed["cpm_rows"]}
    assert len(rows) == 3
    assert rows[("Germany", "53340 AMEX")]["category"] == "card"
    v = rows[("Germany", "130913 vouchers")]
    assert v["category"] == "voucher"
    assert v["comment"] == "checking if applicable"      # odd text preserved
    assert rows[("Austria", "53341 VISA")]["category"] == "card"  # lowercase x, AU=Austria
    # 'Offline Transactions' rows are excluded — covered by the OFFLINE
    # requirements in Sales/Return (user decision 2026-07-04)
    assert ("United Kingdom", "53430 Offline Transactions") not in rows


def test_tab4_fixed_tests_parsed_in_kind_order(parsed):
    tests = {t["test_kind"]: t for t in parsed["tab4_tests"]}
    assert tests["sale_card"]["excel_test_ref"] == "GKP2019MU01"
    assert tests["sale_card"]["test_name"] == "sale of multi tender"
    assert tests["sale_voucher"]["excel_test_ref"] == "GKP2021MU01"
    assert tests["return_card"]["excel_test_ref"] == "GKP2020MU01"
    assert tests["return_voucher"]["excel_test_ref"] == "GKP2022MU01"


def test_comment_country_lists_become_named_targets():
    # user decision 2026-07-04: "UK, IE" means THOSE specific countries
    assert parse_comment_targets("UK, IE") == ["United Kingdom", "Ireland"]
    # conservative: anything that isn't purely a country list is NOT targets
    assert parse_comment_targets("CORE (6)") == []
    assert parse_comment_targets("ALL countries") == []
    assert parse_comment_targets("Special case portugal") == []
    assert parse_comment_targets(None) == []


def test_special_case_in_name_becomes_target():
    # user decision 2026-07-04: "B2B invoice (special case PL)" must pass in Poland
    assert parse_name_targets("B2B invoice (special case PL)") == ["Poland"]
    assert parse_name_targets("B2B invoice") == []
    assert parse_name_targets("Special Case ch something") == ["Switzerland"]
    assert parse_name_targets(None) == []


def test_targets_propagate_across_areas_by_name():
    # user decision 2026-07-04: Kids articles "UK, IE" applies to sale AND return
    recs = [
        {"area": "sales",  "excel_row": 12, "name": "Kids articles", "comment": None},
        {"area": "return", "excel_row": 12, "name": "Kids articles", "comment": "UK, IE"},
        {"area": "sales",  "excel_row": 29, "name": "B2B invoice (special case PL)", "comment": None},
        {"area": "sales",  "excel_row": 30, "name": "B2B invoice", "comment": None},
    ]
    targets = collect_targets(recs)
    assert targets[("sales", 12)] == ["Ireland", "United Kingdom"]   # propagated
    assert targets[("return", 12)] == ["Ireland", "United Kingdom"]
    assert targets[("sales", 29)] == ["Poland"]                      # from the name
    assert ("sales", 30) not in targets                              # plain line: any-1


def test_resolution_fills_ids_and_flags_unresolved(parsed):
    res = resolve_requirements(parsed["requirements"], LOOKUP)
    assert res["resolved"] == 9
    assert [r["name"] for r in res["unresolved"]] == ["suspend"]  # flagged, not dropped
    assert _by_name(parsed, "sales", "Book")["test_case_id"] == "GKPMU000005"
    assert _by_name(parsed, "return", "even exchange")["test_case_id"] == "GKPMU000053"
