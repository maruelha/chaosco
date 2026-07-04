"""Retail — list, detail, annotations, status report (HTML/Excel/PPT), diagnostics

Routes module (refactoring step 4) — registers on the shared app from
app.web_core; endpoint names and URLs are unchanged from the old monolith.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from flask import jsonify, redirect, render_template, request, send_from_directory, url_for
from werkzeug.utils import secure_filename

from app import database
from app.web_core import (app, _cfg, _get_conn, _not_found,
                          _UPLOAD_FOLDER, _IMAGE_EXTS, _ALLOWED_EXTS)

from app.ppt_retail import build_retail_ppt
from app.reporter import compute_retail_report, load_status_mappings

@app.route("/retail")
def retail_list():
    statuses      = request.args.getlist("status")
    assignees     = request.args.getlist("assigned_to")
    countries     = request.args.getlist("country")
    scenarios     = request.args.getlist("scenario")
    action_needed = request.args.get("action_needed", "")
    search_defect  = request.args.get("search_defect", "").strip() or None
    search_order   = request.args.get("search_order", "").strip() or None
    search_billing = request.args.get("search_billing", "").strip() or None

    conn = _get_conn()
    try:
        rows    = database.get_retail(
            conn,
            statuses=statuses or None,
            assignees=assignees or None,
            countries=countries or None,
            scenarios=scenarios or None,
            search_defect=search_defect,
            search_order=search_order,
            search_billing=search_billing,
            action_needed=action_needed or None,
        )
        options = database.get_retail_filter_options(conn)
    finally:
        conn.close()

    return render_template(
        "retail.html",
        rows=rows,
        options=options,
        statuses=statuses,
        assignees=assignees,
        countries=countries,
        scenarios=scenarios,
        action_needed=action_needed,
        search_defect=search_defect or "",
        search_order=search_order or "",
        search_billing=search_billing or "",
    )


@app.route("/retail/<int:retail_id>/annotation", methods=["POST"])
def retail_annotation_save(retail_id: int):
    next_step = request.form.get("next_step", "").strip() or None
    conn = _get_conn()
    try:
        existing        = database.get_retail_annotation(conn, retail_id)
        comment_history = existing["comment_history"] if existing else None
        action_needed   = existing["action_needed"] if existing else 0
        database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
        ann = database.get_retail_annotation(conn, retail_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "next_step": (ann["next_step"] or "") if ann else "",
    })


@app.route("/retail/<int:retail_id>/comment", methods=["POST"])
def retail_comment_save(retail_id: int):
    comment_history = request.form.get("comment_history", "").strip() or None
    conn = _get_conn()
    try:
        existing      = database.get_retail_annotation(conn, retail_id)
        next_step     = existing["next_step"] if existing else None
        action_needed = existing["action_needed"] if existing else 0
        database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
        ann = database.get_retail_annotation(conn, retail_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "comment_history": (ann["comment_history"] or "") if ann else "",
    })


def _get_retail_report():
    conn = _get_conn()
    try:
        status_counts = database.get_retail_status_counts(conn)
    finally:
        conn.close()
    mappings = load_status_mappings()
    return compute_retail_report(status_counts, mappings)


@app.route("/retail/report")
def retail_status_report():
    report = _get_retail_report()
    conn = _get_conn()
    try:
        blocked_defects = database.get_retail_defects_blocked(conn)
        report_comments = database.list_report_comments(conn, "retail")
    finally:
        conn.close()
    blocked_total   = sum(d["blocked_tc_count"] for d in blocked_defects)
    dtco2c_total    = sum(d["blocked_tc_count"] for d in blocked_defects if d["dtco2c"])
    sales_total     = sum(d["blocked_tc_count"] for d in blocked_defects if not d["dtco2c"])
    return render_template(
        "retail_report.html",
        report=report,
        today=date.today().isoformat(),
        blocked_defects=blocked_defects,
        blocked_defects_total=blocked_total,
        dtco2c_total=dtco2c_total,
        sales_total=sales_total,
        report_comments=report_comments,
        total_test_cases=_cfg.get("retail_total_test_cases", 646),
        missing_categories=_cfg.get("retail_missing_categories", []),
    )


_RETAIL_REPORT_HEADERS = [
    "Date",
    "Back with Sales",
    "With DTC",
    "In Progress with DTC",
    "Passed with DTC",
    "Incoming (Gatekeeper)",
    "Ready for validation",
    "In Progress",
    "In Clarification",
    "Blocked",
]


def _append_retail_report_to_excel(report: dict, today: str) -> str:
    """Append one data row to the retail report Excel log. Returns the file path."""
    xlsx_path = Path(_cfg.get("retail_report_xlsx", "output/retail_report_log.xlsx"))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default blank sheet openpyxl creates
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Retail" not in wb.sheetnames:
        ws = wb.create_sheet("Retail")
    else:
        ws = wb["Retail"]

    if ws.cell(1, 1).value is None:
        for col, header in enumerate(_RETAIL_REPORT_HEADERS, 1):
            ws.cell(row=1, column=col).value = header

    b = report["buckets"]
    next_row = ws.max_row + 1
    for col, val in enumerate([
        today,
        b["back_with_sales"],
        b["with_dtc"],
        b["in_progress_with_dtc"],
        b["passed_with_dtc"],
        b["incoming_gatekeeper"],
        b["ready_for_validation"],
        b["in_progress"],
        b["in_clarification"],
        b["blocked"],
    ], 1):
        ws.cell(row=next_row, column=col).value = val

    wb.save(xlsx_path)
    return str(xlsx_path)


@app.route("/retail/report/save-excel", methods=["POST"])
def retail_report_save_excel():
    try:
        report     = _get_retail_report()
        save_date  = request.form.get("date") or date.today().isoformat()
        path       = _append_retail_report_to_excel(report, save_date)
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)})
    return jsonify({"ok": True, "path": path, "date": save_date})


@app.route("/retail/report/download")
def retail_report_download():
    report = _get_retail_report()
    today  = date.today().isoformat()
    conn   = _get_conn()
    try:
        blocked_defects = database.get_retail_defects_blocked(conn)
        report_comments = database.list_report_comments(conn, "retail")
    finally:
        conn.close()
    blocked_total = sum(d["blocked_tc_count"] for d in blocked_defects)
    dtco2c_total  = sum(d["blocked_tc_count"] for d in blocked_defects if d["dtco2c"])
    sales_total   = sum(d["blocked_tc_count"] for d in blocked_defects if not d["dtco2c"])
    html = render_template(
        "retail_report_download.html", report=report, today=today,
        blocked_defects=blocked_defects,
        blocked_defects_total=blocked_total,
        dtco2c_total=dtco2c_total,
        sales_total=sales_total,
        report_comments=report_comments,
        total_test_cases=_cfg.get("retail_total_test_cases", 646),
        missing_categories=_cfg.get("retail_missing_categories", []),
    )
    return html, 200, {
        "Content-Type": "text/html; charset=utf-8",
        "Content-Disposition": f'attachment; filename="retail_report_{today}.html"',
    }


@app.route("/retail/report/diagnostics")
def retail_report_diagnostics():
    conn = _get_conn()
    try:
        status_counts   = database.get_retail_status_counts(conn)
        blocked_defects = database.get_retail_defects_blocked(conn)
    finally:
        conn.close()
    mappings = load_status_mappings()
    report   = compute_retail_report(status_counts, mappings)

    buckets_cfg    = mappings["buckets"]
    known_unmapped = set(mappings.get("known_unmapped", []))
    known_statuses = set(mappings.get("known_statuses", []))
    bucketed       = {s for bkt in buckets_cfg.values() for s in bkt["statuses"]}

    bucket_order = [
        "back_with_sales", "with_dtc", "in_progress_with_dtc", "passed_with_dtc",
        "incoming_gatekeeper", "ready_for_validation", "in_progress",
        "in_clarification", "blocked",
    ]
    grouped = []
    for key in bucket_order:
        bkt      = buckets_cfg[key]
        statuses = [{"status": s, "count": status_counts.get(s, 0)} for s in bkt["statuses"]]
        grouped.append({
            "bucket_key":   key,
            "bucket_label": bkt["label"],
            "bucket_total": sum(r["count"] for r in statuses),
            "statuses":     statuses,
        })

    unmapped_known, unknown = [], []
    for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
        if status in bucketed:
            continue
        label = status if status else "(blank)"
        if status in known_unmapped or status in known_statuses:
            unmapped_known.append({"status": label, "count": count})
        else:
            unknown.append({"status": label, "count": count})

    blocked_total = sum(d["blocked_tc_count"] for d in blocked_defects)
    dtco2c_total  = sum(d["blocked_tc_count"] for d in blocked_defects if d["dtco2c"])
    sales_total   = sum(d["blocked_tc_count"] for d in blocked_defects if not d["dtco2c"])

    return render_template(
        "retail_report_diagnostics.html",
        report=report,
        today=date.today().isoformat(),
        grouped=grouped,
        unmapped_known=unmapped_known,
        unknown=unknown,
        blocked_defects=blocked_defects,
        blocked_defects_total=blocked_total,
        dtco2c_total=dtco2c_total,
        sales_total=sales_total,
        total_test_cases=_cfg.get("retail_total_test_cases", 646),
    )


@app.route("/retail/report/ppt")
def retail_report_ppt():
    report = _get_retail_report()
    today  = date.today().isoformat()
    conn   = _get_conn()
    try:
        blocked_defects = database.get_retail_defects_blocked(conn)
    finally:
        conn.close()
    blocked_total = sum(d["blocked_tc_count"] for d in blocked_defects)
    dtco2c_total  = sum(d["blocked_tc_count"] for d in blocked_defects if d["dtco2c"])
    sales_total   = sum(d["blocked_tc_count"] for d in blocked_defects if not d["dtco2c"])
    pptx_bytes = build_retail_ppt(
        report=report,
        blocked_defects=blocked_defects,
        dtco2c_total=dtco2c_total,
        sales_total=sales_total,
        blocked_total=blocked_total,
        total_test_cases=_cfg.get("retail_total_test_cases", 646),
        today=today,
        missing_categories=_cfg.get("retail_missing_categories", []),
    )
    return pptx_bytes, 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "Content-Disposition": f'attachment; filename="retail_report_{today}.pptx"',
    }


# ---------------------------------------------------------------------------
# Report export â€” save HTML + PDF of both reports to disk for automation


@app.route("/retail/<int:retail_id>", methods=["GET", "POST"])
def retail_detail(retail_id: int):
    conn = _get_conn()
    try:
        row = database.get_retail_by_id(conn, retail_id)
        if row is None:
            conn.close()
            return render_template("404.html", defect_id=str(retail_id)), 404
        if request.method == "POST":
            next_step       = request.form.get("next_step", "").strip() or None
            comment_history = request.form.get("comment_history", "").strip() or None
            action_needed   = 1 if request.form.get("action_needed") == "1" else 0
            database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
            conn.close()
            return redirect(url_for("retail_detail", retail_id=retail_id, saved="1"))
        notes = database.list_notes(conn, "retail", retail_id)
        attachments_by_note = database.get_attachments_for_notes(
            conn, [n["id"] for n in notes]
        )
    finally:
        conn.close()
    saved        = request.args.get("saved") == "1"
    note_added   = request.args.get("note_added") == "1"
    note_saved   = request.args.get("note_saved") == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    added_to_meeting = request.args.get("added_to_meeting") == "1"
    return render_template(
        "retail_detail.html",
        row=row, notes=notes,
        attachments_by_note=attachments_by_note,
        saved=saved, note_added=note_added,
        note_saved=note_saved, note_deleted=note_deleted,
        added_to_meeting=added_to_meeting,
        meetings=database.MEETING_OPTIONS,
    )

