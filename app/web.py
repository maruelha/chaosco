"""Local web UI — run with:  python -m app.web

Opens http://127.0.0.1:5000 in the browser automatically.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, url_for
from werkzeug.utils import secure_filename

from app import database
from app.config_loader import load_config
from app.importer import run_import
from app.reporter import compute_retail_report, load_status_mappings

_HERE = Path(__file__).parent

app = Flask(
    __name__,
    template_folder=str(_HERE / "templates"),
    static_folder=str(_HERE / "static"),
)

_cfg = load_config()
_db_path = Path(_cfg["database_path"])
_UPLOAD_FOLDER = _HERE.parent / "data" / "uploads"
_ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}

# Create schema once at startup; routes use get_connection() after this.
database.init_db(_db_path).close()


def _get_conn():
    return database.get_connection(_db_path)


def _not_found(defect_id: str):
    return render_template("404.html", defect_id=defect_id), 404


def _render_note_add_form(defect_id, solman_name, return_to, *, heading="", error=None):
    label = defect_id + (f" — {solman_name}" if solman_name else "")
    return render_template(
        "note_form.html", mode="add",
        entity_label=label, list_label="Defects",
        list_url=url_for("defects_list"),
        detail_url=url_for("defect_detail", defect_id=defect_id),
        action_url=url_for("note_add", defect_id=defect_id),
        cancel_url=(url_for("defects_list") if return_to == "list"
                    else url_for("defect_detail", defect_id=defect_id)),
        return_to=return_to, heading=heading, error=error,
    )


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def dashboard():
    conn = _get_conn()
    try:
        inbox_count       = database.count_inbox_items(conn)
        open_enhancements = len(database.get_enhancements(conn))
        to_deliver        = database.count_encouragements_to_deliver(conn)
    finally:
        conn.close()
    return render_template("dashboard.html", inbox_count=inbox_count,
                           open_enhancements=open_enhancements,
                           to_deliver=to_deliver)


@app.route("/import", methods=["POST"])
def do_import():
    result = run_import(_cfg)
    return render_template("import_result.html", r=result)


@app.route("/defects")
def defects_list():
    search        = request.args.get("search", "").strip()
    channel       = request.args.get("channel", "")
    statuses      = request.args.getlist("status")
    action_needed = request.args.get("action_needed", "no")
    dtco2c        = request.args.get("dtco2c", "")
    daily         = request.args.get("daily", "")
    show_all      = request.args.get("show_all") == "1"
    note_added    = request.args.get("note_added") == "1"

    hidden = _cfg.get("defects_hidden_statuses", [])
    exclude = [] if (show_all or statuses) else hidden

    conn = _get_conn()
    try:
        defects = database.list_defects(
            conn,
            search=search or None,
            channel=channel or None,
            statuses=statuses or None,
            action_needed=action_needed or None,
            exclude_statuses=exclude or None,
            dtco2c=dtco2c or None,
            daily=daily or None,
        )
        options = database.get_filter_options(conn)
    finally:
        conn.close()

    return render_template(
        "defects.html",
        defects=defects,
        options=options,
        search=search,
        channel=channel,
        statuses=statuses,
        action_needed=action_needed,
        dtco2c=dtco2c,
        daily=daily,
        show_all=show_all,
        hidden=hidden,
        note_added=note_added,
    )


@app.route("/defects/<defect_id>", methods=["GET", "POST"])
def defect_detail(defect_id: str):
    conn = _get_conn()
    try:
        defect = database.get_defect(conn, defect_id)
        if defect is None:
            return _not_found(defect_id)

        if request.method == "POST":
            def _field(name: str) -> str | None:
                v = request.form.get(name, "").strip()
                return v or None

            database.upsert_defect_annotation(
                conn,
                defect_id,
                description=_field("description"),
                business_impact=_field("business_impact"),
                reach=_field("reach"),
                retest_needs=_field("retest_needs"),
                next_step=_field("next_step"),
                action_needed=bool(request.form.get("action_needed")),
                comments=_field("comments"),
                dtco2c=bool(request.form.get("dtco2c")),
                dtco2c_resp=_field("dtco2c_resp"),
                daily=bool(request.form.get("daily")),
            )
            return redirect(url_for("defect_detail", defect_id=defect_id, saved="1"))

        notes = database.list_notes(conn, "defect", defect_id)
        attachments_by_note = database.get_attachments_for_notes(
            conn, [n["id"] for n in notes]
        )
    finally:
        conn.close()

    saved = request.args.get("saved") == "1"
    note_added = request.args.get("note_added") == "1"
    note_saved = request.args.get("note_saved") == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    added_to_meeting = request.args.get("added_to_meeting") == "1"
    return render_template(
        "defect_detail.html",
        defect=defect,
        saved=saved,
        notes=notes,
        attachments_by_note=attachments_by_note,
        note_added=note_added,
        note_saved=note_saved,
        note_deleted=note_deleted,
        added_to_meeting=added_to_meeting,
        meetings=database.MEETING_OPTIONS,
    )


# ---------------------------------------------------------------------------
# Spillover routes
# ---------------------------------------------------------------------------

@app.route("/spillover")
def spillover_list():
    area        = request.args.getlist("area")
    type_       = request.args.getlist("type")
    status      = request.args.getlist("status")
    assigned_to = request.args.getlist("assigned_to")
    critical    = request.args.getlist("critical")
    show_all    = request.args.get("show_all") == "1"

    hidden = _cfg.get("spillover_hidden_statuses", [])
    exclude = [] if (show_all or status) else hidden

    conn = _get_conn()
    try:
        rows    = database.get_spillover(
            conn,
            statuses=status or None,
            areas=area or None,
            types=type_ or None,
            assignees=assigned_to or None,
            critical=critical or None,
            exclude_statuses=exclude or None,
        )
        options          = database.get_spillover_filter_options(conn)
        docs_s4_ids      = database.get_docs_s4_spillover_ids(conn)
        report_comments  = database.list_report_comments(conn, "spillover")
    finally:
        conn.close()

    return render_template(
        "spillover.html",
        rows=rows,
        options=options,
        area=area,
        type_=type_,
        status=status,
        assigned_to=assigned_to,
        critical=critical,
        show_all=show_all,
        hidden_statuses=hidden,
        docs_s4_ids=docs_s4_ids,
        report_comments=report_comments,
    )


@app.route("/spillover/<int:spillover_id>/annotation", methods=["POST"])
def spillover_annotation_save(spillover_id: int):
    importance          = request.form.get("importance_for_signoff", "").strip() or None
    next_step           = request.form.get("next_step", "").strip() or None
    comment_for_signoff = request.form.get("comment_for_signoff", "").strip() or None
    signoff_group       = request.form.get("signoff_group", "").strip() or None
    conn = _get_conn()
    try:
        existing        = database.get_spillover_annotation(conn, spillover_id)
        comment_history = existing["comment_history"] if existing else None
        critical        = existing["critical_for_signoff"] if existing else None
        database.upsert_spillover_annotation(
            conn, spillover_id, importance, next_step, comment_history, critical, comment_for_signoff, signoff_group)
        ann = database.get_spillover_annotation(conn, spillover_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "importance_for_signoff": ann.get("importance_for_signoff") or "",
        "next_step":              ann.get("next_step") or "",
        "comment_for_signoff":    ann.get("comment_for_signoff") or "",
        "signoff_group":          ann.get("signoff_group") or "",
    })


@app.route("/spillover/<int:spillover_id>/comment", methods=["POST"])
def spillover_comment_save(spillover_id: int):
    comment_history = request.form.get("comment_history", "").strip() or None
    conn = _get_conn()
    try:
        existing            = database.get_spillover_annotation(conn, spillover_id)
        importance          = existing["importance_for_signoff"] if existing else None
        next_step           = existing["next_step"] if existing else None
        critical            = existing["critical_for_signoff"] if existing else None
        comment_for_signoff = existing["comment_for_signoff"] if existing else None
        signoff_group       = existing["signoff_group"] if existing else None
        database.upsert_spillover_annotation(
            conn, spillover_id, importance, next_step, comment_history, critical, comment_for_signoff, signoff_group)
        ann = database.get_spillover_annotation(conn, spillover_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "comment_history": ann["comment_history"] or "",
    })


@app.route("/spillover/<int:spillover_id>/critical", methods=["POST"])
def spillover_critical_save(spillover_id: int):
    critical = request.form.get("critical_for_signoff", "").strip() or None
    conn = _get_conn()
    try:
        existing            = database.get_spillover_annotation(conn, spillover_id)
        importance          = existing["importance_for_signoff"] if existing else None
        next_step           = existing["next_step"] if existing else None
        comment_history     = existing["comment_history"] if existing else None
        comment_for_signoff = existing["comment_for_signoff"] if existing else None
        signoff_group       = existing["signoff_group"] if existing else None
        database.upsert_spillover_annotation(
            conn, spillover_id, importance, next_step, comment_history, critical, comment_for_signoff, signoff_group)
        ann = database.get_spillover_annotation(conn, spillover_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "critical_for_signoff": ann["critical_for_signoff"] or "",
    })


# ---------------------------------------------------------------------------
# Spillover detail + note routes
# ---------------------------------------------------------------------------

@app.route("/spillover/<int:spillover_id>")
def spillover_detail(spillover_id: int):
    note_added   = request.args.get("note_added")   == "1"
    note_saved   = request.args.get("note_saved")   == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    conn = _get_conn()
    try:
        row = database.get_spillover_by_id(conn, spillover_id)
        if row is None:
            return render_template("404.html"), 404
        notes = database.list_notes(conn, "spillover", str(spillover_id))
        attachments_by_note = database.get_attachments_for_notes(conn, [n["id"] for n in notes])
    finally:
        conn.close()
    return render_template(
        "spillover_detail.html", row=row,
        notes=notes, attachments_by_note=attachments_by_note,
        note_added=note_added, note_saved=note_saved, note_deleted=note_deleted,
    )


@app.route("/spillover/<int:spillover_id>/notes/add", methods=["GET", "POST"])
def spillover_note_add(spillover_id: int):
    conn = _get_conn()
    try:
        row = database.get_spillover_by_id(conn, spillover_id)
        if row is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if not note_text:
                return render_template(
                    "note_form.html", mode="add",
                    entity_label=row.get("name") or f"Spillover #{spillover_id}",
                    list_url=url_for("spillover_list"), list_label="Core South Spillover",
                    detail_url=url_for("spillover_detail", spillover_id=spillover_id),
                    action_url=url_for("spillover_note_add", spillover_id=spillover_id),
                    cancel_url=url_for("spillover_detail", spillover_id=spillover_id),
                    heading=heading or "", note_text="", error="Note text is required.",
                )
            database.add_note(conn, "spillover", str(spillover_id), heading, note_text)
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("spillover_detail", spillover_id=spillover_id, note_added="1"))
    return render_template(
        "note_form.html", mode="add",
        entity_label=row.get("name") or f"Spillover #{spillover_id}",
        list_url=url_for("spillover_list"), list_label="Core South Spillover",
        detail_url=url_for("spillover_detail", spillover_id=spillover_id),
        action_url=url_for("spillover_note_add", spillover_id=spillover_id),
        cancel_url=url_for("spillover_detail", spillover_id=spillover_id),
        heading="", note_text="", error=None,
    )


@app.route("/spillover/<int:spillover_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def spillover_note_edit(spillover_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "spillover" or note["entity_id"] != str(spillover_id):
            return render_template("404.html"), 404
        row = database.get_spillover_by_id(conn, spillover_id)
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                return redirect(url_for("spillover_detail", spillover_id=spillover_id, note_saved="1"))
    finally:
        conn.close()
    label = (row.get("name") if row else None) or f"Spillover #{spillover_id}"
    kwargs = dict(
        mode="edit", entity_label=label,
        list_url=url_for("spillover_list"), list_label="Core South Spillover",
        detail_url=url_for("spillover_detail", spillover_id=spillover_id),
        action_url=url_for("spillover_note_edit", spillover_id=spillover_id, note_id=note_id),
        cancel_url=url_for("spillover_detail", spillover_id=spillover_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/spillover/<int:spillover_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def spillover_note_delete(spillover_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "spillover" or note["entity_id"] != str(spillover_id):
            return render_template("404.html"), 404
        if request.method == "POST":
            database.delete_note(conn, note_id)
            return redirect(url_for("spillover_detail", spillover_id=spillover_id, note_deleted="1"))
        row = database.get_spillover_by_id(conn, spillover_id)
    finally:
        conn.close()
    label = (row.get("name") if row else None) or f"Spillover #{spillover_id}"
    return render_template(
        "note_confirm_delete.html", note=note,
        entity_label=label,
        cancel_url=url_for("spillover_detail", spillover_id=spillover_id),
        delete_url=url_for("spillover_note_delete", spillover_id=spillover_id, note_id=note_id),
    )


# ---------------------------------------------------------------------------
# Spillover report — selection + view
# ---------------------------------------------------------------------------

@app.route("/spillover/report")
def spillover_report_select():
    area        = request.args.getlist("area")
    type_       = request.args.getlist("type")
    status      = request.args.getlist("status")
    assigned_to = request.args.getlist("assigned_to")
    critical    = request.args.getlist("critical")
    show_all    = request.args.get("show_all") == "1"
    hidden      = _cfg.get("spillover_hidden_statuses", [])
    exclude     = [] if (show_all or status) else hidden
    conn = _get_conn()
    try:
        rows            = database.get_spillover(conn, statuses=status or None, areas=area or None,
                              types=type_ or None, assignees=assigned_to or None,
                              critical=critical or None, exclude_statuses=exclude or None)
        options         = database.get_spillover_filter_options(conn)
        selected_ids    = database.get_spillover_report_ids(conn)
        report_comments = database.list_report_comments(conn, "spillover")
    finally:
        conn.close()
    return render_template(
        "spillover_report_select.html",
        rows=rows, options=options, selected_ids=selected_ids,
        area=area, type_=type_, status=status, assigned_to=assigned_to,
        critical=critical, show_all=show_all,
        report_comments=report_comments,
    )


@app.route("/spillover/<int:spillover_id>/report-toggle", methods=["POST"])
def spillover_report_toggle(spillover_id: int):
    conn = _get_conn()
    try:
        included = database.toggle_spillover_report_item(conn, spillover_id)
        total    = len(database.get_spillover_report_ids(conn))
    finally:
        conn.close()
    return jsonify({"ok": True, "included": included, "total": total})


@app.route("/spillover/report/include-ids", methods=["POST"])
def spillover_report_include_ids():
    ids  = [int(i) for i in request.form.getlist("ids") if i.isdigit()]
    conn = _get_conn()
    try:
        database.include_spillover_report_ids(conn, ids)
        total = len(database.get_spillover_report_ids(conn))
    finally:
        conn.close()
    return jsonify({"ok": True, "total": total})


@app.route("/spillover/report/clear", methods=["POST"])
def spillover_report_clear():
    conn = _get_conn()
    try:
        database.clear_spillover_report(conn)
    finally:
        conn.close()
    return jsonify({"ok": True, "total": 0})


@app.route("/spillover/report/view")
def spillover_report_view():
    conn = _get_conn()
    try:
        items           = database.get_spillover_report_items(conn)
        order_details   = {}
        for item in items:
            od = database.list_order_details(conn, "spillover", str(item["spillover_id"]))
            if od:
                order_details[item["spillover_id"]] = od
        report_comments = database.list_report_comments(conn, "spillover")
    finally:
        conn.close()
    _crit_order = {"yes": 0, "slightly": 1, "no": 2}
    items = sorted(items, key=lambda r: _crit_order.get(r.get("critical_for_signoff") or "", 3))
    from datetime import date
    return render_template(
        "spillover_report_view.html",
        items=items, order_details=order_details,
        report_comments=report_comments,
        today=date.today().strftime("%Y-%m-%d"),
    )


# ---------------------------------------------------------------------------
# Generic order_details routes — work for any entity type
# ---------------------------------------------------------------------------

@app.route("/order-details/<entity_type>/<entity_id>")
def order_details_list(entity_type: str, entity_id: str):
    conn = _get_conn()
    try:
        rows = database.list_order_details(conn, entity_type, entity_id)
    finally:
        conn.close()
    return jsonify(rows)


@app.route("/order-details/<entity_type>/<entity_id>/add", methods=["POST"])
def order_details_add(entity_type: str, entity_id: str):
    conn = _get_conn()
    try:
        detail_id = database.add_order_detail(conn, entity_type, entity_id)
    finally:
        conn.close()
    return jsonify({"ok": True, "id": detail_id})


@app.route("/order-details/<int:detail_id>/update", methods=["POST"])
def order_detail_update(detail_id: int):
    order_type   = request.form.get("order_type",   "").strip()
    order_number = request.form.get("order_number", "").strip()
    comment      = request.form.get("comment",      "").strip()
    docs_in_s4   = 1 if request.form.get("docs_in_s4") == "1" else 0
    conn = _get_conn()
    try:
        database.update_order_detail(conn, detail_id, order_type, order_number, comment, docs_in_s4)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/order-details/<int:detail_id>/delete", methods=["POST"])
def order_detail_delete(detail_id: int):
    conn = _get_conn()
    try:
        database.delete_order_detail(conn, detail_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# ECOM Gatekeeper Check
# ---------------------------------------------------------------------------

@app.route("/ecom-gatekeeper")
def ecom_gatekeeper_list():
    conn = _get_conn()
    try:
        rows = database.list_ecom_gatekeeper_rows(conn)
    finally:
        conn.close()
    return render_template("ecom_gatekeeper.html", rows=rows)


@app.route("/ecom-gatekeeper/add", methods=["POST"])
def ecom_gatekeeper_add():
    conn = _get_conn()
    try:
        row_id = database.add_ecom_gatekeeper_row(conn)
    finally:
        conn.close()
    return jsonify({"ok": True, "id": row_id})


@app.route("/ecom-gatekeeper/<int:row_id>/update", methods=["POST"])
def ecom_gatekeeper_update(row_id: int):
    jira_id       = request.form.get("jira_id",       "").strip()
    solman_id     = request.form.get("solman_id",     "").strip()
    testcase_name = request.form.get("testcase_name", "").strip()
    status        = request.form.get("status",        "open").strip()
    next_step     = request.form.get("next_step",     "").strip()
    conn = _get_conn()
    try:
        database.update_ecom_gatekeeper_row(conn, row_id, jira_id, solman_id, testcase_name, status, next_step)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/ecom-gatekeeper/<int:row_id>/delete", methods=["POST"])
def ecom_gatekeeper_delete(row_id: int):
    conn = _get_conn()
    try:
        database.delete_ecom_gatekeeper_row(conn, row_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/ecom-gatekeeper/<int:row_id>")
def ecom_gatekeeper_detail(row_id: int):
    note_added   = request.args.get("note_added")   == "1"
    note_saved   = request.args.get("note_saved")   == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    conn = _get_conn()
    try:
        row = database.get_ecom_gatekeeper_row(conn, row_id)
        if row is None:
            return render_template("404.html"), 404
        notes = database.list_notes(conn, "ecom_gatekeeper", str(row_id))
        attachments_by_note = database.get_attachments_for_notes(conn, [n["id"] for n in notes])
    finally:
        conn.close()
    return render_template(
        "ecom_gatekeeper_detail.html", row=row,
        notes=notes, attachments_by_note=attachments_by_note,
        note_added=note_added, note_saved=note_saved, note_deleted=note_deleted,
    )


@app.route("/ecom-gatekeeper/<int:row_id>/notes/add", methods=["GET", "POST"])
def ecom_gatekeeper_note_add(row_id: int):
    conn = _get_conn()
    try:
        row = database.get_ecom_gatekeeper_row(conn, row_id)
        if row is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if not note_text:
                return render_template(
                    "note_form.html", mode="add",
                    entity_label=row.get("testcase_name") or f"GK #{row_id}",
                    list_url=url_for("ecom_gatekeeper_list"), list_label="ECOM Gatekeeper",
                    detail_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
                    action_url=url_for("ecom_gatekeeper_note_add", row_id=row_id),
                    cancel_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
                    heading=heading or "", note_text="", error="Note text is required.",
                )
            database.add_note(conn, "ecom_gatekeeper", str(row_id), heading, note_text)
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("ecom_gatekeeper_detail", row_id=row_id, note_added="1"))
    return render_template(
        "note_form.html", mode="add",
        entity_label=row.get("testcase_name") or f"GK #{row_id}",
        list_url=url_for("ecom_gatekeeper_list"), list_label="ECOM Gatekeeper",
        detail_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
        action_url=url_for("ecom_gatekeeper_note_add", row_id=row_id),
        cancel_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
        heading="", note_text="", error=None,
    )


@app.route("/ecom-gatekeeper/<int:row_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def ecom_gatekeeper_note_edit(row_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "ecom_gatekeeper" or note["entity_id"] != str(row_id):
            return render_template("404.html"), 404
        row = database.get_ecom_gatekeeper_row(conn, row_id)
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                return redirect(url_for("ecom_gatekeeper_detail", row_id=row_id, note_saved="1"))
    finally:
        conn.close()
    label = (row.get("testcase_name") if row else None) or f"GK #{row_id}"
    kwargs = dict(
        mode="edit", entity_label=label,
        list_url=url_for("ecom_gatekeeper_list"), list_label="ECOM Gatekeeper",
        detail_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
        action_url=url_for("ecom_gatekeeper_note_edit", row_id=row_id, note_id=note_id),
        cancel_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/ecom-gatekeeper/<int:row_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def ecom_gatekeeper_note_delete(row_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "ecom_gatekeeper" or note["entity_id"] != str(row_id):
            return render_template("404.html"), 404
        if request.method == "POST":
            database.delete_note(conn, note_id)
            return redirect(url_for("ecom_gatekeeper_detail", row_id=row_id, note_deleted="1"))
        row = database.get_ecom_gatekeeper_row(conn, row_id)
    finally:
        conn.close()
    label = (row.get("testcase_name") if row else None) or f"GK #{row_id}"
    return render_template(
        "note_confirm_delete.html", note=note,
        entity_label=label,
        cancel_url=url_for("ecom_gatekeeper_detail", row_id=row_id),
        delete_url=url_for("ecom_gatekeeper_note_delete", row_id=row_id, note_id=note_id),
    )


# ---------------------------------------------------------------------------
# Report comments — bullet points shown in the status reports
# ---------------------------------------------------------------------------

@app.route("/report-comments/<report>/add", methods=["POST"])
def report_comment_add(report: str):
    if report not in ("spillover", "retail"):
        return jsonify({"ok": False}), 400
    comment = request.form.get("comment", "").strip()
    conn = _get_conn()
    try:
        new_id = database.add_report_comment(conn, report, comment)
    finally:
        conn.close()
    return jsonify({"ok": True, "row": {"id": new_id, "comment": comment}})


@app.route("/report-comments/<int:comment_id>/update", methods=["POST"])
def report_comment_update(comment_id: int):
    comment = request.form.get("comment", "").strip()
    conn = _get_conn()
    try:
        database.update_report_comment(conn, comment_id, comment)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/report-comments/<int:comment_id>/delete", methods=["POST"])
def report_comment_delete(comment_id: int):
    conn = _get_conn()
    try:
        database.delete_report_comment(conn, comment_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Retail routes
# ---------------------------------------------------------------------------

@app.route("/retail")
def retail_list():
    statuses      = request.args.getlist("status")
    assignees     = request.args.getlist("assigned_to")
    countries     = request.args.getlist("country")
    scenarios     = request.args.getlist("scenario")
    action_needed = request.args.get("action_needed", "")
    search_defect  = request.args.get("search_defect", "").strip() or None
    search_order   = request.args.get("search_order", "").strip() or None
    search_billing = request.args.get("search_billing", "").strip() or None

    conn = _get_conn()
    try:
        rows    = database.get_retail(
            conn,
            statuses=statuses or None,
            assignees=assignees or None,
            countries=countries or None,
            scenarios=scenarios or None,
            search_defect=search_defect,
            search_order=search_order,
            search_billing=search_billing,
            action_needed=action_needed or None,
        )
        options = database.get_retail_filter_options(conn)
    finally:
        conn.close()

    return render_template(
        "retail.html",
        rows=rows,
        options=options,
        statuses=statuses,
        assignees=assignees,
        countries=countries,
        scenarios=scenarios,
        action_needed=action_needed,
        search_defect=search_defect or "",
        search_order=search_order or "",
        search_billing=search_billing or "",
    )


@app.route("/retail/<int:retail_id>/annotation", methods=["POST"])
def retail_annotation_save(retail_id: int):
    next_step = request.form.get("next_step", "").strip() or None
    conn = _get_conn()
    try:
        existing        = database.get_retail_annotation(conn, retail_id)
        comment_history = existing["comment_history"] if existing else None
        action_needed   = existing["action_needed"] if existing else 0
        database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
        ann = database.get_retail_annotation(conn, retail_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "next_step": (ann["next_step"] or "") if ann else "",
    })


@app.route("/retail/<int:retail_id>/comment", methods=["POST"])
def retail_comment_save(retail_id: int):
    comment_history = request.form.get("comment_history", "").strip() or None
    conn = _get_conn()
    try:
        existing      = database.get_retail_annotation(conn, retail_id)
        next_step     = existing["next_step"] if existing else None
        action_needed = existing["action_needed"] if existing else 0
        database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
        ann = database.get_retail_annotation(conn, retail_id)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "comment_history": (ann["comment_history"] or "") if ann else "",
    })


def _get_retail_report():
    conn = _get_conn()
    try:
        status_counts = database.get_retail_status_counts(conn)
    finally:
        conn.close()
    mappings = load_status_mappings()
    return compute_retail_report(status_counts, mappings)


@app.route("/retail/report")
def retail_status_report():
    report = _get_retail_report()
    conn = _get_conn()
    try:
        blocked_defects = database.get_retail_defects_blocked(conn)
        report_comments = database.list_report_comments(conn, "retail")
    finally:
        conn.close()
    blocked_total   = sum(d["blocked_tc_count"] for d in blocked_defects)
    dtco2c_total    = sum(d["blocked_tc_count"] for d in blocked_defects if d["dtco2c"])
    sales_total     = sum(d["blocked_tc_count"] for d in blocked_defects if not d["dtco2c"])
    return render_template(
        "retail_report.html",
        report=report,
        today=date.today().isoformat(),
        blocked_defects=blocked_defects,
        blocked_defects_total=blocked_total,
        dtco2c_total=dtco2c_total,
        sales_total=sales_total,
        report_comments=report_comments,
    )


_RETAIL_REPORT_HEADERS = [
    "Date",
    "Back with Sales",
    "With DTC",
    "In Progress with DTC",
    "Passed with DTC",
    "Incoming (Gatekeeper)",
    "Ready for validation",
    "In Progress",
    "In Clarification",
    "Blocked",
]


def _append_retail_report_to_excel(report: dict, today: str) -> str:
    """Append one data row to the retail report Excel log. Returns the file path."""
    xlsx_path = Path(_cfg.get("retail_report_xlsx", "output/retail_report_log.xlsx"))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default blank sheet openpyxl creates
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Retail" not in wb.sheetnames:
        ws = wb.create_sheet("Retail")
    else:
        ws = wb["Retail"]

    if ws.cell(1, 1).value is None:
        for col, header in enumerate(_RETAIL_REPORT_HEADERS, 1):
            ws.cell(row=1, column=col).value = header

    b = report["buckets"]
    next_row = ws.max_row + 1
    for col, val in enumerate([
        today,
        b["back_with_sales"],
        b["with_dtc"],
        b["in_progress_with_dtc"],
        b["passed_with_dtc"],
        b["incoming_gatekeeper"],
        b["ready_for_validation"],
        b["in_progress"],
        b["in_clarification"],
        b["blocked"],
    ], 1):
        ws.cell(row=next_row, column=col).value = val

    wb.save(xlsx_path)
    return str(xlsx_path)


@app.route("/retail/report/save-excel", methods=["POST"])
def retail_report_save_excel():
    try:
        report     = _get_retail_report()
        save_date  = request.form.get("date") or date.today().isoformat()
        path       = _append_retail_report_to_excel(report, save_date)
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)})
    return jsonify({"ok": True, "path": path, "date": save_date})


@app.route("/retail/report/download")
def retail_report_download():
    report = _get_retail_report()
    today  = date.today().isoformat()
    conn   = _get_conn()
    try:
        report_comments = database.list_report_comments(conn, "retail")
    finally:
        conn.close()
    html = render_template(
        "retail_report_download.html", report=report, today=today,
        report_comments=report_comments,
    )
    return html, 200, {
        "Content-Type": "text/html; charset=utf-8",
        "Content-Disposition": f'attachment; filename="retail_report_{today}.html"',
    }


# ---------------------------------------------------------------------------
# Signoff reports
# ---------------------------------------------------------------------------

def _prepare_report(rows: list) -> list[tuple[str, list]]:
    """Group rows by critical_for_signoff, sort each group by (signoff_group, name)."""
    order = [
        ("yes",      "Critical for sign-off: Yes"),
        ("slightly", "Critical for sign-off: Slightly"),
        ("no",       "Critical for sign-off: No"),
        ("",         "Critical for sign-off: Not set"),
    ]
    groups: dict[str, list] = {k: [] for k, _ in order}
    for r in rows:
        key = (r.get("critical_for_signoff") or "").lower()
        groups[key if key in groups else ""].append(r)
    for grp in groups.values():
        grp.sort(key=lambda r: (
            (r.get("signoff_group") or "").lower() or "\xff",
            (r.get("name") or "").lower(),
        ))
    return [(label, groups[key]) for key, label in order if groups[key]]


@app.route("/report/retail")
def retail_report():
    hidden = _cfg.get("spillover_hidden_statuses", [])
    exclude_areas = {"ecom", "omni"}
    conn = _get_conn()
    try:
        rows = database.get_spillover(conn, exclude_statuses=hidden or None)
    finally:
        conn.close()
    rows = [r for r in rows if (r.get("area") or "").lower() not in exclude_areas]
    return render_template("report.html",
        title="Retail Spillover Report",
        report_date=date.today().isoformat(),
        sections=_prepare_report(rows),
        total=len(rows),
        prod_defects=[])


@app.route("/report/ecom")
def ecom_report():
    hidden = _cfg.get("spillover_hidden_statuses", [])
    ecom_areas = {"ecom", "omni"}
    conn = _get_conn()
    try:
        rows = database.get_spillover(conn, exclude_statuses=hidden or None)
        prod_defects = database.list_known_prod_defects(conn)
    finally:
        conn.close()
    rows = [r for r in rows if (r.get("area") or "").lower() in ecom_areas]
    return render_template("report.html",
        title="ECOM / Omni Spillover Report",
        report_date=date.today().isoformat(),
        sections=_prepare_report(rows),
        total=len(rows),
        prod_defects=prod_defects)


# ---------------------------------------------------------------------------
# Known production defects routes
# ---------------------------------------------------------------------------

@app.route("/prod_defects")
def prod_defects_list():
    conn = _get_conn()
    try:
        rows = database.list_known_prod_defects(conn)
    finally:
        conn.close()
    return render_template("prod_defects.html", rows=rows)


@app.route("/prod_defects/new", methods=["GET", "POST"])
def prod_defect_new():
    if request.method == "POST":
        def _f(name): return request.form.get(name, "").strip() or None
        conn = _get_conn()
        try:
            row = database.create_known_prod_defect(
                conn,
                short_description=_f("short_description"),
                scenario=_f("scenario"),
                description=_f("description"),
                biz_impact=_f("biz_impact"),
                numbers=_f("numbers"),
                refs=_f("refs"),
                next_steps=_f("next_steps"),
                comments=_f("comments"),
                confluence=_f("confluence"),
            )
        finally:
            conn.close()
        return redirect(url_for("prod_defect_detail", record_id=row["id"], saved="1"))
    return render_template("prod_defect_detail.html", record={}, is_new=True, saved=False)


@app.route("/prod_defects/<int:record_id>", methods=["GET", "POST"])
def prod_defect_detail(record_id: int):
    conn = _get_conn()
    try:
        record = database.get_known_prod_defect(conn, record_id)
        if record is None:
            return _not_found(str(record_id))
        if request.method == "POST":
            def _f(name): return request.form.get(name, "").strip() or None
            database.update_known_prod_defect(
                conn, record_id,
                short_description=_f("short_description"),
                scenario=_f("scenario"),
                description=_f("description"),
                biz_impact=_f("biz_impact"),
                numbers=_f("numbers"),
                refs=_f("refs"),
                next_steps=_f("next_steps"),
                comments=_f("comments"),
                confluence=_f("confluence"),
            )
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("prod_defect_detail", record_id=record_id, saved="1"))
    saved = request.args.get("saved") == "1"
    return render_template("prod_defect_detail.html", record=record, is_new=False, saved=saved)


@app.route("/prod_defects/<int:record_id>/delete", methods=["POST"])
def prod_defect_delete(record_id: int):
    conn = _get_conn()
    try:
        database.delete_known_prod_defect(conn, record_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Links routes
# ---------------------------------------------------------------------------

@app.route("/links")
def links_list():
    areas       = request.args.getlist("area")
    tools       = request.args.getlist("tool")
    tags        = request.args.getlist("tag")
    search      = request.args.get("search", "").strip()
    conn = _get_conn()
    try:
        rows    = database.list_links(conn, areas=areas or None, tools=tools or None,
                                      tags=tags or None, search=search or None)
        options = database.get_link_options(conn)
    finally:
        conn.close()
    return render_template("links.html", rows=rows, options=options,
                           areas=areas, tools=tools, tags=tags, search=search)


@app.route("/links/new", methods=["GET", "POST"])
def link_new():
    if request.method == "POST":
        def _f(name): return request.form.get(name, "").strip() or None
        conn = _get_conn()
        try:
            row = database.create_link(
                conn,
                description=_f("description"),
                url=_f("url"),
                area=_f("area"),
                tool=_f("tool"),
                tags=_f("tags"),
            )
        finally:
            conn.close()
        return redirect(url_for("link_detail", link_id=row["id"], saved="1"))
    return render_template("link_detail.html", record={}, is_new=True, saved=False)


@app.route("/links/<int:link_id>", methods=["GET", "POST"])
def link_detail(link_id: int):
    saved = request.args.get("saved") == "1"
    conn = _get_conn()
    try:
        record = database.get_link(conn, link_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            def _f(name): return request.form.get(name, "").strip() or None
            database.update_link(
                conn, link_id,
                description=_f("description"),
                url=_f("url"),
                area=_f("area"),
                tool=_f("tool"),
                tags=_f("tags"),
            )
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("link_detail", link_id=link_id, saved="1"))
    return render_template("link_detail.html", record=record, is_new=False, saved=saved)


@app.route("/links/<int:link_id>/delete", methods=["POST"])
def link_delete(link_id: int):
    conn = _get_conn()
    try:
        database.delete_link(conn, link_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Contacts routes
# ---------------------------------------------------------------------------

@app.route("/contacts")
def contacts_list():
    areas  = request.args.getlist("area")
    topics = request.args.getlist("topic")
    tags   = request.args.getlist("tag")
    search = request.args.get("search", "").strip()
    conn = _get_conn()
    try:
        rows    = database.list_contacts(conn, areas=areas or None, topics=topics or None,
                                         tags=tags or None, search=search or None)
        options = database.get_contact_options(conn)
    finally:
        conn.close()
    return render_template("contacts.html", rows=rows, options=options,
                           areas=areas, topics=topics, tags=tags, search=search)


@app.route("/contacts/new", methods=["GET", "POST"])
def contact_new():
    if request.method == "POST":
        def _f(name): return request.form.get(name, "").strip() or None
        conn = _get_conn()
        try:
            row = database.create_contact(
                conn,
                name=_f("name") or "",
                email=_f("email"),
                area=_f("area"),
                topic=_f("topic"),
                comments=_f("comments"),
                tags=_f("tags"),
            )
        finally:
            conn.close()
        return redirect(url_for("contact_detail", contact_id=row["id"], saved="1"))
    return render_template("contact_detail.html", record={}, is_new=True, saved=False)


@app.route("/contacts/<int:contact_id>", methods=["GET", "POST"])
def contact_detail(contact_id: int):
    saved = request.args.get("saved") == "1"
    conn = _get_conn()
    try:
        record = database.get_contact(conn, contact_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            def _f(name): return request.form.get(name, "").strip() or None
            database.update_contact(
                conn, contact_id,
                name=_f("name") or "",
                email=_f("email"),
                area=_f("area"),
                topic=_f("topic"),
                comments=_f("comments"),
                tags=_f("tags"),
            )
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("contact_detail", contact_id=contact_id, saved="1"))
    return render_template("contact_detail.html", record=record, is_new=False, saved=saved)


@app.route("/contacts/<int:contact_id>/delete", methods=["POST"])
def contact_delete(contact_id: int):
    conn = _get_conn()
    try:
        database.delete_contact(conn, contact_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Encouragements
# ---------------------------------------------------------------------------

@app.route("/encouragements")
def encouragements_list():
    person_id = request.args.get("person_id", type=int)
    added     = request.args.get("added") == "1"
    conn = _get_conn()
    try:
        people = database.list_encouragement_people(conn)
        items  = database.list_encouragements(conn, person_id=person_id)
        person = next((p for p in people if p["id"] == person_id), None) if person_id else None
    finally:
        conn.close()
    return render_template(
        "encouragements.html",
        people=people,
        items=items,
        person=person,
        person_id=person_id,
        today=date.today().isoformat(),
        added=added,
    )


@app.route("/encouragements/add", methods=["POST"])
def encouragement_add():
    name     = request.form.get("person_name", "").strip()
    text     = request.form.get("text", "").strip()
    enc_date = request.form.get("date", "").strip() or date.today().isoformat()
    if not name or not text:
        return redirect(url_for("encouragements_list"))
    conn = _get_conn()
    try:
        person_id = database.get_or_create_encouragement_person(conn, name)
        database.add_encouragement(conn, person_id, text, enc_date)
    finally:
        conn.close()
    return redirect(url_for("encouragements_list", person_id=person_id, added="1"))


@app.route("/encouragements/<int:enc_id>/delivered", methods=["POST"])
def encouragement_toggle_delivered(enc_id: int):
    value = request.json.get("value", False) if request.is_json else bool(request.form.get("value"))
    conn = _get_conn()
    try:
        database.set_encouragement_delivered(conn, enc_id, value)
    finally:
        conn.close()
    return {"ok": True}


@app.route("/encouragements/<int:enc_id>/delete", methods=["POST"])
def encouragement_delete(enc_id: int):
    conn = _get_conn()
    try:
        database.delete_encouragement(conn, enc_id)
    finally:
        conn.close()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Test Learnings routes
# ---------------------------------------------------------------------------

@app.route("/test_learnings")
def test_learning_list():
    channels = request.args.getlist("channel")
    tags     = request.args.getlist("tag")
    conn = _get_conn()
    try:
        rows    = database.list_test_learnings(conn, channels=channels or None, tags=tags or None)
        options = database.get_test_learning_options(conn)
    finally:
        conn.close()
    return render_template("test_learning_list.html", rows=rows, options=options,
                           channels=channels, tags=tags)


@app.route("/test_learnings/new", methods=["GET", "POST"])
def test_learning_new():
    if request.method == "POST":
        def _f(n): return request.form.get(n, "").strip() or None
        channel  = request.form.get("channel", "").strip() or "Retail"
        learning = request.form.get("learning", "").strip()
        conn = _get_conn()
        try:
            row = database.create_test_learning(
                conn, channel=channel, topic=_f("topic"), learning=learning,
                scenario=_f("scenario"), tags=_f("tags"),
            )
        finally:
            conn.close()
        return redirect(url_for("test_learning_detail", learning_id=row["id"], saved="1"))
    prefill_channel = request.args.get("channel", "Retail")
    return render_template("test_learning_detail.html", record={"channel": prefill_channel},
                           is_new=True, saved=False)


@app.route("/test_learnings/<int:learning_id>", methods=["GET", "POST"])
def test_learning_detail(learning_id: int):
    saved        = request.args.get("saved")        == "1"
    note_added   = request.args.get("note_added")   == "1"
    note_saved   = request.args.get("note_saved")   == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    conn = _get_conn()
    try:
        record = database.get_test_learning(conn, learning_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            def _f(n): return request.form.get(n, "").strip() or None
            channel  = request.form.get("channel", "").strip() or "Retail"
            learning = request.form.get("learning", "").strip()
            database.update_test_learning(
                conn, learning_id, channel=channel, topic=_f("topic"), learning=learning,
                scenario=_f("scenario"), tags=_f("tags"),
            )
        notes = database.list_notes(conn, "test_learning", str(learning_id))
        attachments_by_note = database.get_attachments_for_notes(conn, [n["id"] for n in notes])
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("test_learning_detail", learning_id=learning_id, saved="1"))
    return render_template(
        "test_learning_detail.html", record=record, is_new=False, saved=saved,
        notes=notes, attachments_by_note=attachments_by_note,
        note_added=note_added, note_saved=note_saved, note_deleted=note_deleted,
    )


@app.route("/test_learnings/<int:learning_id>/delete", methods=["POST"])
def test_learning_delete(learning_id: int):
    conn = _get_conn()
    try:
        database.delete_test_learning(conn, learning_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/test_learnings/<int:learning_id>/notes")
def test_learning_notes(learning_id: int):
    conn = _get_conn()
    try:
        notes = database.list_notes(conn, "test_learning", str(learning_id))
    finally:
        conn.close()
    return jsonify(notes)


@app.route("/test_learnings/<int:learning_id>/notes/add", methods=["GET", "POST"])
def test_learning_note_add(learning_id: int):
    conn = _get_conn()
    try:
        record = database.get_test_learning(conn, learning_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if not note_text:
                return render_template(
                    "note_form.html", mode="add",
                    entity_label=f"Learning #{learning_id}",
                    list_url=url_for("test_learning_list"), list_label="Test Learnings",
                    detail_url=url_for("test_learning_detail", learning_id=learning_id),
                    action_url=url_for("test_learning_note_add", learning_id=learning_id),
                    cancel_url=url_for("test_learning_detail", learning_id=learning_id),
                    heading=heading or "", note_text="", error="Note text is required.",
                )
            database.add_note(conn, "test_learning", str(learning_id), heading, note_text)
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("test_learning_detail", learning_id=learning_id, note_added="1"))
    return render_template(
        "note_form.html", mode="add",
        entity_label=f"Learning #{learning_id}",
        list_url=url_for("test_learning_list"), list_label="Test Learnings",
        detail_url=url_for("test_learning_detail", learning_id=learning_id),
        action_url=url_for("test_learning_note_add", learning_id=learning_id),
        cancel_url=url_for("test_learning_detail", learning_id=learning_id),
        heading="", note_text="", error=None,
    )


@app.route("/test_learnings/<int:learning_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def test_learning_note_edit(learning_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "test_learning" or note["entity_id"] != str(learning_id):
            return render_template("404.html"), 404
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                return redirect(url_for("test_learning_detail", learning_id=learning_id, note_saved="1"))
    finally:
        conn.close()
    kwargs = dict(
        mode="edit", entity_label=f"Learning #{learning_id}",
        list_url=url_for("test_learning_list"), list_label="Test Learnings",
        detail_url=url_for("test_learning_detail", learning_id=learning_id),
        action_url=url_for("test_learning_note_edit", learning_id=learning_id, note_id=note_id),
        cancel_url=url_for("test_learning_detail", learning_id=learning_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/test_learnings/<int:learning_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def test_learning_note_delete(learning_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "test_learning" or note["entity_id"] != str(learning_id):
            return render_template("404.html"), 404
        if request.method == "POST":
            database.delete_note(conn, note_id)
            return redirect(url_for("test_learning_detail", learning_id=learning_id, note_deleted="1"))
    finally:
        conn.close()
    return render_template(
        "note_confirm_delete.html", note=note,
        entity_label=f"Learning #{learning_id}",
        cancel_url=url_for("test_learning_detail", learning_id=learning_id),
        delete_url=url_for("test_learning_note_delete", learning_id=learning_id, note_id=note_id),
    )


# ---------------------------------------------------------------------------
# Test Limitations routes
# ---------------------------------------------------------------------------

@app.route("/test_limitations")
def test_limitation_list():
    channels = request.args.getlist("channel")
    conn = _get_conn()
    try:
        rows    = database.list_test_limitations(conn, channels=channels or None)
        options = database.get_test_limitation_options(conn)
    finally:
        conn.close()
    return render_template("test_limitation_list.html", rows=rows, options=options,
                           channels=channels)


@app.route("/test_limitations/new", methods=["GET", "POST"])
def test_limitation_new():
    if request.method == "POST":
        def _f(n): return request.form.get(n, "").strip() or None
        channel = request.form.get("channel", "").strip() or "Retail"
        limitation = request.form.get("limitation", "").strip()
        conn = _get_conn()
        try:
            row = database.create_test_limitation(
                conn, channel=channel, limitation=limitation,
                scenario=_f("scenario"), comment=_f("comment"),
            )
        finally:
            conn.close()
        return redirect(url_for("test_limitation_detail", limitation_id=row["id"], saved="1"))
    prefill_channel = request.args.get("channel", "Retail")
    return render_template("test_limitation_detail.html", record={"channel": prefill_channel},
                           is_new=True, saved=False)


@app.route("/test_limitations/<int:limitation_id>", methods=["GET", "POST"])
def test_limitation_detail(limitation_id: int):
    saved = request.args.get("saved") == "1"
    conn = _get_conn()
    try:
        record = database.get_test_limitation(conn, limitation_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            def _f(n): return request.form.get(n, "").strip() or None
            channel    = request.form.get("channel", "").strip() or "Retail"
            limitation = request.form.get("limitation", "").strip()
            database.update_test_limitation(
                conn, limitation_id, channel=channel, limitation=limitation,
                scenario=_f("scenario"), comment=_f("comment"),
            )
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("test_limitation_detail", limitation_id=limitation_id, saved="1"))
    return render_template("test_limitation_detail.html", record=record, is_new=False, saved=saved)


@app.route("/test_limitations/<int:limitation_id>/delete", methods=["POST"])
def test_limitation_delete(limitation_id: int):
    conn = _get_conn()
    try:
        database.delete_test_limitation(conn, limitation_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/test_limitations/<int:limitation_id>/notes")
def test_limitation_notes(limitation_id: int):
    conn = _get_conn()
    try:
        notes = database.list_notes(conn, "test_limitation", str(limitation_id))
    finally:
        conn.close()
    return jsonify(notes)


@app.route("/test_limitations/<int:limitation_id>/notes/add", methods=["POST"])
def test_limitation_note_add(limitation_id: int):
    note = request.form.get("note", "").strip()
    if not note:
        return jsonify({"ok": False, "error": "empty"})
    conn = _get_conn()
    try:
        database.add_note(conn, "test_limitation", str(limitation_id), None, note)
        notes = database.list_notes(conn, "test_limitation", str(limitation_id))
    finally:
        conn.close()
    return jsonify({"ok": True, "notes": notes})


# ---------------------------------------------------------------------------
# Core South Follow-Up Tracker routes
# ---------------------------------------------------------------------------

@app.route("/cs_followups")
def cs_followup_list():
    areas     = request.args.getlist("area")
    with_whom = request.args.getlist("with_whom")
    statuses  = request.args.getlist("status")
    show_done = request.args.get("done") == "1"
    conn = _get_conn()
    try:
        rows    = database.list_cs_followups(conn, areas=areas or None,
                                             with_whom=with_whom or None,
                                             statuses=statuses or None,
                                             include_done=show_done)
        options = database.get_cs_followup_options(conn)
    finally:
        conn.close()
    return render_template("cs_followup_list.html", rows=rows, options=options,
                           areas=areas, with_whom=with_whom, statuses=statuses,
                           show_done=show_done,
                           all_statuses=database.CS_FOLLOWUP_STATUSES,
                           today=date.today().isoformat())


@app.route("/cs_followups/new", methods=["GET", "POST"])
def cs_followup_new():
    if request.method == "POST":
        def _f(n): return request.form.get(n, "").strip() or None
        conn = _get_conn()
        try:
            row = database.create_cs_followup(
                conn,
                area=_f("area"), jira_id=_f("jira_id"), topic=request.form.get("topic", "").strip(),
                description=_f("description"), next_step=_f("next_step"), with_whom=_f("with_whom"),
            )
        finally:
            conn.close()
        return redirect(url_for("cs_followup_detail", followup_id=row["id"], saved="1"))
    return render_template("cs_followup_detail.html", record={}, is_new=True, saved=False)


@app.route("/cs_followups/<int:followup_id>", methods=["GET", "POST"])
def cs_followup_detail(followup_id: int):
    saved = request.args.get("saved") == "1"
    conn = _get_conn()
    try:
        record = database.get_cs_followup(conn, followup_id)
        if record is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            def _f(n): return request.form.get(n, "").strip() or None
            database.update_cs_followup(
                conn, followup_id,
                area=_f("area"), jira_id=_f("jira_id"),
                topic=request.form.get("topic", "").strip(),
                description=_f("description"), next_step=_f("next_step"),
                with_whom=_f("with_whom"),
            )
    finally:
        conn.close()
    if request.method == "POST":
        return redirect(url_for("cs_followup_detail", followup_id=followup_id, saved="1"))
    return render_template("cs_followup_detail.html", record=record, is_new=False, saved=saved)


@app.route("/cs_followups/<int:followup_id>/status", methods=["POST"])
def cs_followup_status(followup_id: int):
    status = request.form.get("status", "")
    if status not in database.CS_FOLLOWUP_STATUSES:
        return jsonify({"ok": False})
    conn = _get_conn()
    try:
        database.set_cs_followup_status(conn, followup_id, status)
    finally:
        conn.close()
    return jsonify({"ok": True, "status": status})


@app.route("/cs_followups/<int:followup_id>/delete", methods=["POST"])
def cs_followup_delete(followup_id: int):
    conn = _get_conn()
    try:
        database.delete_cs_followup(conn, followup_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/cs_followups/<int:followup_id>/notes")
def cs_followup_notes(followup_id: int):
    conn = _get_conn()
    try:
        notes = database.list_notes(conn, "cs_followup", str(followup_id))
    finally:
        conn.close()
    return jsonify(notes)


@app.route("/cs_followups/<int:followup_id>/notes/add", methods=["POST"])
def cs_followup_note_add(followup_id: int):
    note = request.form.get("note", "").strip()
    if not note:
        return jsonify({"ok": False, "error": "empty"})
    conn = _get_conn()
    try:
        database.add_note(conn, "cs_followup", str(followup_id), None, note)
        notes = database.list_notes(conn, "cs_followup", str(followup_id))
    finally:
        conn.close()
    return jsonify({"ok": True, "notes": notes})


# ---------------------------------------------------------------------------
# Note routes
# ---------------------------------------------------------------------------

@app.route("/defects/<defect_id>/dtco2c", methods=["POST"])
def defect_toggle_dtco2c(defect_id: str):
    value = request.json.get("value", False) if request.is_json else bool(request.form.get("value"))
    conn = _get_conn()
    try:
        database.set_defect_dtco2c(conn, defect_id, value)
    finally:
        conn.close()
    return {"ok": True}


@app.route("/defects/<defect_id>/daily", methods=["POST"])
def defect_toggle_daily(defect_id: str):
    value = request.json.get("value", False) if request.is_json else bool(request.form.get("value"))
    conn = _get_conn()
    try:
        database.set_defect_daily(conn, defect_id, value)
    finally:
        conn.close()
    return {"ok": True}


@app.route("/defects/<defect_id>/notes/add")
def note_add_form(defect_id: str):
    conn = _get_conn()
    try:
        defect = database.get_defect(conn, defect_id)
    finally:
        conn.close()
    if defect is None:
        return _not_found(defect_id)
    return_to = request.args.get("return_to", "detail")
    return _render_note_add_form(defect_id, defect.get("solman_name"), return_to)


@app.route("/defects/<defect_id>/notes", methods=["POST"])
def note_add(defect_id: str):
    conn = _get_conn()
    try:
        defect = database.get_defect(conn, defect_id)
        if defect is None:
            return _not_found(defect_id)
        heading = request.form.get("heading", "").strip() or None
        note_text = request.form.get("note", "").strip() or None
        return_to = request.form.get("return_to", "detail")
        if not note_text:
            return _render_note_add_form(
                defect_id, defect.get("solman_name"), return_to,
                heading=heading or "", error="Note text is required.",
            )
        database.add_note(conn, "defect", defect_id, heading, note_text)
    finally:
        conn.close()
    if return_to == "list":
        return redirect(url_for("defects_list", note_added="1"))
    return redirect(url_for("defect_detail", defect_id=defect_id, note_added="1"))


@app.route("/defects/<defect_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def note_edit(defect_id: str, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "defect" or note["entity_id"] != defect_id:
            return _not_found(defect_id)
        defect = database.get_defect(conn, defect_id)
        if request.method == "POST":
            heading = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                return redirect(url_for("defect_detail", defect_id=defect_id, note_saved="1"))
    finally:
        conn.close()
    solman_name  = defect.get("solman_name") if defect else None
    entity_label = defect_id + (f" — {solman_name}" if solman_name else "")
    kwargs = dict(
        mode="edit", entity_label=entity_label,
        list_url=url_for("defects_list"), list_label="Defects",
        detail_url=url_for("defect_detail", defect_id=defect_id),
        action_url=url_for("note_edit", defect_id=defect_id, note_id=note_id),
        cancel_url=url_for("defect_detail", defect_id=defect_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/defects/<defect_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def note_delete(defect_id: str, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "defect" or note["entity_id"] != defect_id:
            return _not_found(defect_id)
        if request.method == "POST":
            database.delete_note(conn, note_id)
            return redirect(url_for("defect_detail", defect_id=defect_id, note_deleted="1"))
    finally:
        conn.close()
    return render_template(
        "note_confirm_delete.html", note=note,
        entity_label=defect_id,
        cancel_url=url_for("defect_detail", defect_id=defect_id),
        delete_url=url_for("note_delete", defect_id=defect_id, note_id=note_id),
    )


# ---------------------------------------------------------------------------
# Retail detail + note routes
# ---------------------------------------------------------------------------

@app.route("/retail/<int:retail_id>", methods=["GET", "POST"])
def retail_detail(retail_id: int):
    conn = _get_conn()
    try:
        row = database.get_retail_by_id(conn, retail_id)
        if row is None:
            conn.close()
            return render_template("404.html", defect_id=str(retail_id)), 404
        if request.method == "POST":
            next_step       = request.form.get("next_step", "").strip() or None
            comment_history = request.form.get("comment_history", "").strip() or None
            action_needed   = 1 if request.form.get("action_needed") == "1" else 0
            database.upsert_retail_annotation(conn, retail_id, next_step, comment_history, action_needed)
            conn.close()
            return redirect(url_for("retail_detail", retail_id=retail_id, saved="1"))
        notes = database.list_notes(conn, "retail", retail_id)
        attachments_by_note = database.get_attachments_for_notes(
            conn, [n["id"] for n in notes]
        )
    finally:
        conn.close()
    saved        = request.args.get("saved") == "1"
    note_added   = request.args.get("note_added") == "1"
    note_saved   = request.args.get("note_saved") == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    added_to_meeting = request.args.get("added_to_meeting") == "1"
    return render_template(
        "retail_detail.html",
        row=row, notes=notes,
        attachments_by_note=attachments_by_note,
        saved=saved, note_added=note_added,
        note_saved=note_saved, note_deleted=note_deleted,
        added_to_meeting=added_to_meeting,
        meetings=database.MEETING_OPTIONS,
    )


@app.route("/retail/<int:retail_id>/notes/add")
def retail_note_add_form(retail_id: int):
    conn = _get_conn()
    try:
        row = database.get_retail_by_id(conn, retail_id)
    finally:
        conn.close()
    if row is None:
        return render_template("404.html", defect_id=str(retail_id)), 404
    entity_label = f"{row['test_case_id']} / {row['country']}"
    return render_template(
        "note_form.html", mode="add",
        entity_label=entity_label,
        list_url=url_for("retail_list"), list_label="Retail",
        detail_url=url_for("retail_detail", retail_id=retail_id),
        action_url=url_for("retail_note_add", retail_id=retail_id),
        cancel_url=url_for("retail_detail", retail_id=retail_id),
        heading="", return_to="detail", error=None,
    )


@app.route("/retail/<int:retail_id>/notes", methods=["POST"])
def retail_note_add(retail_id: int):
    conn = _get_conn()
    try:
        row = database.get_retail_by_id(conn, retail_id)
        if row is None:
            conn.close()
            return render_template("404.html", defect_id=str(retail_id)), 404
        heading   = request.form.get("heading", "").strip() or None
        note_text = request.form.get("note", "").strip() or None
        if not note_text:
            entity_label = f"{row['test_case_id']} / {row['country']}"
            conn.close()
            return render_template(
                "note_form.html", mode="add",
                entity_label=entity_label,
                list_url=url_for("retail_list"), list_label="Retail",
                detail_url=url_for("retail_detail", retail_id=retail_id),
                action_url=url_for("retail_note_add", retail_id=retail_id),
                cancel_url=url_for("retail_detail", retail_id=retail_id),
                heading=heading or "", return_to="detail",
                error="Note text is required.",
            )
        database.add_note(conn, "retail", retail_id, heading, note_text)
    finally:
        conn.close()
    return redirect(url_for("retail_detail", retail_id=retail_id, note_added="1"))


@app.route("/retail/<int:retail_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def retail_note_edit(retail_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "retail" or note["entity_id"] != str(retail_id):
            conn.close()
            return render_template("404.html", defect_id=str(retail_id)), 404
        row = database.get_retail_by_id(conn, retail_id)
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                conn.close()
                return redirect(url_for("retail_detail", retail_id=retail_id, note_saved="1"))
    finally:
        conn.close()
    entity_label = f"{row['test_case_id']} / {row['country']}" if row else str(retail_id)
    kwargs = dict(
        mode="edit", entity_label=entity_label,
        list_url=url_for("retail_list"), list_label="Retail",
        detail_url=url_for("retail_detail", retail_id=retail_id),
        action_url=url_for("retail_note_edit", retail_id=retail_id, note_id=note_id),
        cancel_url=url_for("retail_detail", retail_id=retail_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/retail/<int:retail_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def retail_note_delete(retail_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "retail" or note["entity_id"] != str(retail_id):
            conn.close()
            return render_template("404.html", defect_id=str(retail_id)), 404
        if request.method == "POST":
            database.delete_note(conn, note_id)
            conn.close()
            return redirect(url_for("retail_detail", retail_id=retail_id, note_deleted="1"))
    finally:
        conn.close()
    return render_template(
        "note_confirm_delete.html", note=note,
        cancel_url=url_for("retail_detail", retail_id=retail_id),
        delete_url=url_for("retail_note_delete", retail_id=retail_id, note_id=note_id),
        entity_label=f"retail #{retail_id}",
    )


# ---------------------------------------------------------------------------
# Meeting prep
# ---------------------------------------------------------------------------

@app.route("/meeting-prep")
def meeting_prep_list():
    meeting_filter = request.args.get("meeting", "")
    status_filter  = request.args.get("status", "planned")
    conn = _get_conn()
    try:
        items = database.get_meeting_prep(
            conn,
            meeting=meeting_filter or None,
            status=status_filter or None,
        )
    finally:
        conn.close()
    return render_template(
        "meeting_prep.html",
        items=items,
        meetings=database.MEETING_OPTIONS,
        overall_topics=database.MEETING_OVERALL_TOPICS,
        meeting_filter=meeting_filter,
        status_filter=status_filter,
    )


@app.route("/meeting-prep/agenda")
def meeting_prep_agenda():
    meeting_filter = request.args.get("meeting", "")
    status_filter  = request.args.get("status", "planned")
    conn = _get_conn()
    try:
        items = database.get_meeting_prep(
            conn,
            meeting=meeting_filter or None,
            status=status_filter or None,
        )
    finally:
        conn.close()
    order = {t: i for i, t in enumerate(database.MEETING_OVERALL_TOPICS)}
    items.sort(key=lambda r: (order.get(r.get("overall_topic") or "", 999), r.get("id", 0)))
    return render_template(
        "meeting_agenda.html",
        items=items,
        meeting_filter=meeting_filter,
        status_filter=status_filter,
        today=date.today().strftime("%d %B %Y"),
        overall_topics=database.MEETING_OVERALL_TOPICS,
    )


@app.route("/meeting-prep/dtco2c-daily")
def dtco2c_daily_report():
    conn = _get_conn()
    try:
        topics    = database.get_meeting_prep(conn, meeting="DTC O2C Daily", status="planned")
        defects   = database.list_daily_defects(conn)
        followups = database.get_followups(conn, with_whom="DTC O2C", include_done=False)
    finally:
        conn.close()
    return render_template(
        "dtco2c_daily_report.html",
        topics=topics,
        defects=defects,
        followups=followups,
        today=date.today().strftime("%d %B %Y"),
    )


@app.route("/meeting-prep/add", methods=["POST"])
def meeting_prep_add():
    meeting             = request.form.get("meeting", "").strip()
    topic               = request.form.get("topic", "").strip()
    source_entity_type  = request.form.get("source_entity_type", "").strip() or None
    source_entity_id    = request.form.get("source_entity_id", "").strip() or None
    overall_topic       = request.form.get("overall_topic", "").strip() or None
    if meeting and topic:
        conn = _get_conn()
        try:
            database.add_meeting_prep(conn, meeting, topic,
                                      source_entity_type, source_entity_id,
                                      overall_topic)
        finally:
            conn.close()
    back = request.form.get("back_url")
    if back:
        return redirect(back + "?added_to_meeting=1")
    return redirect(url_for("meeting_prep_list",
                            meeting=request.form.get("meeting_filter", "")))


@app.route("/meeting-prep/<int:item_id>/status", methods=["POST"])
def meeting_prep_status(item_id: int):
    status = request.form.get("status", "planned")
    conn = _get_conn()
    try:
        database.set_meeting_prep_status(conn, item_id, status)
    finally:
        conn.close()
    return jsonify({"ok": True, "status": status})


@app.route("/meeting-prep/<int:item_id>/note", methods=["POST"])
def meeting_prep_note(item_id: int):
    note = request.form.get("note", "").strip()
    conn = _get_conn()
    try:
        database.set_meeting_prep_note(conn, item_id, note)
    finally:
        conn.close()
    return jsonify({"ok": True, "note": note})


@app.route("/meeting-prep/<int:item_id>/notes")
def meeting_prep_notes(item_id: int):
    conn = _get_conn()
    try:
        notes = database.list_notes(conn, "meeting_prep", str(item_id))
    finally:
        conn.close()
    return jsonify(notes)


@app.route("/meeting-prep/<int:item_id>/notes/add", methods=["POST"])
def meeting_prep_note_add(item_id: int):
    note = request.form.get("note", "").strip()
    if not note:
        return jsonify({"ok": False, "error": "empty"})
    conn = _get_conn()
    try:
        database.add_note(conn, "meeting_prep", str(item_id), None, note)
        notes = database.list_notes(conn, "meeting_prep", str(item_id))
    finally:
        conn.close()
    return jsonify({"ok": True, "notes": notes})


@app.route("/meeting-prep/<int:item_id>/topic", methods=["POST"])
def meeting_prep_topic(item_id: int):
    topic = request.form.get("topic", "").strip()
    if not topic:
        return jsonify({"ok": False, "error": "empty"})
    conn = _get_conn()
    try:
        database.set_meeting_prep_topic(conn, item_id, topic)
    finally:
        conn.close()
    return jsonify({"ok": True, "topic": topic})


@app.route("/meeting-prep/<int:item_id>/overall_topic", methods=["POST"])
def meeting_prep_overall_topic(item_id: int):
    overall_topic = request.form.get("overall_topic", "").strip() or None
    conn = _get_conn()
    try:
        database.set_meeting_prep_overall_topic(conn, item_id, overall_topic)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/meeting-prep/<int:item_id>/delete", methods=["POST"])
def meeting_prep_delete(item_id: int):
    conn = _get_conn()
    try:
        database.delete_meeting_prep(conn, item_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# To-do list
# ---------------------------------------------------------------------------

@app.route("/todos")
def todo_list():
    f_area    = request.args.get("area", "")
    f_status  = request.args.get("status", "")
    f_prio    = request.args.get("priority", "")
    f_whom    = request.args.get("for_whom", "")
    f_due     = request.args.get("due_date", "")
    f_closed  = request.args.get("closed", "") == "1"
    conn = _get_conn()
    try:
        items   = database.get_todos(conn,
                    area=f_area or None, status=f_status or None,
                    priority=f_prio or None, for_whom=f_whom or None,
                    due_date=f_due or None, include_closed=f_closed)
        opts    = database.get_todo_filter_options(conn)
    finally:
        conn.close()
    today = date.today().isoformat()
    return render_template("todo_list.html",
        items=items, opts=opts, today=today,
        statuses=database.TODO_STATUSES,
        priorities=database.TODO_PRIORITIES,
        f_area=f_area, f_status=f_status, f_prio=f_prio,
        f_whom=f_whom, f_due=f_due, f_closed=f_closed)


@app.route("/todos/add", methods=["POST"])
def todo_add():
    topic    = request.form.get("topic", "").strip()
    area     = request.form.get("area", "").strip()
    kind     = request.form.get("kind", "").strip()
    priority = request.form.get("priority", "Medium")
    due_date = request.form.get("due_date", "").strip()
    for_whom = request.form.get("for_whom", "").strip()
    if topic:
        conn = _get_conn()
        try:
            database.add_todo(conn, area, kind, topic, priority, due_date, for_whom)
        finally:
            conn.close()
    return redirect(url_for("todo_list"))


@app.route("/todos/<int:todo_id>/edit", methods=["POST"])
def todo_edit(todo_id: int):
    topic    = request.form.get("topic", "").strip()
    area     = request.form.get("area", "").strip()
    kind     = request.form.get("kind", "").strip()
    priority = request.form.get("priority", "Medium")
    due_date = request.form.get("due_date", "").strip()
    for_whom = request.form.get("for_whom", "").strip()
    if topic:
        conn = _get_conn()
        try:
            database.update_todo(conn, todo_id, area, kind, topic, priority, due_date, for_whom)
        finally:
            conn.close()
    return jsonify({"ok": True})


@app.route("/todos/<int:todo_id>/status", methods=["POST"])
def todo_status(todo_id: int):
    status = request.form.get("status", "open")
    conn = _get_conn()
    try:
        database.set_todo_status(conn, todo_id, status)
    finally:
        conn.close()
    return jsonify({"ok": True, "status": status})


@app.route("/todos/<int:todo_id>/notes")
def todo_notes(todo_id: int):
    conn = _get_conn()
    try:
        notes = database.list_notes(conn, "todo", str(todo_id))
    finally:
        conn.close()
    return jsonify(notes)


@app.route("/todos/<int:todo_id>/notes/add", methods=["POST"])
def todo_note_add(todo_id: int):
    note = request.form.get("note", "").strip()
    if not note:
        return jsonify({"ok": False, "error": "empty"})
    conn = _get_conn()
    try:
        database.add_note(conn, "todo", str(todo_id), None, note)
        notes = database.list_notes(conn, "todo", str(todo_id))
    finally:
        conn.close()
    return jsonify({"ok": True, "notes": notes})


# ---------------------------------------------------------------------------
# Follow-ups
# ---------------------------------------------------------------------------

@app.route("/followups")
def followup_list():
    f_whom      = request.args.getlist("with_whom")
    f_group     = request.args.getlist("group_name")
    f_when      = request.args.get("when_next", "")
    f_done      = request.args.get("done", "") == "1"
    conn = _get_conn()
    try:
        items = database.get_followups(conn,
                    with_whom=f_whom or None,
                    when_next=f_when or None,
                    group_name=f_group or None,
                    include_done=f_done)
        opts  = database.get_followup_filter_options(conn)
    finally:
        conn.close()
    today = date.today().isoformat()
    return render_template("followup_list.html",
        items=items, opts=opts, today=today,
        statuses=database.FOLLOWUP_STATUSES,
        f_whom=f_whom, f_group=f_group, f_when=f_when, f_done=f_done)


@app.route("/followups/add", methods=["POST"])
def followup_add():
    with_whom  = request.form.get("with_whom", "").strip()
    topic      = request.form.get("topic", "").strip()
    when_next  = request.form.get("when_next", "").strip() or date.today().isoformat()
    group_name = request.form.get("group_name", "").strip() or None
    if with_whom and topic:
        conn = _get_conn()
        try:
            database.add_followup(conn, with_whom, topic, when_next, group_name)
        finally:
            conn.close()
    return redirect(url_for("followup_list"))


@app.route("/followups/<int:followup_id>/edit", methods=["POST"])
def followup_edit(followup_id: int):
    with_whom  = request.form.get("with_whom", "").strip()
    topic      = request.form.get("topic", "").strip()
    when_next  = request.form.get("when_next", "").strip() or None
    group_name = request.form.get("group_name", "").strip() or None
    if with_whom and topic:
        conn = _get_conn()
        try:
            database.update_followup(conn, followup_id, with_whom, topic, when_next, group_name)
        finally:
            conn.close()
    return redirect(url_for("followup_list"))


@app.route("/followups/<int:followup_id>/delete", methods=["POST"])
def followup_delete(followup_id: int):
    conn = _get_conn()
    try:
        database.delete_followup(conn, followup_id)
    finally:
        conn.close()
    return redirect(url_for("followup_list"))


@app.route("/followups/<int:followup_id>/status", methods=["POST"])
def followup_status(followup_id: int):
    status = request.form.get("status", "open")
    conn = _get_conn()
    try:
        database.set_followup_status(conn, followup_id, status)
    finally:
        conn.close()
    return jsonify({"ok": True, "status": status})


@app.route("/followups/<int:followup_id>")
def followup_detail(followup_id: int):
    note_added   = request.args.get("note_added") == "1"
    note_saved   = request.args.get("note_saved") == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    conn = _get_conn()
    try:
        row = database.get_followup_by_id(conn, followup_id)
        if row is None:
            return render_template("404.html"), 404
        notes = database.list_notes(conn, "followup", str(followup_id))
        attachments_by_note = database.get_attachments_for_notes(conn, [n["id"] for n in notes])
    finally:
        conn.close()
    return render_template(
        "followup_detail.html", row=row,
        notes=notes, attachments_by_note=attachments_by_note,
        note_added=note_added, note_saved=note_saved, note_deleted=note_deleted,
        statuses=database.FOLLOWUP_STATUSES,
    )


@app.route("/followups/<int:followup_id>/notes/add", methods=["GET", "POST"])
def followup_note_add(followup_id: int):
    conn = _get_conn()
    try:
        row = database.get_followup_by_id(conn, followup_id)
        if row is None:
            return render_template("404.html"), 404
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.add_note(conn, "followup", str(followup_id), heading, note_text)
                return redirect(url_for("followup_detail", followup_id=followup_id, note_added="1"))
    finally:
        conn.close()
    label = f"{row['with_whom']} — {row['topic']}" if row else str(followup_id)
    return render_template(
        "note_form.html", mode="add",
        entity_label=label, list_label="Follow-ups",
        list_url=url_for("followup_list"),
        detail_url=url_for("followup_detail", followup_id=followup_id),
        action_url=url_for("followup_note_add", followup_id=followup_id),
        cancel_url=url_for("followup_detail", followup_id=followup_id),
        heading="", note_text="", error=None,
    )


@app.route("/followups/<int:followup_id>/notes/<int:note_id>/edit", methods=["GET", "POST"])
def followup_note_edit(followup_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "followup" or note["entity_id"] != str(followup_id):
            return render_template("404.html"), 404
        row = database.get_followup_by_id(conn, followup_id)
        if request.method == "POST":
            heading   = request.form.get("heading", "").strip() or None
            note_text = request.form.get("note", "").strip() or None
            if note_text:
                database.update_note(conn, note_id, heading, note_text)
                return redirect(url_for("followup_detail", followup_id=followup_id, note_saved="1"))
    finally:
        conn.close()
    label = f"{row['with_whom']} — {row['topic']}" if row else str(followup_id)
    kwargs = dict(
        mode="edit", entity_label=label, list_label="Follow-ups",
        list_url=url_for("followup_list"),
        detail_url=url_for("followup_detail", followup_id=followup_id),
        action_url=url_for("followup_note_edit", followup_id=followup_id, note_id=note_id),
        cancel_url=url_for("followup_detail", followup_id=followup_id),
        created_at=note["created_at"],
    )
    if request.method == "POST":
        return render_template("note_form.html", **kwargs, heading=heading or "", note_text="", error="Note text is required.")
    return render_template("note_form.html", **kwargs, heading=note["heading"] or "", note_text=note["note"] or "")


@app.route("/followups/<int:followup_id>/notes/<int:note_id>/delete", methods=["GET", "POST"])
def followup_note_delete(followup_id: int, note_id: int):
    conn = _get_conn()
    try:
        note = database.get_note(conn, note_id)
        if note is None or note["entity_type"] != "followup" or note["entity_id"] != str(followup_id):
            return render_template("404.html"), 404
        if request.method == "POST":
            database.delete_note(conn, note_id)
            return redirect(url_for("followup_detail", followup_id=followup_id, note_deleted="1"))
    finally:
        conn.close()
    return render_template(
        "note_confirm_delete.html", note=note,
        entity_label=f"Follow-up #{followup_id}",
        cancel_url=url_for("followup_detail", followup_id=followup_id),
        delete_url=url_for("followup_note_delete", followup_id=followup_id, note_id=note_id),
    )


# ---------------------------------------------------------------------------
# Enhancements (floating panel, JSON API)
# ---------------------------------------------------------------------------

@app.route("/enhancements")
def enhancements_list():
    area           = request.args.get("area", "")
    priority       = request.args.get("priority", "")
    status         = request.args.get("status", "")
    include_closed = request.args.get("closed", "") == "1"
    conn = _get_conn()
    try:
        items = database.get_enhancements(
            conn,
            area=area or None,
            priority=priority or None,
            status=status or None,
            include_closed=include_closed,
        )
        areas = database.get_enhancement_areas(conn)
    finally:
        conn.close()
    return jsonify({
        "items": items,
        "areas": areas,
        "priorities": database.ENHANCEMENT_PRIORITIES,
    })


@app.route("/enhancements/add", methods=["POST"])
def enhancements_add():
    area        = request.form.get("area", "").strip()
    enhancement = request.form.get("enhancement", "").strip()
    priority    = request.form.get("priority", "Medium")
    if not enhancement:
        return jsonify({"ok": False, "error": "enhancement required"})
    conn = _get_conn()
    try:
        new_id = database.add_enhancement(conn, area, enhancement, priority)
    finally:
        conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/enhancements/<int:item_id>/status", methods=["POST"])
def enhancements_status(item_id: int):
    status = request.form.get("status", "not_started")
    conn = _get_conn()
    try:
        database.set_enhancement_status(conn, item_id, status)
    finally:
        conn.close()
    return jsonify({"ok": True, "status": status})


@app.route("/enhancements/page")
def enhancements_page():
    include_closed = request.args.get("closed", "") == "1"
    sort  = request.args.get("sort", "priority")
    dirn  = request.args.get("dir", "asc")
    conn = _get_conn()
    try:
        items = database.get_enhancements(conn, include_closed=include_closed)
    finally:
        conn.close()
    prio_order = {"High": 0, "Medium": 1, "Low": 2}
    status_order = {"not_started": 0, "in_progress": 1, "closed": 2}
    reverse = dirn == "desc"
    if sort == "priority":
        items.sort(key=lambda r: prio_order.get(r["priority"], 9), reverse=reverse)
    elif sort == "status":
        items.sort(key=lambda r: status_order.get(r["status"], 9), reverse=reverse)
    elif sort == "area":
        items.sort(key=lambda r: (r["area"] or "").lower(), reverse=reverse)
    return render_template(
        "enhancements_page.html",
        items=items,
        include_closed=include_closed,
        priorities=database.ENHANCEMENT_PRIORITIES,
        sort=sort,
        dirn=dirn,
    )


@app.route("/enhancements/<int:item_id>/delete", methods=["POST"])
def enhancements_delete(item_id: int):
    conn = _get_conn()
    try:
        database.delete_enhancement(conn, item_id)
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/enhancements/<int:item_id>/update", methods=["POST"])
def enhancements_update(item_id: int):
    area        = request.form.get("area", "").strip()
    enhancement = request.form.get("enhancement", "").strip()
    priority    = request.form.get("priority", "Medium")
    if not enhancement:
        return jsonify({"ok": False, "error": "enhancement text required"})
    conn = _get_conn()
    try:
        database.update_enhancement(conn, item_id, area, enhancement, priority)
    finally:
        conn.close()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Inbox — daily capture notes (entity_type='input', entity_id='inbox')
# ---------------------------------------------------------------------------

@app.route("/inbox")
def inbox():
    conn = _get_conn()
    try:
        items = database.list_inbox_items(conn)
        attachments_by_note = database.get_attachments_for_notes(conn, [n["id"] for n in items])
    finally:
        conn.close()
    note_added   = request.args.get("note_added") == "1"
    note_saved   = request.args.get("note_saved") == "1"
    note_deleted = request.args.get("note_deleted") == "1"
    note_filed   = request.args.get("note_filed")
    return render_template(
        "inbox.html", items=items, attachments_by_note=attachments_by_note,
        note_added=note_added, note_saved=note_saved,
        note_deleted=note_deleted, note_filed=note_filed,
    )


@app.route("/inbox/add", methods=["POST"])
def inbox_add():
    heading   = request.form.get("heading", "").strip() or None
    note_text = request.form.get("note", "").strip() or None
    if note_text:
        conn = _get_conn()
        try:
            database.add_inbox_item(conn, heading, note_text)
        finally:
            conn.close()
    return redirect(url_for("inbox", note_added="1"))


@app.route("/inbox/<int:note_id>/edit", methods=["POST"])
def inbox_edit(note_id: int):
    heading   = request.form.get("heading", "").strip() or None
    note_text = request.form.get("note", "").strip() or None
    if note_text:
        conn = _get_conn()
        try:
            note = database.get_note(conn, note_id)
            if note and note["entity_type"] == "input":
                database.update_note(conn, note_id, heading, note_text)
        finally:
            conn.close()
    return redirect(url_for("inbox", note_saved="1"))


@app.route("/inbox/<int:note_id>/delete", methods=["POST"])
def inbox_delete(note_id: int):
    conn = _get_conn()
    try:
        filenames = database.delete_inbox_item(conn, note_id)
    finally:
        conn.close()
    for fname in filenames:
        fp = _UPLOAD_FOLDER / fname
        if fp.exists():
            fp.unlink()
    return redirect(url_for("inbox", note_deleted="1"))


@app.route("/inbox/<int:note_id>/file", methods=["POST"])
def inbox_file(note_id: int):
    target_type = request.form.get("target_type", "").strip()
    target_id   = request.form.get("target_id", "").strip()
    conn = _get_conn()
    try:
        ok = database.file_inbox_item(conn, note_id, target_type, target_id)
    finally:
        conn.close()
    if ok:
        return redirect(url_for("inbox", note_filed=target_type))
    return redirect(url_for("inbox"))


@app.route("/inbox/targets")
def inbox_targets():
    target_type = request.args.get("type", "").strip()
    q           = request.args.get("q", "").strip()
    conn = _get_conn()
    try:
        results = database.search_targets(conn, target_type, q)
    finally:
        conn.close()
    return jsonify(results)


# ---------------------------------------------------------------------------
# Attachments (screenshots linked to notes)
# ---------------------------------------------------------------------------

@app.route("/uploads/<path:filename>")
def serve_upload(filename: str):
    return send_from_directory(str(_UPLOAD_FOLDER), filename)


@app.route("/notes/<int:note_id>/attachments/add", methods=["POST"])
def attachment_add(note_id: int):
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "no file"})
    ext = Path(f.filename).suffix.lower()
    if ext not in _ALLOWED_EXTS:
        return jsonify({"ok": False, "error": f"File type {ext!r} not allowed. Use PNG, JPG, GIF or WEBP."})
    _UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
    from datetime import datetime as _dt
    timestamp = _dt.now().strftime("%Y%m%d%H%M%S")
    filename = f"{note_id}_{timestamp}_{secure_filename(f.filename)}"
    f.save(str(_UPLOAD_FOLDER / filename))
    conn = _get_conn()
    try:
        att = database.add_attachment(conn, note_id, filename, f.filename)
    finally:
        conn.close()
    return jsonify({"ok": True, "attachment": att})


@app.route("/notes/<int:note_id>/attachments/<int:attachment_id>/delete", methods=["POST"])
def attachment_delete(note_id: int, attachment_id: int):
    conn = _get_conn()
    try:
        filename = database.delete_attachment(conn, attachment_id)
    finally:
        conn.close()
    if filename:
        fp = _UPLOAD_FOLDER / filename
        if fp.exists():
            fp.unlink()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import threading
    import webbrowser

    threading.Timer(1.2, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    app.run(debug=False, host="127.0.0.1", port=5000)
