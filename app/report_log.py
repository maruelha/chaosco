"""Shared report-log workbook writer — one dated row per saved report.

All bucket reports (Retail, ECOM, Manual Retail, Manual ECOM) share the same
column layout; each report writes to its own sheet of the workbook at
cfg['retail_report_xlsx'].
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

REPORT_LOG_HEADERS = [
    "Date",
    "Back with Sales",
    "With DTC",
    "In Progress with DTC",
    "Passed with DTC",
    "Incoming (Gatekeeper)",
    "Ready for validation",
    "In Progress",
    "In Clarification",
    "Blocked",
]


def append_report_row(cfg: dict, report: dict, day: str, sheet_name: str) -> str:
    """Append one dated bucket row to `sheet_name` of the report log workbook.

    Creates the workbook/sheet/header row on first use. Returns the file path.
    """
    xlsx_path = Path(cfg.get("retail_report_xlsx", "output/retail_report_log.xlsx"))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default blank sheet openpyxl creates
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    if ws.cell(1, 1).value is None:
        for col, header in enumerate(REPORT_LOG_HEADERS, 1):
            ws.cell(row=1, column=col).value = header

    b = report["buckets"]
    next_row = ws.max_row + 1
    for col, val in enumerate([
        day,
        b["back_with_sales"],
        b["with_dtc"],
        b["in_progress_with_dtc"],
        b["passed_with_dtc"],
        b["incoming_gatekeeper"],
        b["ready_for_validation"],
        b["in_progress"],
        b["in_clarification"],
        b["blocked"],
    ], 1):
        ws.cell(row=next_row, column=col).value = val

    wb.save(xlsx_path)
    return str(xlsx_path)
