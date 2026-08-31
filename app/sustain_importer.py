"""Core South Sustainphase Monitoring importer (build plan step 2,
2026-08-27).

Parses the daily GBS Operations checklist workbook (file picked in the
browser, suffix-matched `DTC_GBS Operations_checklist.xlsx` — see
docs/claude/sustain.md) and writes sustain_tasks/sustain_task_details via
db_sustain.replace_day_stream, one call per `Retail_<date>`/`eCom_<date>`
tab. Loaded with data_only=True: we import the cached cell values but
never the workbook's aggregations — those are recomputed in
app/db/sustain.py. Parent task = row with a Task ID; detail row = outline
level ≥ 1 under the last parent.

The header row is LOCATED, not hardcoded [2026-08-31]: the September file
dropped the "Duplicate this sheet…" instruction line, moving the header
from row 6 to row 5, and added column M "Comments/Observations". Both
layouts import; the Comments column is likewise found by header name and
is simply absent (None) in the older one.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from app.db import sustain as db_sustain

_TAB_RE = re.compile(r"^(Retail|eCom)_(\d{4}-\d{2}-\d{2})$", re.IGNORECASE)
_STREAM_CANON = {"retail": db_sustain.STREAM_RETAIL,
                 "ecom": db_sustain.STREAM_ECOM}

# The header row moved 6 -> 5 between file versions, so scan for it
# instead of hardcoding. Column M "Comments/Observations" is new and is
# located by header name (absent in pre-2026-08-31 files).
_HEADER_SEARCH_ROWS = 12
_LAST_FIXED_COLUMN = 12          # A..L, stable across both layouts
_COMMENTS_HEADER = "comment"     # substring match, case-insensitive


class ParseError(Exception):
    """Raised when the workbook cannot be read or a day tab does not have
    the expected checklist structure."""


def _clean(val):
    """None stays None; strings stripped (blank -> None); numbers to their
    shortest text form (Task IDs arrive as text OR numbers)."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val or None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val)


def _find_header_row(ws) -> int:
    """Row whose column A is 'Task ID' — row 6 in the pre-2026-08-31
    files, row 5 since the instruction line was dropped."""
    for r in range(1, _HEADER_SEARCH_ROWS + 1):
        if _clean(ws.cell(r, 1).value) == "Task ID":
            return r
    raise ParseError(
        f"Tab '{ws.title}' looks like a checklist day tab by name but no"
        f" 'Task ID' header was found in column A of rows"
        f" 1-{_HEADER_SEARCH_ROWS} — structure changed?")


def _find_comments_column(ws, header_row: int):
    """Column index of 'Comments/Observations' (new 2026-08-31), or None
    for the older layout that doesn't have it."""
    for c in range(_LAST_FIXED_COLUMN + 1, ws.max_column + 1):
        header = _clean(ws.cell(header_row, c).value)
        if header and _COMMENTS_HEADER in header.casefold():
            return c
    return None


def _parse_sheet(ws) -> list[dict]:
    """One day tab -> parent-task dicts with their 'details' attached."""
    header_row = _find_header_row(ws)
    comments_col = _find_comments_column(ws, header_row)
    tasks: list[dict] = []
    current: dict | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        values = [_clean(ws.cell(r, c).value)
                  for c in range(1, _LAST_FIXED_COLUMN + 1)]
        comments = (_clean(ws.cell(r, comments_col).value)
                    if comments_col else None)
        if not any(values) and comments is None:
            continue
        (task_id, taxonomy, process, cadence, due_today, country, provider,
         result_fr, result_it, result_pt, result_es, overall) = values
        dim = ws.row_dimensions.get(r)
        outline_level = dim.outline_level if dim is not None else 0
        row = {
            "excel_row": r, "cadence": cadence, "due_today": due_today,
            "country": country, "provider": provider,
            "result_fr": result_fr, "result_it": result_it,
            "result_pt": result_pt, "result_es": result_es,
            "overall": overall, "comments": comments,
        }
        if task_id is not None:
            current = {**row, "task_id": task_id, "taxonomy": taxonomy,
                       "process": process, "details": []}
            tasks.append(current)
        elif outline_level >= 1 and current is not None:
            current["details"].append(row)
        # level-0 rows without a Task ID (stray notes) are dropped
    return tasks


def parse_sustain_workbook(xlsx_path: Path) -> list[dict]:
    """Return tab dicts {'day', 'stream', 'tasks'} for every
    Retail_<date>/eCom_<date> tab, in workbook order. Non-matching tabs
    are ignored; a matching tab with the wrong structure raises
    ParseError."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        raise ParseError(f"Could not read workbook: {exc}") from exc
    try:
        tabs: list[dict] = []
        for ws in wb.worksheets:
            m = _TAB_RE.match(ws.title.strip())
            if not m:
                continue
            tabs.append({
                "day": m.group(2),
                "stream": _STREAM_CANON[m.group(1).lower()],
                "tasks": _parse_sheet(ws),
            })
        return tabs
    finally:
        wb.close()


def run_sustain_import(cfg: dict, xlsx_path: Path) -> dict:
    """Parse + replace each contained (day, stream) tab. Result dict
    mirrors run_smoke_import's shape: ok/error plus counts."""
    result: dict = {"ok": False, "error": None, "xlsx_path": str(xlsx_path),
                    "tabs": 0, "tasks": 0, "details": 0}
    try:
        tabs = parse_sustain_workbook(xlsx_path)
    except ParseError as exc:
        result["error"] = str(exc)
        return result
    if not tabs:
        result["error"] = ("No Retail_<date>/eCom_<date> day tabs found —"
                           " is this the right workbook?")
        return result

    from app import database
    db_path = Path(cfg["database_path"])
    db_sustain.init_schema(db_path)
    conn = database.get_connection(db_path)
    try:
        for tab in tabs:
            counts = db_sustain.replace_day_stream(
                conn, tab["day"], tab["stream"], tab["tasks"])
            result["tabs"] += 1
            result["tasks"] += counts["tasks"]
            result["details"] += counts["details"]
    finally:
        conn.close()
    result["ok"] = True
    return result
