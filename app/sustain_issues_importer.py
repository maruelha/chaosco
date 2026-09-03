"""Sustainphase Issues importer (rewritten 2026-09-03 [USER]).

Parses the **Go-Live defect tracker** workbook (file picked in the
browser, name-contains guard in the web layer) — three tabs, each mapped
by NORMALIZED HEADER NAME (never by position):

- "ASPEN Incidents"        → upsert_incidents (key: Incident Number; rows
                              without one are SKIPPED and counted [USER])
- "Issue Solution tracker" → replace_solutions (no row identity → wholesale)
- "Total"                  → replace_interfaces (the interface list only:
                              header row located by its "Namespace" cell,
                              the sheet's own "Total Issue #" column is
                              IGNORED — totals are computed in storage)

Header rows: ASPEN Incidents + Issue Solution tracker in row 1; the Total
sheet's list header is on row 3 (row 2 carries the "Interfaces" /
"Total Issue #" group titles) — located, not assumed. Date cells arrive
as datetimes → ISO dates.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from app.db import sustain_issues as db_si

SHEET_INCIDENTS = "aspen incidents"
SHEET_SOLUTIONS = "issue solution tracker"
SHEET_TOTAL = "total"

# normalized header prefix -> field
_INCIDENT_HEADERS = [
    ("incident number", "incident_number"),
    ("date", "date"),
    ("requestor", "requestor"),
    ("title", "title"),
    ("status", "status"),
    ("assigned to", "assigned_to"),
    ("latest comment", "latest_comment"),
]
_SOLUTION_HEADERS = [
    ("owner", "owner"),
    ("interface", "interface"),
    ("msg", "msg"),
    ("text", "text"),
    ("external reference", "external_reference"),
    ("inc reference", "inc_reference"),
    ("reason", "reason"),
    ("solution", "solution"),
    ("status", "status"),
]
_INTERFACE_HEADERS = [
    ("namespace", "namespace"),
    ("interface", "interface"),
    ("version", "version"),
    ("name", "name"),
    ("variant", "variant"),
    ("index tables", "index_tables"),
]


class ParseError(Exception):
    """Raised when a tab is missing or its headers don't look like the
    Go-Live defect tracker template."""


def _norm_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _clean(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        val = re.sub(r"[ \t]+", " ", val).strip()
        return val or None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val)


def _sheet(wb, title_norm: str):
    return next((ws for ws in wb.worksheets
                 if ws.title.strip().casefold() == title_norm), None)


def _map_columns(sheet, header_row: int, header_map) -> dict[int, str]:
    col_map: dict[int, str] = {}
    for c in range(1, sheet.max_column + 1):
        header = _norm_header(sheet.cell(header_row, c).value)
        if not header:
            continue
        for prefix, name in header_map:
            if header.startswith(prefix) and name not in col_map.values():
                col_map[c] = name
                break
    return col_map


def _rows(sheet, header_row: int, col_map: dict[int, str]) -> list[dict]:
    out: list[dict] = []
    for r in range(header_row + 1, sheet.max_row + 1):
        row = {name: _clean(sheet.cell(r, c).value) for c, name in col_map.items()}
        if not any(row.values()):
            continue
        row["excel_row"] = r
        out.append(row)
    return out


def _find_header_row(sheet, marker: str, scan: int = 10) -> int | None:
    for r in range(1, min(scan, sheet.max_row) + 1):
        if any(_norm_header(sheet.cell(r, c).value) == marker
               for c in range(1, sheet.max_column + 1)):
            return r
    return None


def parse_go_live_tracker(xlsx_path: Path) -> dict:
    """{'incidents': [...], 'skipped_incidents': n, 'solutions': [...],
    'interfaces': [...]} — raises ParseError on a missing tab or headers
    that do not fit."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        raise ParseError(f"Could not read workbook: {exc}") from exc
    try:
        # --- ASPEN Incidents
        ws = _sheet(wb, SHEET_INCIDENTS)
        if ws is None:
            raise ParseError(
                f"No 'ASPEN Incidents' tab found. Sheets present: {wb.sheetnames}")
        col_map = _map_columns(ws, 1, _INCIDENT_HEADERS)
        if "incident_number" not in col_map.values():
            raise ParseError("'ASPEN Incidents' tab has no 'Incident Number'"
                             " column in row 1 — structure changed?")
        raw = _rows(ws, 1, col_map)
        incidents = [r for r in raw if r.get("incident_number")]
        skipped = len(raw) - len(incidents)

        # --- Issue Solution tracker
        ws = _sheet(wb, SHEET_SOLUTIONS)
        if ws is None:
            raise ParseError(
                f"No 'Issue Solution tracker' tab found. Sheets present: {wb.sheetnames}")
        col_map = _map_columns(ws, 1, _SOLUTION_HEADERS)
        if "interface" not in col_map.values():
            raise ParseError("'Issue Solution tracker' tab has no 'Interface'"
                             " column in row 1 — structure changed?")
        solutions = _rows(ws, 1, col_map)

        # --- Total (interface list)
        ws = _sheet(wb, SHEET_TOTAL)
        if ws is None:
            raise ParseError(
                f"No 'Total' tab found. Sheets present: {wb.sheetnames}")
        hdr = _find_header_row(ws, "namespace")
        if hdr is None:
            raise ParseError("'Total' tab: no 'Namespace' header found in the"
                             " first 10 rows — structure changed?")
        col_map = _map_columns(ws, hdr, _INTERFACE_HEADERS)
        interfaces = [r for r in _rows(ws, hdr, col_map) if r.get("interface")]
        return {"incidents": incidents, "skipped_incidents": skipped,
                "solutions": solutions, "interfaces": interfaces}
    finally:
        wb.close()


def run_sustain_issues_import(cfg: dict, xlsx_path: Path) -> dict:
    """Parse + store all three tabs. Result: ok/error plus counts —
    incidents (rows, inserted, updated, new_comments, skipped), solutions,
    interfaces. Empty tabs import fine (the template starts empty)."""
    result: dict = {"ok": False, "error": None, "xlsx_path": str(xlsx_path),
                    "incidents": 0, "inserted": 0, "updated": 0,
                    "new_comments": 0, "skipped": 0,
                    "solutions": 0, "interfaces": 0}
    try:
        parsed = parse_go_live_tracker(xlsx_path)
    except ParseError as exc:
        result["error"] = str(exc)
        return result
    from app import database
    db_path = Path(cfg["database_path"])
    db_si.init_schema(db_path)
    conn = database.get_connection(db_path)
    try:
        counts = db_si.upsert_incidents(conn, parsed["incidents"])
        result.update(counts)
        result["incidents"] = len(parsed["incidents"])
        result["skipped"] = parsed["skipped_incidents"]
        result["solutions"] = db_si.replace_solutions(conn, parsed["solutions"])
        result["interfaces"] = db_si.replace_interfaces(conn, parsed["interfaces"])
    finally:
        conn.close()
    result["ok"] = True
    return result
