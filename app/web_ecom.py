"""ECOM vertical — routes (Flask Blueprint, day plan 05.07 step 8).

Own Blueprint by design (CLAUDE.md new-module pattern). No SQL here —
storage in app/db/ecom.py; Jira context is read-only from the shared jira
store (app/db/jira.py), joined by jira id. Excel fields and Jira fields
stay strictly separate on the pages too.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from flask import Blueprint, jsonify, redirect, render_template, request, url_for

from app import database
from app.config_loader import load_config
from app.db import ecom as db_ecom
from app.db import jira as db_jira
from app.jira_importer import run_jira_import
from app.reporter import (compute_impacted_totals, compute_retail_report,
                          load_status_mappings, passed_family)
from app.row_validations import validate_rows

bp = Blueprint("ecom", __name__, url_prefix="/ecom")

_cfg = load_config()
_db_path = Path(_cfg["database_path"])


def _get_conn():
    return database.get_connection(_db_path)


@bp.route("/")
def ecom_list():
    statuses  = request.args.getlist("status")
    countries = request.args.getlist("country")
    scenarios = request.args.getlist("scenario")
    q         = request.args.get("q", "").strip() or None

    conn = _get_conn()
    try:
        rows = db_ecom.get_ecom_rows(conn, statuses=statuses or None,
                                     countries=countries or None,
                                     scenarios=scenarios or None, q=q)
        distincts = db_ecom.get_ecom_distincts(conn)
        jira_info = {i["jira_key"]: i for i in db_jira.list_jira_issues(conn)}
        jira_keys = set(jira_info)
        from app.db import teams_chats as db_tc
        chats_by_entity = db_tc.chats_by_entity(conn, "jira")
        from app.db import gatekeeper as db_gk
        track_sales_keys = db_gk.get_track_sales_keys(conn)
        # per-row expander [USER 2026-07-12]: jira comments + the notes made
        # on the gatekeeper side (entity 'jira' — same key, shared history)
        jira_comments = {r["jira_id"]: db_jira.list_jira_comments(conn, r["jira_id"])
                         for r in rows if r["jira_id"] in jira_keys}
        jira_notes = {r["jira_id"]: database.list_notes(conn, "jira", r["jira_id"])
                      for r in rows if r["jira_id"] in jira_keys}
    finally:
        conn.close()
    return render_template(
        "ecom.html", rows=rows, distincts=distincts, jira_keys=jira_keys,
        validations=validate_rows("ecom", rows, "ecom_id"),
        jira_info=jira_info, chats_by_entity=chats_by_entity,
        track_sales_keys=track_sales_keys,
        jira_comments=jira_comments, jira_notes=jira_notes,
        sel_statuses=statuses, sel_countries=countries, sel_scenarios=scenarios,
        q=q or "",
        jira_ok=request.args.get("jira_ok"),
        jira_msg=request.args.get("jira_msg"),
    )


@bp.route("/import-jira", methods=["POST"])
def ecom_import_jira():
    """'Update from Jira' — runs the ECOM-folder XML import (step 2 code)."""
    result = run_jira_import(_cfg)
    if result["ok"]:
        msg = (f"{Path(result['xml_path']).name}: {result['parsed']} in file — "
               f"{result['refreshed']} tracked refreshed · "
               f"{result['new_gatekeeper']} new (assigned to Marina) · "
               f"{result['new_board']} new (on the board) · "
               f"{result['ignored']} ignored · {result['comments']} comments")
        return redirect(url_for("ecom.ecom_list", jira_ok="1", jira_msg=msg))
    return redirect(url_for("ecom.ecom_list", jira_ok="0", jira_msg=result["error"]))


def _report_context(conn) -> dict:
    """Shared context for the report page/download: buckets (same definitions
    as Retail [USER 2026-07-09]) + impacted ECOM-channel defects."""
    mappings = load_status_mappings()
    report = compute_retail_report(db_ecom.get_ecom_status_counts(conn), mappings)
    defects = db_ecom.get_ecom_defects_impacted(conn, passed_family(mappings))
    totals = compute_impacted_totals(defects)
    return {
        "report": report,
        "impacted_defects": defects,
        "impacted_total": totals["total"],
        "mb_total": totals["mb"],
        "sales_total": totals["sales"],
        "report_comments": database.list_report_comments(conn, "ecom"),
        "today": date.today().isoformat(),
    }


@bp.route("/report")
def ecom_report():
    conn = _get_conn()
    try:
        ctx = _report_context(conn)
    finally:
        conn.close()
    return render_template("ecom_report.html", **ctx)


@bp.route("/report/download")
def ecom_report_download():
    """Dated standalone snapshot — the page itself made self-contained
    (CSS inlined, buttons/scripts stripped), same as the email attachment."""
    from app.emailer import standalone_html
    from app.web_core import app as flask_app
    resp = flask_app.test_client().get(url_for("ecom.ecom_report"))
    today = date.today().isoformat()
    return standalone_html(resp.get_data(as_text=True)), 200, {
        "Content-Type": "text/html; charset=utf-8",
        "Content-Disposition": f'attachment; filename="ecom_report_{today}.html"',
    }


_ECOM_REPORT_HEADERS = [
    "Date", "Back with Sales", "With DTC", "In Progress with DTC",
    "Passed with DTC", "Incoming (Gatekeeper)", "Ready for validation",
    "In Progress", "In Clarification", "Blocked",
]


@bp.route("/report/save-excel", methods=["POST"])
def ecom_report_save_excel():
    """Append one dated row to the ECOM sheet of the report log workbook
    (same file as the Retail log, own sheet)."""
    try:
        conn = _get_conn()
        try:
            report = compute_retail_report(
                db_ecom.get_ecom_status_counts(conn), load_status_mappings())
        finally:
            conn.close()
        save_date = request.form.get("date") or date.today().isoformat()

        xlsx_path = Path(_cfg.get("retail_report_xlsx", "output/retail_report_log.xlsx"))
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = (openpyxl.load_workbook(xlsx_path) if xlsx_path.exists()
              else openpyxl.Workbook())
        if not xlsx_path.exists() and "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb["ECOM"] if "ECOM" in wb.sheetnames else wb.create_sheet("ECOM")
        if ws.cell(1, 1).value is None:
            for col, header in enumerate(_ECOM_REPORT_HEADERS, 1):
                ws.cell(row=1, column=col).value = header
        b = report["buckets"]
        next_row = ws.max_row + 1
        for col, val in enumerate([
            save_date, b["back_with_sales"], b["with_dtc"],
            b["in_progress_with_dtc"], b["passed_with_dtc"],
            b["incoming_gatekeeper"], b["ready_for_validation"],
            b["in_progress"], b["in_clarification"], b["blocked"],
        ], 1):
            ws.cell(row=next_row, column=col).value = val
        wb.save(xlsx_path)
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)})
    return jsonify({"ok": True, "path": str(xlsx_path), "date": save_date})


@bp.route("/<int:ecom_id>", methods=["GET", "POST"])
def ecom_detail(ecom_id: int):
    conn = _get_conn()
    try:
        row = db_ecom.get_ecom_by_id(conn, ecom_id)
        if row is None:
            conn.close()
            return render_template("404.html", defect_id=f"ECOM #{ecom_id}"), 404
        if request.method == "POST":
            db_ecom.upsert_ecom_annotation(
                conn, row["jira_id"],
                next_step=request.form.get("next_step", "").strip() or None,
                comment_history=request.form.get("comment_history", "").strip() or None,
                action_needed=request.form.get("action_needed") == "1")
            conn.close()
            return redirect(url_for("ecom.ecom_detail", ecom_id=ecom_id, saved="1"))
        jira = db_jira.get_jira_issue(conn, row["jira_id"])
        jira_comments = (db_jira.list_jira_comments(conn, row["jira_id"])
                         if jira else [])
        notes = database.list_notes(conn, "ecom", ecom_id)
        attachments_by_note = database.get_attachments_for_notes(
            conn, [n["id"] for n in notes])
    finally:
        conn.close()
    return render_template(
        "ecom_detail.html",
        row=row, jira=jira, jira_comments=jira_comments,
        notes=notes, attachments_by_note=attachments_by_note,
        saved=request.args.get("saved") == "1",
        note_added=request.args.get("note_added") == "1",
        note_saved=request.args.get("note_saved") == "1",
        note_deleted=request.args.get("note_deleted") == "1",
        orders_moved=request.args.get("orders_moved"),
    )


@bp.route("/<int:ecom_id>/comment", methods=["POST"])
def ecom_comment_save(ecom_id: int):
    """Shared Comments dialog on the board (_comment_history_dialog.html)."""
    comment_history = request.form.get("comment_history", "").strip() or None
    conn = _get_conn()
    try:
        row = db_ecom.get_ecom_by_id(conn, ecom_id)
        if row is None:
            return jsonify({"ok": False, "error": "unknown row"}), 404
        db_ecom.set_ecom_comment_history(conn, row["jira_id"], comment_history)
    finally:
        conn.close()
    return jsonify({"ok": True, "comment_history": comment_history or ""})


@bp.route("/<int:ecom_id>/pull-orders", methods=["POST"])
def ecom_pull_orders(ecom_id: int):
    """LEGACY — superseded by the shared ('jira', key) order address
    (2026-07-16); the button was removed, route kept for URL stability."""
    conn = _get_conn()
    try:
        row = db_ecom.get_ecom_by_id(conn, ecom_id)
        moved = (db_ecom.relink_gatekeeper_orders(conn, row["jira_id"], ecom_id)
                 if row else 0)
    finally:
        conn.close()
    return redirect(url_for("ecom.ecom_detail", ecom_id=ecom_id, orders_moved=moved))
