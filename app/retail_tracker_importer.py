"""Retail Requirements Tracker — one-time import of the tracking Excel (tabs 1-3).

Parsing rules (verified against the real file 2026-07-04, see
retail-tracker-handoff.md decisions #8/#9):

- A row is a requirement when its test-id cell contains a GKP id; otherwise a
  non-empty Scenarios cell is a heading. Numbered headings ("1. Retail Sale")
  start a section, lettered ones ("a. normal") a sub-section; the two flatten
  into one scenario_label per requirement.
- Return tab: the countable test is the "Test ID Return" column when present,
  else the "Test ID Sale" column (exchanges, PROMAT RETURN live there). When
  both exist, the sale id is kept as informational sale_test_ref only.
- Matching to the dashboard is by the test NAME after the first underscore
  (whitespace-collapsed, case-insensitive) — the id prefixes are two unrelated
  numbering schemes. Resolution happens here, once, at import time.
- "DTC required": 18 -> all-countries sentinel; '1 or 5' -> 5 (user decision
  2026-07-04); other integers -> required_dtc; anything else is stored raw and
  flagged as needing a decision.
- Duplicates (same requirement name + same test name) across tabs are skipped
  and reported — the Excel repeats several Sales rows on the Return and
  Payment tabs.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from app import database
from app import db_retail_tracker as db

# Excel header -> canonical country name (join key vs retail.country).
# Only mappings we are sure about; everything else keeps the header name and
# can be renamed in tracker_countries later.
_COUNTRY_FIX = {"UK": "United Kingdom", "Belguim": "Belgium"}

_TOP_RE = re.compile(r"^\d+[\.\)]")
_SUB_RE = re.compile(r"^[a-z][\.\)]\s")

# There is NO payment area [USER 2026-07-04]: the Payment methods tab almost
# entirely duplicates section 8 of the Sales/Return tabs, so those tabs win the
# dedup and the payment tab's few unique rows (OFFLINE Return, Blind return)
# are folded into sales/return based on their section heading (area=None below
# means "dynamic by heading").
_TABS = [
    # (sheet name, area, sale-id col header, countable-id col header)
    ("Tracking Sales",  "sales",  "Test ID Sale", None),
    ("Tracking Return", "return", "Test ID Sale", "Test ID Return"),
    ("Payment methods", None,     "Solman ID",    None),
]

# payment-tab rows keep their own excel_row space to avoid colliding with
# (area, excel_row) keys of the tab they are folded into
_PAYMENT_TAB_ROW_OFFSET = 1000

_TAB4_SHEET = "Country Specific PaymentM"

# Tab 4 uses 2-letter codes; tabs 1-3 use names. AU = Austria (not Australia),
# BU = Bulgaria. Cyprus (CY) appears ONLY in tab 4 and is deliberately NOT added
# to tracker_countries (it would silently turn "all = 18" into 19).
_TAB4_CODE_MAP = {
    "DE": "Germany", "GB": "United Kingdom", "NO": "Norway", "FI": "Finland",
    "SE": "Sweden", "DK": "Denmark", "HU": "Hungary", "PL": "Poland",
    "CZ": "Czech", "IE": "Ireland", "CH": "Switzerland", "AU": "Austria",
    "GR": "Greece", "BU": "Bulgaria", "CY": "Cyprus", "RO": "Romania",
    "HR": "Croatia", "BE": "Belgium", "NL": "Netherlands",
}

# order of the four "Tested" columns in tab 4 (row 1 headers, row 2 test ids)
_TAB4_KINDS = ["sale_card", "sale_voucher", "return_card", "return_voucher"]

# Comment country-list codes (user decision 2026-07-04: a comment like "UK, IE"
# means THOSE specific countries must pass — currently one row; parsed
# conservatively: only when EVERY comma-separated token is a known code).
_COMMENT_CODE_MAP = {**_TAB4_CODE_MAP, "UK": "United Kingdom", "AT": "Austria"}


def parse_comment_targets(comment: str | None) -> list[str]:
    if not comment:
        return []
    tokens = [t.strip().upper() for t in comment.split(",") if t.strip()]
    if not tokens or any(t not in _COMMENT_CODE_MAP for t in tokens):
        return []
    return [_COMMENT_CODE_MAP[t] for t in tokens]


# "B2B invoice (special case PL)" -> the named country IS the requirement
# [USER 2026-07-04: that line must be tested in Poland specifically]
_SPECIAL_CASE_RE = re.compile(r"special case\s+([A-Za-z]{2})\b", re.IGNORECASE)


def parse_name_targets(name: str | None) -> list[str]:
    m = _SPECIAL_CASE_RE.search(name or "")
    code = m.group(1).upper() if m else None
    return [_COMMENT_CODE_MAP[code]] if code in _COMMENT_CODE_MAP else []


def collect_targets(recs: list[dict]) -> dict[tuple[str, int], list[str]]:
    """Named targets per (area, excel_row): comment lists ∪ name 'special
    case XX', then PROPAGATED across areas to same-named requirements
    [USER 2026-07-04: Kids articles' 'UK, IE' applies to sale AND return]."""
    by_name: dict[str, list[str]] = {}
    for r in recs:
        targets = sorted(set(parse_comment_targets(r["comment"]))
                         | set(parse_name_targets(r["name"])))
        if targets:
            by_name.setdefault(r["name"].strip().lower(), targets)
    return {(r["area"], r["excel_row"]): by_name[r["name"].strip().lower()]
            for r in recs if r["name"].strip().lower() in by_name}


def _norm(cell) -> str:
    return " ".join(str(cell).split()).strip() if cell is not None else ""


def _is_gkp(s: str) -> bool:
    return s.upper().startswith("GKP")


def _split_test(s: str) -> tuple[str | None, str | None]:
    """'GKP0005MU01_Sale of the Book - Cash' -> ('GKP0005MU01', 'sale of the book - cash')"""
    if "_" not in s:
        return (s or None), None
    ref, name = s.split("_", 1)
    name = " ".join(name.split()).lower()
    return ref.strip() or None, name or None


def _parse_required(raw: str) -> tuple[int | None, int]:
    """-> (required_dtc, all_countries). Unknown non-integers -> (None, 0)."""
    if " ".join(str(raw).split()).lower() == "1 or 5":
        return 5, 0  # user decision 2026-07-04: '1 or 5' counts as 5
    try:
        n = int(float(raw))
    except (ValueError, TypeError):
        return None, 0
    if n == 18:  # the Excel's "all countries" convention (decision #4)
        return None, 1
    return n, 0


def _header_index(headers: list[str], label: str) -> int | None:
    wanted = label.lower()
    for i, h in enumerate(headers):
        if h and wanted in h.lower():
            return i
    return None


def parse_tracking_workbook(xlsx_path: Path) -> dict:
    """Parse tabs 1-3. Returns {'requirements': [...], 'skipped_dups': [...],
    'countries': [(name, excel_header), ...], 'per_tab': {area: count}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        requirements: list[dict] = []
        skipped_dups: list[dict] = []
        countries: list[tuple[str, str]] = []
        per_tab: dict[str, int] = {}
        seen: dict[tuple, dict] = {}  # (name, test_name) -> first requirement

        for sheet, area, sale_col_label, return_col_label in _TABS:
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)
            headers = [_norm(c) for c in next(rows)]

            sale_idx = _header_index(headers, sale_col_label)
            return_idx = _header_index(headers, return_col_label) if return_col_label else None
            req_idx = _header_index(headers, "required")
            comment_idx = _header_index(headers, "comment")
            number_idx = _header_index(headers, "number")
            if sale_idx is None or req_idx is None or number_idx is None:
                raise ValueError(
                    f"Sheet '{sheet}': expected columns not found (headers: {headers})")

            if not countries:  # country columns sit after 'number'; same on all tabs
                for h in headers[number_idx + 1:]:
                    if h:
                        countries.append((_COUNTRY_FIX.get(h, h), h))

            dynamic_area = area is None  # payment tab: fold rows into sales/return
            top = sub = None
            if not dynamic_area:
                per_tab.setdefault(area, 0)
            for row_no, row in enumerate(rows, start=2):
                cells = [_norm(c) for c in row]
                scen = cells[0] if len(cells) > 0 else ""
                sale_cell = cells[sale_idx] if sale_idx < len(cells) else ""
                return_cell = cells[return_idx] if return_idx is not None and return_idx < len(cells) else ""

                # pick the countable test id (return col wins on the return tab)
                if _is_gkp(return_cell):
                    test_cell, sale_ref = return_cell, (sale_cell if _is_gkp(sale_cell) else None)
                elif _is_gkp(sale_cell):
                    test_cell, sale_ref = sale_cell, None
                else:
                    if scen:  # heading row
                        if dynamic_area or _TOP_RE.match(scen):
                            top, sub = scen, None
                        elif _SUB_RE.match(scen):
                            sub = scen
                        else:
                            sub = scen
                    continue

                # a requirement row whose name carries its own sub-heading letter
                # ("d. sales cancellation") REPLACES the current sub-heading —
                # otherwise it would inherit the previous section's scenario
                if scen and _SUB_RE.match(scen):
                    sub = scen
                    scen = re.sub(r"^[a-z][\.\)]\s*", "", scen)

                if dynamic_area:
                    row_area = "sales" if (top or "").lower().startswith("sale") else "return"
                    row_no_key = row_no + _PAYMENT_TAB_ROW_OFFSET
                    label = f"Payment methods — {top}" if top else "Payment methods"
                else:
                    row_area, row_no_key = area, row_no
                    label = f"{top} — {sub}" if (top and sub) else (top or sub)

                excel_ref, test_name = _split_test(test_cell)
                required_raw = cells[req_idx] if req_idx < len(cells) else ""
                required_dtc, all_countries = _parse_required(required_raw)
                rec = {
                    "area": row_area,
                    "excel_row": row_no_key,
                    "scenario_label": label,
                    "name": scen or (test_name or excel_ref or "?"),
                    "excel_test_ref": excel_ref,
                    "test_name": test_name,
                    "test_case_id": None,
                    "sale_test_ref": sale_ref,
                    "required_raw": required_raw or None,
                    "required_dtc": required_dtc,
                    "all_countries": all_countries,
                    "comment": (cells[comment_idx] if comment_idx is not None and comment_idx < len(cells) else "") or None,
                }

                key = (rec["name"].lower(), test_name)
                if key in seen:
                    first = seen[key]
                    skipped_dups.append({
                        "area": row_area, "excel_row": row_no, "name": rec["name"],
                        "excel_test_ref": excel_ref, "sheet": sheet,
                        "dup_of": f"{first['area']} row {first['excel_row']}",
                    })
                    continue
                seen[key] = rec
                requirements.append(rec)
                per_tab[row_area] = per_tab.get(row_area, 0) + 1

        tab4 = _parse_tab4(wb)
        return {"requirements": requirements, "skipped_dups": skipped_dups,
                "countries": countries, "per_tab": per_tab,
                "cpm_rows": tab4["cpm_rows"], "tab4_tests": tab4["tab4_tests"]}
    finally:
        wb.close()


def _parse_tab4(wb) -> dict:
    """Tab 4: one row per (country x payment method). Row 2 holds the four
    fixed test ids under the 'Tested …' headers. Category comes from an 'X'
    in the Card/Voucher column; any other text there is kept as a comment
    and the category is left NULL for a manual decision."""
    ws = wb[_TAB4_SHEET]
    rows = ws.iter_rows(values_only=True)
    next(rows)                                   # row 1: headers
    id_row = [_norm(c) for c in next(rows)]      # row 2: the four test ids

    tab4_tests = []
    gkp_cells = [c for c in id_row if _is_gkp(c)]
    for kind, cell in zip(_TAB4_KINDS, gkp_cells):
        ref, name = _split_test(cell)
        tab4_tests.append({"test_kind": kind, "excel_test_ref": ref,
                           "test_name": name, "test_case_id": None})

    cpm_rows = []
    for row_no, row in enumerate(rows, start=3):
        cells = [_norm(c) for c in row]
        code = cells[0] if len(cells) > 0 else ""
        method = cells[1] if len(cells) > 1 else ""
        if not code or not method:
            continue
        # already covered by the sales/return OFFLINE requirements [USER 2026-07-04]
        if "offline transactions" in method.lower():
            continue
        card_cell = cells[2] if len(cells) > 2 else ""
        voucher_cell = cells[3] if len(cells) > 3 else ""
        if card_cell.upper() == "X":
            category = "card"
        elif voucher_cell.upper() == "X":
            category = "voucher"
        else:
            category = None
        notes = [c for c in (card_cell, voucher_cell) if c and c.upper() != "X"]
        cpm_rows.append({
            "country": _TAB4_CODE_MAP.get(code, code),
            "method_name": method,
            "category": category,
            "comment": " / ".join(notes) or None,
            "excel_row": row_no,
        })
    return {"cpm_rows": cpm_rows, "tab4_tests": tab4_tests}


def resolve_requirements(recs: list[dict], name_lookup: dict[str, str]) -> dict:
    """Fill test_case_id from the name lookup. Returns counts + unresolved list."""
    resolved = 0
    unresolved: list[dict] = []
    for r in recs:
        if r["test_name"] and r["test_name"] in name_lookup:
            r["test_case_id"] = name_lookup[r["test_name"]]
            resolved += 1
        else:
            unresolved.append(r)
    return {"resolved": resolved, "unresolved": unresolved}


def run_tracker_import(cfg: dict) -> dict:
    """Parse + resolve + upsert. Returns a result dict for the import screen."""
    result: dict = {"ok": False, "error": None, "xlsx_path": None}
    default_path = "report_export/TRACKING_Retail ROE UAT Testcases DTC (1).xlsx"
    xlsx_path = Path(cfg.get("retail_tracker_excel", default_path))
    result["xlsx_path"] = str(xlsx_path)
    if not xlsx_path.exists():
        result["error"] = f"Tracking Excel not found: {xlsx_path}"
        return result

    try:
        parsed = parse_tracking_workbook(xlsx_path)
    except Exception as exc:
        result["error"] = f"Parse failed: {exc}"
        return result

    conn = database.get_connection(Path(cfg["database_path"]))
    try:
        lookup = db.build_retail_name_lookup(conn)
        res = resolve_requirements(parsed["requirements"], lookup)
        counts = db.upsert_requirements(conn, parsed["requirements"])
        removed = db.delete_requirements_not_in(
            conn, {(r["area"], r["excel_row"]) for r in parsed["requirements"]})
        seeded = db.seed_countries(conn, parsed["countries"])
        # named country targets: comment lists + "special case XX" names,
        # propagated across areas by requirement name
        targets_by_key = collect_targets(parsed["requirements"])
        target_rows = []
        all_reqs = db.list_requirements(conn)
        by_key = {(x["area"], x["excel_row"]): x["id"] for x in all_reqs}
        for key, targets in targets_by_key.items():
            db.replace_requirement_targets(conn, by_key[key], targets)
            target_rows.append({"area": key[0], "excel_row": key[1],
                                "targets": targets})
        # tab 4: resolve the four fixed tests by name, then upsert the matrix
        for t in parsed["tab4_tests"]:
            t["test_case_id"] = lookup.get(t["test_name"])
        db.upsert_tab4_tests(conn, parsed["tab4_tests"])
        cpm_counts = db.upsert_cpm_rows(conn, parsed["cpm_rows"])
        db.delete_cpm_not_in(
            conn, {(r["country"], r["method_name"]) for r in parsed["cpm_rows"]})
    except Exception as exc:
        result["error"] = f"Database write failed: {exc}"
        return result
    finally:
        conn.close()

    result.update({
        "ok": True,
        "per_tab": parsed["per_tab"],
        "inserted": counts["inserted"],
        "updated": counts["updated"],
        "resolved": res["resolved"],
        "unresolved": [
            {"area": r["area"], "excel_row": r["excel_row"], "name": r["name"],
             "excel_test_ref": r["excel_test_ref"], "test_name": r["test_name"]}
            for r in res["unresolved"]],
        "skipped_dups": parsed["skipped_dups"],
        "countries_seeded": seeded,
        "removed_stale": removed,
        "cpm_inserted": cpm_counts["inserted"],
        "cpm_updated": cpm_counts["updated"],
        "cpm_unknown_category": sum(1 for r in parsed["cpm_rows"] if r["category"] is None),
        "tab4_tests": parsed["tab4_tests"],
        "named_targets": target_rows,
    })
    return result
